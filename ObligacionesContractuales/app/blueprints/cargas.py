"""
Blueprint de cargas masivas.

Responsabilidades:
- Carga masiva de evidencias desde Excel.
- Carga masiva de evidencias por mes.
- Procesamiento en segundo plano.
- Progreso mediante Server-Sent Events (SSE).
- Rate limiter para Gemini.
- Generación de plantilla Excel por reporte.
- Generación de plantilla Excel para carga masiva mensual.
"""

import os
import io
import calendar
import threading
import time
import uuid
import json

from datetime import datetime, date
from app.services.excel_service import ExcelService
from app.services.carga_masiva_service import CargaMasivaService
from app.services.reporte_service import ReporteService
from app.services.evidencia_service import EvidenciaService

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    Response,
    jsonify,
    current_app
)

from flask_login import (
    login_required,
    current_user
)

from werkzeug.utils import secure_filename

from openpyxl import load_workbook

from models import (
    db,
    Contrato,
    Obligacion,
    ReporteMensual,
    Evidencia
)

from app.blueprints.configuracion import (
    _obtener_api_key
)

from app.services.excel_service import (
    ExcelService
)

from app.services.carga_masiva_service import (
    CargaMasivaService
)

from app.services.reporte_service import (
    ReporteService
)

from app.services.evidencia_service import (
    EvidenciaService
)


# ============================================================
# BLUEPRINT
# ============================================================

cargas_bp = Blueprint(
    'cargas',
    __name__
)


# ============================================================
# CONFIGURACIÓN
# ============================================================

ALLOWED_EXTENSIONS = {
    'png',
    'jpg',
    'jpeg',
    'gif',
    'bmp',
    'webp'
}


def allowed_file(filename):
    """
    Verifica si un archivo tiene una extensión de imagen
    permitida.
    """

    return (
        '.'
        in filename
        and
        filename.rsplit(
            '.',
            1
        )[1].lower()
        in ALLOWED_EXTENSIONS
    )


# ============================================================
# GENERAR MESES DEL CONTRATO
# ============================================================

def generar_meses_contrato(
    fecha_inicio,
    fecha_fin
):
    """
    Genera los meses comprendidos entre las fechas
    de inicio y fin del contrato.
    """

    meses = []

    current = date(
        fecha_inicio.year,
        fecha_inicio.month,
        1
    )

    end = date(
        fecha_fin.year,
        fecha_fin.month,
        1
    )

    nombres_meses = [
        '',
        'Enero',
        'Febrero',
        'Marzo',
        'Abril',
        'Mayo',
        'Junio',
        'Julio',
        'Agosto',
        'Septiembre',
        'Octubre',
        'Noviembre',
        'Diciembre'
    ]

    while current <= end:

        meses.append(
            (
                current.month,
                current.year,
                nombres_meses[
                    current.month
                ]
            )
        )

        if current.month == 12:

            current = date(
                current.year + 1,
                1,
                1
            )

        else:

            current = date(
                current.year,
                current.month + 1,
                1
            )

    return meses


# ============================================================
# ESTADO GLOBAL DE JOBS
# ============================================================

jobs_lock = threading.Lock()

jobs_progreso = {}


# ============================================================
# RATE LIMITER GEMINI
# ============================================================

_gemini_last_call = 0.0

_gemini_lock = threading.Lock()

# Aproximadamente 15 solicitudes por minuto.
# Se deja margen de seguridad.
GEMINI_MIN_INTERVAL = 4.1


def _esperar_rate_limit_gemini():
    """
    Espera el tiempo necesario entre llamadas a Gemini.
    """

    global _gemini_last_call

    with _gemini_lock:

        ahora = time.time()

        transcurrido = (
            ahora
            - _gemini_last_call
        )

        if (
            transcurrido
            < GEMINI_MIN_INTERVAL
        ):

            esperar = (
                GEMINI_MIN_INTERVAL
                - transcurrido
            )

            time.sleep(
                esperar
            )

        _gemini_last_call = time.time()


# ============================================================
# ACTUALIZAR JOB
# ============================================================

def _actualizar_job(
    job_id,
    estado,
    porcentaje,
    mensaje,
    resultado=None,
    errores=None
):
    """
    Actualiza de manera thread-safe el estado de un
    proceso de carga masiva.
    """

    with jobs_lock:

        if job_id not in jobs_progreso:

            jobs_progreso[
                job_id
            ] = {}

        jobs_progreso[
            job_id
        ].update(
            {
                'estado': estado,
                'porcentaje': porcentaje,
                'mensaje': mensaje,
                'timestamp': time.time()
            }
        )

        if resultado is not None:

            jobs_progreso[
                job_id
            ]['resultado'] = resultado

        if errores is not None:

            jobs_progreso[
                job_id
            ]['errores'] = errores


# ============================================================
# PROCESAMIENTO DE CARGA MASIVA
# ============================================================

def _procesar_carga_masiva_job(
    app,
    job_id,
    contrato_id,
    mes,
    anio,
    excel_path,
    imagenes_subidas,
    api_key
):
    """
    Procesa la carga masiva mensual en segundo plano.

    El blueprint solamente coordina el proceso.
    La lógica de negocio está delegada a:

        ExcelService
        CargaMasivaService
        ReporteService
        EvidenciaService
        GeminiService
    """

    try:

        with app.app_context():

            # ====================================================
            # RECUPERAR CONTRATO DENTRO DEL CONTEXTO FLASK
            # ====================================================

            contrato = (
                Contrato.query
                .filter_by(
                    id=contrato_id
                )
                .first()
            )

            if not contrato:

                _actualizar_job(
                    job_id,
                    'error',
                    0,
                    f'No se encontró el contrato {contrato_id}.'
                )

                return

            # ====================================================
            # RECUPERAR OBLIGACIONES DENTRO DEL CONTEXTO FLASK
            # ====================================================

            obligaciones = (
                Obligacion.query
                .filter_by(
                    contrato_id=contrato.id
                )
                .order_by(
                    Obligacion.numero
                )
                .all()
            )

            # ====================================================
            # SERVICIOS
            # ====================================================

            reporte_service = ReporteService()

            evidencia_service = EvidenciaService()

            carga_service = CargaMasivaService(
                reporte_service=reporte_service,
                evidencia_service=evidencia_service
            )


            # ====================================================
            # LEER EXCEL
            # ====================================================

            try:

                workbook = load_workbook(
                    excel_path,
                    data_only=True
                )

                worksheet = workbook.active

            except Exception as exc:

                _actualizar_job(
                    job_id,
                    'error',
                    0,
                    (
                        'No fue posible abrir el archivo Excel: '
                        f'{str(exc)}'
                    )
                )

                return

            # ====================================================
            # ENCABEZADOS
            # ====================================================

            expected_headers = [
                'Obligacion No.',
                'Descripcion Obligacion',
                'Anuncio / Contexto',
                'Fecha de la actividad',
                'Nombre Imagen'
            ]

            headers = [
                cell.value
                for cell in worksheet[1]
            ]

            headers = [
                str(value).strip()
                if value is not None
                else ''
                for value in headers[:5]
            ]

            if headers != expected_headers:

                _actualizar_job(
                    job_id,
                    'error',
                    0,
                    (
                        'Encabezados incorrectos. '
                        f'Esperado: {expected_headers}. '
                        f'Recibido: {headers}.'
                    )
                )

                return

            # ====================================================
            # OBLIGACIONES POR NÚMERO
            # ====================================================

            obligaciones_por_numero = {}

            for obligacion in obligaciones:

                try:

                    obligaciones_por_numero[
                        obligacion.numero
                    ] = obligacion

                except Exception:

                    continue

            # ====================================================
            # PREPARAR FILAS
            # ====================================================

            filas = []

            for numero_fila, row in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    values_only=True
                ),
                start=2
            ):

                valores = list(row)

                while len(valores) < 5:

                    valores.append(None)

                obligacion_value = valores[0]
                descripcion_value = valores[1]
                anuncio_value = valores[2]
                fecha_value = valores[3]
                imagen_value = valores[4]

                if (
                    obligacion_value is None
                    and
                    anuncio_value is None
                    and
                    imagen_value is None
                ):

                    continue

                filas.append(
                    {
                        'obligacion': obligacion_value,
                        'descripcion': descripcion_value,
                        'anuncio': anuncio_value,
                        'fecha': fecha_value,
                        'nombre_imagen': imagen_value,
                        '_fila_excel': numero_fila
                    }
                )

            total_filas = len(filas)

            if total_filas == 0:

                _actualizar_job(
                    job_id,
                    'error',
                    0,
                    'No se encontraron filas válidas en el Excel.'
                )

                return

            # ====================================================
            # CACHE
            # ====================================================

            reportes_cache = {}

            # ====================================================
            # IMÁGENES
            # ====================================================

            imagenes_disponibles = dict(
                imagenes_subidas or {}
            )

            # ====================================================
            # RESULTADOS
            # ====================================================

            exitosos = 0

            errores = []

            # ====================================================
            # PROCESAR FILAS
            # ====================================================

            for indice, fila in enumerate(
                filas,
                start=1
            ):

                numero_fila = fila.get(
                    '_fila_excel',
                    indice
                )

                porcentaje = int(
                    (
                        (indice - 1)
                        /
                        total_filas
                    )
                    * 100
                )

                _actualizar_job(
                    job_id,
                    'procesando',
                    porcentaje,
                    (
                        f'Procesando fila '
                        f'{numero_fila} '
                        f'({indice}/{total_filas})...'
                    )
                )

                # ------------------------------------------------
                # Rate limit para Gemini (solo si hay imagen)
                # ------------------------------------------------
                if fila.get('nombre_imagen'):
                    _esperar_rate_limit_gemini()

                # ------------------------------------------------
                # PROCESAR FILA MEDIANTE EL SERVICIO
                # ------------------------------------------------

                try:

                    resultado = (
                        carga_service
                        ._procesar_fila(
                            contrato=contrato,
                            mes=mes,
                            anio=anio,
                            fila=fila,
                            imagenes=imagenes_disponibles,
                            obligaciones_por_numero=(
                                obligaciones_por_numero
                            ),
                            api_key=api_key,
                            reportes_cache=reportes_cache
                        )
                    )

                except Exception as exc:
                    resultado = {
                        'exitoso': False,
                        'errores': [
                            (
                                f'Fila {numero_fila}: '
                                f'Error inesperado: '
                                f'{str(exc)}'
                            )
                        ],
                        'evidencia': None
                    }

                # ------------------------------------------------
                # ERRORES
                # ------------------------------------------------

                fila_errores = resultado.get(
                    'errores',
                    []
                )

                if fila_errores:

                    errores.extend(
                        fila_errores
                    )

                # ------------------------------------------------
                # ÉXITO
                # ------------------------------------------------

                if resultado.get(
                    'exitoso',
                    False
                ):

                    exitosos += 1

                # ------------------------------------------------
                # PROGRESO
                # ------------------------------------------------

                porcentaje_actual = int(
                    (
                        indice
                        /
                        total_filas
                    )
                    * 100
                )

                _actualizar_job(
                    job_id,
                    'procesando',
                    porcentaje_actual,
                    (
                        f'Fila {numero_fila} procesada '
                        f'({indice}/{total_filas}).'
                    )
                )
            # ====================================================
            # CONFIRMAR TRANSACCIÓN
            # ====================================================

            try:

                db.session.commit()

            except Exception as exc:

                db.session.rollback()

                raise RuntimeError(
                    'No fue posible guardar la carga masiva '
                    f'en la base de datos: {str(exc)}'
                )

            # ====================================================
            # CERRAR EXCEL
            # ====================================================

            try:

                workbook.close()

            except Exception:

                pass

            # ====================================================
            # LIMPIAR IMÁGENES QUE NO FUERON UTILIZADAS
            # ====================================================

            for tmp_path in (
                imagenes_disponibles.values()
            ):

                try:

                    if (
                        tmp_path
                        and
                        os.path.exists(
                            tmp_path
                        )
                    ):

                        os.remove(
                            tmp_path
                        )

                except Exception:

                    pass

            # ====================================================
            # LIMPIAR EXCEL TEMPORAL
            # ====================================================

            try:

                if (
                    excel_path
                    and
                    os.path.exists(
                        excel_path
                    )
                ):

                    os.remove(
                        excel_path
                    )

            except Exception:

                pass

            # ====================================================
            # RESULTADO FINAL
            # ====================================================

            _actualizar_job(
                job_id,
                'completado',
                100,
                (
                    f'Proceso finalizado. '
                    f'{exitosos} evidencias cargadas.'
                ),
                resultado={
                    'exitosos': exitosos,
                    'errores': len(errores),
                    'total': total_filas,
                    'mes': mes,
                    'anio': anio
                },
                errores=errores
            )

    except Exception as exc:

        import traceback

        traceback.print_exc()

        # ========================================================
        # EL EXCEPT TAMBIÉN NECESITA CONTEXTO FLASK
        # ========================================================

        try:

            with app.app_context():

                _actualizar_job(
                    job_id,
                    'error',
                    0,
                    (
                        'Error inesperado durante la carga masiva: '
                        f'{str(exc)}'
                    )
                )

        except Exception as error_job:

            print(
                '[CargaMasiva] '
                'No fue posible actualizar el estado del job: '
                f'{error_job}'
            )

        # ========================================================
        # LIMPIAR EXCEL TEMPORAL
        # ========================================================

        try:

            if (
                excel_path
                and
                os.path.exists(
                    excel_path
                )
            ):

                os.remove(
                    excel_path
                )

        except Exception:

            pass

# ============================================================
# CARGA MASIVA PARA UN REPORTE
# ============================================================

@cargas_bp.route(
    '/reporte/<int:id>/carga-masiva',
    methods=['GET', 'POST']
)
@login_required
def carga_masiva_evidencias(id):
    """
    Carga masiva de actividades mediante Excel
    para un reporte específico.
    """

    reporte = (
        ReporteMensual.query
        .get_or_404(id)
    )

    obligacion = (
        reporte.obligacion
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    if (
        not contrato
        or
        contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # Contrato cerrado
    # --------------------------------------------------------

    if contrato.etapa == 'Reporte Cerrado':

        flash(
            'Este contrato esta finalizado '
            '(Reporte Cerrado). '
            'No se pueden agregar mas evidencias.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # ========================================================
    # POST
    # ========================================================

    if request.method == 'POST':

        # ----------------------------------------------------
        # Excel
        # ----------------------------------------------------

        if 'archivo_excel' not in request.files:

            flash(
                'No se selecciono el archivo Excel.',
                'danger'
            )

            return redirect(
                url_for(
                    'cargas.carga_masiva_evidencias',
                    id=id
                )
            )

        archivo_excel = request.files[
            'archivo_excel'
        ]

        if archivo_excel.filename == '':

            flash(
                'No se selecciono archivo Excel.',
                'danger'
            )

            return redirect(
                url_for(
                    'cargas.carga_masiva_evidencias',
                    id=id
                )
            )

        # ----------------------------------------------------
        # Validar extensión
        # ----------------------------------------------------

        if not archivo_excel.filename.lower().endswith(
            (
                '.xlsx',
                '.xls'
            )
        ):

            flash(
                'El archivo debe ser Excel '
                '(.xlsx o .xls).',
                'danger'
            )

            return redirect(
                url_for(
                    'cargas.carga_masiva_evidencias',
                    id=id
                )
            )

        # ====================================================
        # LEER EXCEL
        # ====================================================

        try:

            wb = load_workbook(
                archivo_excel
            )

            ws = wb.active

            headers = [
                cell.value
                for cell in ws[1]
            ]

            expected = [
                'Anuncio / Contexto',
                'Fecha de la actividad'
            ]

            if headers[:2] != expected:

                flash(
                    (
                        'Encabezados incorrectos. '
                        f'Se esperaba: {expected}. '
                        f'Encontrado: {headers[:2]}'
                    ),
                    'danger'
                )

                return redirect(
                    url_for(
                        'cargas.carga_masiva_evidencias',
                        id=id
                    )
                )

            api_key = _obtener_api_key()

            exitosos = 0

            errores = []

            # =================================================
            # FILAS
            # =================================================

            for idx, row in enumerate(
                ws.iter_rows(
                    min_row=2,
                    values_only=True
                ),
                start=2
            ):

                anuncio = str(
                    row[0] or ''
                ).strip()

                fecha_str = str(
                    row[1] or ''
                ).strip()

                if not anuncio:

                    errores.append(
                        f'Fila {idx}: '
                        f'Anuncio vacio.'
                    )

                    continue

                if not fecha_str:

                    errores.append(
                        f'Fila {idx}: '
                        f'Fecha vacia.'
                    )

                    continue

                # ------------------------------------------------
                # Fecha
                # ------------------------------------------------

                try:

                    fecha_actividad = (
                        datetime.strptime(
                            fecha_str,
                            '%Y-%m-%d'
                        ).date()
                    )

                except ValueError:

                    try:

                        fecha_actividad = (
                            datetime.strptime(
                                fecha_str,
                                '%d/%m/%Y'
                            ).date()
                        )

                    except ValueError:

                        errores.append(
                            f'Fila {idx}: '
                            f'Fecha invalida '
                            f'({fecha_str}). '
                            f'Use YYYY-MM-DD '
                            f'o DD/MM/YYYY.'
                        )

                        continue

                # ------------------------------------------------
                # Validar período
                # ------------------------------------------------

                if (
                    fecha_actividad
                    < reporte.fecha_inicio_reporte
                    or
                    fecha_actividad
                    > reporte.fecha_fin_reporte
                ):

                    errores.append(
                        f'Fila {idx}: '
                        f'Fecha {fecha_str} '
                        f'fuera del periodo del reporte.'
                    )

                    continue

                # ------------------------------------------------
                # Número actividad
                # ------------------------------------------------

                ultima_evidencia = (
                    Evidencia.query
                    .filter_by(
                        reporte_id=reporte.id
                    )
                    .order_by(
                        Evidencia
                        .numero_actividad
                        .desc()
                    )
                    .first()
                )

                numero_actividad = (
                    ultima_evidencia
                    .numero_actividad + 1
                    if ultima_evidencia
                    else 1
                )

                # ------------------------------------------------
                # Crear evidencia
                # ------------------------------------------------

                evidencia = Evidencia(
                    numero_actividad=(
                        numero_actividad
                    ),

                    imagen_path='',

                    anuncio_usuario=(
                        anuncio
                    ),

                    descripcion_visual_ia=None,

                    descripcion_actividad='',

                    fecha_actividad=(
                        fecha_actividad
                    ),

                    reporte_id=(
                        reporte.id
                    )
                )

                # ------------------------------------------------
                # Generar descripción
                # ------------------------------------------------

                evidencia.descripcion_actividad = (
                    evidencia
                    .generar_descripcion_automatica(
                        obligacion
                    )
                )

                db.session.add(
                    evidencia
                )

                exitosos += 1

            db.session.commit()

            # ----------------------------------------------------
            # Resultado
            # ----------------------------------------------------

            if exitosos:

                flash(
                    f'{exitosos} actividades '
                    f'cargadas exitosamente.',
                    'success'
                )

            if errores:

                flash(
                    'Se encontraron errores: '
                    +
                    ' | '.join(
                        errores[:5]
                    ),
                    'warning'
                )

            return redirect(
                url_for(
                    'reportes.ver_reporte',
                    id=id
                )
            )

        except Exception as e:

            flash(
                f'Error procesando Excel: {str(e)}',
                'danger'
            )

            return redirect(
                url_for(
                    'cargas.carga_masiva_evidencias',
                    id=id
                )
            )

    # ========================================================
    # GET
    # ========================================================

    return render_template(
        'carga_masiva.html',

        reporte=reporte,

        obligacion=obligacion,

        contrato=contrato
    )


# ============================================================
# GENERAR PLANTILLA MASIVA POR MES
# ============================================================

def generar_plantilla_masiva(
    contrato,
    obligaciones,
    mes,
    anio
):
    """
    Genera una plantilla Excel para cargar evidencias
    de todas las obligaciones de un mes.
    """

    from openpyxl import Workbook

    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side
    )

    # --------------------------------------------------------
    # Crear workbook
    # --------------------------------------------------------

    wb = Workbook()

    ws = wb.active

    ws.title = (
        f'Carga_{mes:02d}_{anio}'
    )

    # --------------------------------------------------------
    # Encabezados
    # --------------------------------------------------------

    headers = [
        'Obligacion No.',
        'Descripcion Obligacion',
        'Anuncio / Contexto',
        'Fecha de la actividad',
        'Nombre Imagen'
    ]

    ws.append(
        headers
    )

    # --------------------------------------------------------
    # Estilos
    # --------------------------------------------------------

    header_font = Font(
        bold=True,
        color='FFFFFF',
        size=11
    )

    header_fill = PatternFill(
        start_color='2c3e50',
        end_color='2c3e50',
        fill_type='solid'
    )

    header_align = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )

    thin_border = Border(
        left=Side(
            style='thin'
        ),
        right=Side(
            style='thin'
        ),
        top=Side(
            style='thin'
        ),
        bottom=Side(
            style='thin'
        )
    )

    for cell in ws[1]:

        cell.font = (
            header_font
        )

        cell.fill = (
            header_fill
        )

        cell.alignment = (
            header_align
        )

        cell.border = (
            thin_border
        )

    # --------------------------------------------------------
    # Obligaciones
    # --------------------------------------------------------

    for obligacion in obligaciones:

        ws.append(
            [
                obligacion.numero,

                obligacion.descripcion,

                '',

                f'{anio}-{mes:02d}-15',

                ''
            ]
        )

    # --------------------------------------------------------
    # Anchos
    # --------------------------------------------------------

    ws.column_dimensions[
        'A'
    ].width = 16

    ws.column_dimensions[
        'B'
    ].width = 55

    ws.column_dimensions[
        'C'
    ].width = 55

    ws.column_dimensions[
        'D'
    ].width = 22

    ws.column_dimensions[
        'E'
    ].width = 28

    ws.freeze_panes = 'A2'

    # ========================================================
    # HOJA DE INSTRUCCIONES
    # ========================================================

    ws_instr = wb.create_sheet(
        'Instrucciones'
    )

    instrucciones = [

        [
            'INSTRUCCIONES DE CARGA MASIVA POR MES'
        ],

        [''],

        [
            '1. NO modifique los encabezados '
            'de columna (fila 1).'
        ],

        [
            '2. NO modifique las columnas A y B '
            '(Obligacion No. y Descripcion).'
        ],

        [
            '3. En la columna C escriba el anuncio '
            'o contexto de la actividad '
            '(solo para el sistema).'
        ],

        [
            '4. En la columna D indique la fecha '
            'en formato YYYY-MM-DD o DD/MM/YYYY.'
        ],

        [
            '5. En la columna E escriba el nombre '
            'EXACTO del archivo de imagen, '
            'incluyendo extension '
            '(ej: evidencia1.jpg).'
        ],

        [
            '6. Puede INSERTAR mas filas para la '
            'misma obligacion si tiene multiples evidencias.'
        ],

        [
            '7. Puede ELIMINAR las filas de obligaciones '
            'que no tengan evidencias este mes.'
        ],

        [
            '8. Las imagenes deben cargarse JUNTO '
            'con el Excel en el formulario web '
            '(campo de archivos multiples).'
        ],

        [''],

        [
            'REGLAS IMPORTANTES:'
        ],

        [
            '- La fecha debe pertenecer al mes '
            'y año seleccionados.'
        ],

        [
            '- El nombre de imagen en el Excel '
            'debe coincidir EXACTAMENTE con el archivo subido.'
        ],

        [
            '- Si no adjunta imagen, deje la columna E vacia; '
            'se creara la actividad sin evidencia visual.'
        ],

        [
            '- El sistema creara automaticamente '
            'los reportes mensuales por obligacion '
            'si no existen.'
        ],

        [
            '- Si tiene API key de Gemini configurada, '
            'analizara automaticamente cada imagen.'
        ],

        [
            '- NOTA: El tier gratuito de Gemini '
            'permite 15 imagenes/minuto. '
            'Si sube mas, el sistema las procesara '
            'automaticamente con pausas.'
        ]
    ]

    for row in instrucciones:

        ws_instr.append(
            row
        )

    ws_instr.column_dimensions[
        'A'
    ].width = 100

    # --------------------------------------------------------
    # Memoria
    # --------------------------------------------------------

    output = io.BytesIO()

    wb.save(
        output
    )

    output.seek(0)

    # --------------------------------------------------------
    # Nombre
    # --------------------------------------------------------

    filename = (
        f'Plantilla_CargaMasiva_'
        f'{mes:02d}_{anio}_'
        f'{contrato.contratista or "Contrato"}.xlsx'
    )

    return send_file(
        output,

        mimetype=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),

        as_attachment=True,

        download_name=filename
    )


# ============================================================
# CARGA MASIVA POR MES
# ============================================================

@cargas_bp.route(
    '/carga-masiva-mes',
    methods=['GET', 'POST']
)
@login_required
def carga_masiva_mes():
    """
    Carga masiva de evidencias para TODAS las obligaciones
    del contrato activo en un mes determinado.
    """

    # --------------------------------------------------------
    # Contrato activo
    # --------------------------------------------------------

    contrato = (
        Contrato.query
        .filter_by(
            activo=True,
            user_id=current_user.id
        )
        .first()
    )

    if not contrato:

        flash(
            'No hay contrato activo configurado.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # Contrato cerrado
    # --------------------------------------------------------

    if contrato.etapa == 'Reporte Cerrado':

        flash(
            'El contrato activo esta finalizado '
            '(Reporte Cerrado). '
            'No se pueden agregar mas evidencias.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.reportes'
            )
        )

    # --------------------------------------------------------
    # Obligaciones
    # --------------------------------------------------------

    obligaciones = (
        Obligacion.query
        .filter_by(
            contrato_id=contrato.id
        )
        .order_by(
            Obligacion.numero
        )
        .all()
    )

    # --------------------------------------------------------
    # Meses
    # --------------------------------------------------------

    meses = generar_meses_contrato(
        contrato.fecha_inicio,
        contrato.fecha_fin
    )

    # ========================================================
    # POST
    # ========================================================

    if request.method == 'POST':

        action = request.form.get(
            'action'
        )

        # ====================================================
        # DESCARGAR PLANTILLA
        # ====================================================

        if action == 'descargar_plantilla':

            try:

                mes = int(
                    request.form.get(
                        'mes',
                        0
                    )
                )

                anio = int(
                    request.form.get(
                        'anio',
                        0
                    )
                )

            except (
                TypeError,
                ValueError
            ):

                flash(
                    'Seleccione mes y año.',
                    'danger'
                )

                return redirect(
                    url_for(
                        'cargas.carga_masiva_mes'
                    )
                )

            if not mes or not anio:

                flash(
                    'Seleccione mes y año.',
                    'danger'
                )

                return redirect(
                    url_for(
                        'cargas.carga_masiva_mes'
                    )
                )

            return generar_plantilla_masiva(
                contrato,
                obligaciones,
                mes,
                anio
            )

        # ====================================================
        # CARGAR MASIVO
        # ====================================================

        elif action == 'cargar_masivo':

            try:

                mes = int(
                    request.form.get(
                        'mes',
                        0
                    )
                )

                anio = int(
                    request.form.get(
                        'anio',
                        0
                    )
                )

            except (
                TypeError,
                ValueError
            ):

                return jsonify(
                    {
                        'error':
                        'Seleccione mes y año.'
                    }
                ), 400

            if not mes or not anio:

                return jsonify(
                    {
                        'error':
                        'Seleccione mes y año.'
                    }
                ), 400

            # ------------------------------------------------
            # Validar mes
            # ------------------------------------------------

            if mes < 1 or mes > 12:

                return jsonify(
                    {
                        'error':
                        'El mes seleccionado no es válido.'
                    }
                ), 400

            # ------------------------------------------------
            # Excel
            # ------------------------------------------------

            if (
                'archivo_excel'
                not in request.files
            ):

                return jsonify(
                    {
                        'error':
                        'No se seleccionó '
                        'el archivo Excel.'
                    }
                ), 400

            archivo_excel = request.files[
                'archivo_excel'
            ]

            if archivo_excel.filename == '':

                return jsonify(
                    {
                        'error':
                        'No se seleccionó '
                        'archivo Excel.'
                    }
                ), 400

            # ------------------------------------------------
            # Extensión Excel
            # ------------------------------------------------

            if not archivo_excel.filename.lower().endswith(
                (
                    '.xlsx',
                    '.xls'
                )
            ):

                return jsonify(
                    {
                        'error':
                        'El archivo debe ser Excel '
                        '(.xlsx o .xls).'
                    }
                ), 400

            # =================================================
            # CREAR JOB
            # =================================================

            job_id = str(
                uuid.uuid4()
            )

            tmp_ts = (
                datetime.now()
                .strftime(
                    '%Y%m%d_%H%M%S'
                )
            )

            excel_tmp_name = secure_filename(
                (
                    f'tmp_excel_'
                    f'{job_id}_'
                    f'{tmp_ts}.xlsx'
                )
            )

            excel_tmp_path = os.path.join(
                current_app.config[
                    'UPLOAD_FOLDER'
                ],
                excel_tmp_name
            )

            archivo_excel.save(
                excel_tmp_path
            )

            # =================================================
            # IMÁGENES TEMPORALES
            # =================================================

            imagenes_subidas = {}

            imagenes_files = request.files.getlist(
                'imagenes'
            )

            for img_file in imagenes_files:

                if (
                    img_file
                    and
                    img_file.filename
                    and
                    allowed_file(
                        img_file.filename
                    )
                ):

                    tmp_name = secure_filename(
                        (
                            f'tmp_'
                            f'{job_id}_'
                            f'{img_file.filename}'
                        )
                    )

                    tmp_path = os.path.join(
                        current_app.config[
                            'UPLOAD_FOLDER'
                        ],
                        tmp_name
                    )

                    img_file.save(
                        tmp_path
                    )

                    imagenes_subidas[
                        img_file.filename
                    ] = tmp_path

                    imagenes_subidas[
                        secure_filename(
                            img_file.filename
                        )
                    ] = tmp_path

            # =================================================
            # API KEY
            # =================================================

            api_key = _obtener_api_key()

            # =================================================
            # INICIALIZAR JOB
            # =================================================

            _actualizar_job(
                job_id,
                'iniciado',
                0,
                'Iniciando procesamiento...'
            )

            # =================================================
            # THREAD
            # =================================================
            app = current_app._get_current_object()
            thread = threading.Thread(
                target=_procesar_carga_masiva_job,
                args=(
                    current_app._get_current_object(),
                    job_id,
                    contrato.id,
                    mes,
                    anio,
                    excel_tmp_path,
                    imagenes_subidas,
                    api_key
                )
            )

            thread.daemon = True

            thread.start()

            return jsonify(
                {
                    'job_id': job_id,
                    'status': 'started'
                }
            )

    # ========================================================
    # GET
    # ========================================================

    api_key_configurada = bool(
        _obtener_api_key()
    )

    return render_template(
        'carga_masiva_mes.html',

        contrato=contrato,

        obligaciones=obligaciones,

        meses=meses,

        api_key_configurada=(
            api_key_configurada
        )
    )


# ============================================================
# PROGRESO SSE
# ============================================================

@cargas_bp.route(
    '/carga-masiva-mes/progreso/<job_id>'
)
@login_required
def carga_masiva_progreso(job_id):
    """
    Server-Sent Events.

    Envía al navegador el progreso de la carga masiva
    en tiempo real.
    """

    def event_stream():

        ultimo_estado = None

        while True:

            # ------------------------------------------------
            # Leer estado
            # ------------------------------------------------

            with jobs_lock:

                job = jobs_progreso.get(
                    job_id,
                    {}
                )

            estado = job.get(
                'estado',
                'desconocido'
            )

            porcentaje = job.get(
                'porcentaje',
                0
            )

            mensaje = job.get(
                'mensaje',
                'Procesando...'
            )

            errores = job.get(
                'errores',
                []
            )

            resultado = job.get(
                'resultado'
            )

            # ------------------------------------------------
            # Datos SSE
            # ------------------------------------------------

            data = {
                'estado': estado,

                'porcentaje': porcentaje,

                'mensaje': mensaje,

                'errores': errores[:5]
            }

            if resultado:

                data[
                    'resultado'
                ] = resultado

            # ------------------------------------------------
            # Enviar
            # ------------------------------------------------

            yield (
                'data: '
                +
                json.dumps(
                    data
                )
                +
                '\n\n'
            )

            ultimo_estado = (
                estado
            )

            # ------------------------------------------------
            # Finalizado
            # ------------------------------------------------

            if estado in (
                'completado',
                'error'
            ):

                time.sleep(
                    2
                )

                with jobs_lock:

                    jobs_progreso.pop(
                        job_id,
                        None
                    )

                break

            # ------------------------------------------------
            # Esperar
            # ------------------------------------------------

            time.sleep(
                0.5
            )

    return Response(
        event_stream(),
        mimetype='text/event-stream'
    )


# ============================================================
# PLANTILLA EXCEL PARA UN REPORTE
# ============================================================

@cargas_bp.route(
    '/reporte/<int:id>/plantilla-excel'
)
@login_required
def descargar_plantilla_excel(id):
    """
    Descarga una plantilla Excel simple para registrar
    actividades de un reporte específico.
    """

    from openpyxl import Workbook

    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment
    )

    # --------------------------------------------------------
    # Validar reporte
    # --------------------------------------------------------

    reporte = (
        ReporteMensual.query
        .get_or_404(id)
    )

    obligacion = (
        reporte.obligacion
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    if (
        not contrato
        or
        contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # ========================================================
    # CREAR WORKBOOK
    # ========================================================

    wb = Workbook()

    ws = wb.active

    ws.title = (
        'Carga Masiva'
    )

    # --------------------------------------------------------
    # Encabezados
    # --------------------------------------------------------

    headers = [
        'Anuncio / Contexto',
        'Fecha de la actividad'
    ]

    ws.append(
        headers
    )

    # --------------------------------------------------------
    # Estilos
    # --------------------------------------------------------

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color='FFFFFF'
        )

        cell.fill = PatternFill(
            start_color='2c3e50',
            end_color='2c3e50',
            fill_type='solid'
        )

        cell.alignment = Alignment(
            horizontal='center'
        )

    # --------------------------------------------------------
    # Ejemplos
    # --------------------------------------------------------

    ws.append(
        [
            'Presentacion del estado de avance '
            'de proyectos',

            '2026-07-15'
        ]
    )

    ws.append(
        [
            'Revision de solicitudes de ajuste '
            'tecnicos',

            '2026-07-20'
        ]
    )

    ws.append(
        [
            'Elaboracion del plan de trabajo '
            'mensual',

            '2026-07-25'
        ]
    )

    # --------------------------------------------------------
    # Anchos
    # --------------------------------------------------------

    ws.column_dimensions[
        'A'
    ].width = 60

    ws.column_dimensions[
        'B'
    ].width = 25

    # --------------------------------------------------------
    # Memoria
    # --------------------------------------------------------

    output = io.BytesIO()

    wb.save(
        output
    )

    output.seek(0)

    # --------------------------------------------------------
    # Nombre
    # --------------------------------------------------------

    filename = (
        f'Plantilla_Carga_Masiva_{id}.xlsx'
    )

    return send_file(
        output,

        mimetype=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),

        as_attachment=True,

        download_name=filename
    )
