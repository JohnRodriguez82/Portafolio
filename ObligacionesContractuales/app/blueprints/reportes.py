"""
Blueprint de reportes y evidencias.

Responsabilidades:
- Listado de reportes.
- Creación de reportes mensuales.
- Visualización de reportes.
- Registro de evidencias.
- Edición de evidencias.
- Eliminación de evidencias.
- Generación de PDF individual.
- Eliminación de reportes.
- Servir archivos de evidencias.
- Descarga masiva de PDFs por mes.
- Generación de Excel consolidado.

IMPORTANTE:
Este Blueprint conserva la lógica de negocio existente
en app.py.

La refactorización cambia la estructura de la aplicación,
pero no pretende cambiar el comportamiento funcional.
"""

import logging
import os
import io
import zipfile
import calendar
import threading

from datetime import datetime, date

from datetime import (
    datetime,
    date
)

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_from_directory,
    session,
    send_file,
    jsonify,
    current_app
)

from werkzeug.exceptions import RequestEntityTooLarge

from flask_login import (
    login_required,
    current_user
)

from werkzeug.utils import (
    secure_filename
)

from werkzeug.exceptions import (
    RequestEntityTooLarge
)

from models import (
    db,
    Contrato,
    Obligacion,
    ReporteMensual,
    Evidencia
)

from pdf_generator import (
    PDFGenerator
)

from vision_analyzer import (
    analizar_imagen,
    analizar_imagen_con_reintentos,
    consolidar_textos_ejecutivo
)

from app.blueprints.configuracion import (
    _obtener_api_key
)
from app.services.evidencia_service import (
    EvidenciaService
)

# ============================================================
# BLUEPRINT
# ============================================================

reportes_bp = Blueprint(
    'reportes',
    __name__
)


# ============================================================
# CONFIGURACIÓN DE ARCHIVOS
# ============================================================

ALLOWED_EXTENSIONS = {
    'png',
    'jpg',
    'jpeg',
    'gif',
    'bmp',
    'webp'
}


def allowed_file(filename):
    """
    Determina si la extensión del archivo de imagen
    está permitida.
    """

    return (
        '.'
        in filename
        and filename.rsplit(
            '.',
            1
        )[1].lower()
        in ALLOWED_EXTENSIONS
    )


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def generar_meses_contrato(
    fecha_inicio,
    fecha_fin
):
    """
    Genera todos los meses comprendidos entre la fecha
    inicial y final del contrato.

    Retorna:

        [
            (mes, anio, nombre_mes),
            ...
        ]

    Ejemplo:

        [
            (1, 2026, 'Enero'),
            (2, 2026, 'Febrero'),
            (3, 2026, 'Marzo')
        ]
    """

    meses = []

    current = date(
        fecha_inicio.year,
        fecha_inicio.month,
        1
    )

    end = date(
        fecha_fin.year,
        fecha_fin.month,
        1
    )

    nombres_meses = [
        '',
        'Enero',
        'Febrero',
        'Marzo',
        'Abril',
        'Mayo',
        'Junio',
        'Julio',
        'Agosto',
        'Septiembre',
        'Octubre',
        'Noviembre',
        'Diciembre'
    ]

    while current <= end:

        meses.append(
            (
                current.month,
                current.year,
                nombres_meses[
                    current.month
                ]
            )
        )

        if current.month == 12:

            current = date(
                current.year + 1,
                1,
                1
            )

        else:

            current = date(
                current.year,
                current.month + 1,
                1
            )

    return meses


def obtener_nombre_mes(mes):
    """
    Devuelve el nombre del mes.
    """

    nombres_meses = [
        '',
        'Enero',
        'Febrero',
        'Marzo',
        'Abril',
        'Mayo',
        'Junio',
        'Julio',
        'Agosto',
        'Septiembre',
        'Octubre',
        'Noviembre',
        'Diciembre'
    ]

    if not mes or mes < 1 or mes > 12:
        return ''

    return nombres_meses[mes]


# ============================================================
# LISTADO DE REPORTES
# ============================================================



# ============================================================
# CACHE DE IA (en memoria con TTL de 24 horas)
# ============================================================

_ia_cache = {}


def _cache_key_ia(image_bytes, contexto, anuncio):
    """Genera clave de cache unica basada en contenido de imagen + contexto."""
    import hashlib
    hasher = hashlib.md5()
    hasher.update(image_bytes)
    hasher.update(str(contexto).encode('utf-8'))
    hasher.update(str(anuncio).encode('utf-8'))
    return hasher.hexdigest()


def _get_cached_ia(image_bytes, contexto, anuncio):
    """Obtiene descripcion cacheada de IA si existe y no ha expirado."""
    from datetime import datetime
    key = _cache_key_ia(image_bytes, contexto, anuncio)
    entry = _ia_cache.get(key)
    if entry:
        ts, desc = entry
        if (datetime.utcnow() - ts).total_seconds() < 86400:  # 24h
            print(f'[CACHE HIT] IA cacheada para hash {key[:8]}...')
            return desc
        else:
            del _ia_cache[key]
    return None


def _set_cached_ia(image_bytes, contexto, anuncio, descripcion):
    """Guarda descripcion de IA en cache."""
    from datetime import datetime
    key = _cache_key_ia(image_bytes, contexto, anuncio)
    _ia_cache[key] = (datetime.utcnow(), descripcion)
    print(f'[CACHE SET] IA cacheada para hash {key[:8]}...')

@reportes_bp.route('/reportes')
@login_required
def reportes():
    """
    Lista los reportes mensuales del contrato activo.

    Permite:
    - búsqueda;
    - filtro por mes;
    - filtro por año;
    - filtro por obligación;
    - paginación.
    """

    contrato = (
        Contrato.query
        .filter_by(
            activo=True,
            user_id=current_user.id
        )
        .first()
    )

    # --------------------------------------------------------
    # Filtros
    # --------------------------------------------------------

    page = request.args.get(
        'page',
        1,
        type=int
    )

    per_page = request.args.get(
        'per_page',
        10,
        type=int
    )

    search = request.args.get(
        'search',
        '',
        type=str
    ).strip()

    filtro_mes = request.args.get(
        'filtro_mes',
        '',
        type=str
    )

    filtro_anio = request.args.get(
        'filtro_anio',
        '',
        type=str
    )

    filtro_obligacion = request.args.get(
        'filtro_obligacion',
        '',
        type=str
    )

    reportes_list = []

    reportes_pag = None

    obligaciones_list = []

    # ========================================================
    # CONTRATO ACTIVO
    # ========================================================

    if contrato:

        # ----------------------------------------------------
        # Obligaciones disponibles para el filtro
        # ----------------------------------------------------

        obligaciones_list = (
            Obligacion.query
            .filter_by(
                contrato_id=contrato.id
            )
            .order_by(
                Obligacion.numero
            )
            .all()
        )

        # ----------------------------------------------------
        # Consulta de reportes
        # ----------------------------------------------------

        query = (
            ReporteMensual.query
            .join(Obligacion)
            .filter(
                Obligacion.contrato_id ==
                contrato.id
            )
        )

        # ----------------------------------------------------
        # Buscar por número o descripción
        # ----------------------------------------------------

        if search:

            query = query.filter(
                db.or_(
                    Obligacion.descripcion.ilike(
                        f'%{search}%'
                    ),

                    Obligacion.numero
                    .cast(db.String)
                    .ilike(
                        f'%{search}%'
                    )
                )
            )

        # ----------------------------------------------------
        # Filtrar por mes
        # ----------------------------------------------------

        if filtro_mes:

            try:

                query = query.filter(
                    ReporteMensual.mes ==
                    int(filtro_mes)
                )

            except ValueError:

                pass

        # ----------------------------------------------------
        # Filtrar por año
        # ----------------------------------------------------

        if filtro_anio:

            try:

                query = query.filter(
                    ReporteMensual.anio ==
                    int(filtro_anio)
                )

            except ValueError:

                pass

        # ----------------------------------------------------
        # Filtrar por obligación
        # ----------------------------------------------------

        if filtro_obligacion:

            try:

                query = query.filter(
                    Obligacion.id ==
                    int(filtro_obligacion)
                )

            except ValueError:

                pass

        # ----------------------------------------------------
        # Orden y paginación
        # ----------------------------------------------------

        reportes_pag = (
            query
            .order_by(
                ReporteMensual.anio.desc(),
                ReporteMensual.mes.desc()
            )
            .paginate(
                page=page,
                per_page=per_page,
                error_out=False
            )
        )

        reportes_list = (
            reportes_pag.items
        )

    # ========================================================
    # RENDER
    # ========================================================

    return render_template(
        'reportes.html',

        reportes=reportes_list,

        reportes_pag=reportes_pag,

        contrato=contrato,

        obligaciones_list=obligaciones_list,

        search=search,

        filtro_mes=filtro_mes,

        filtro_anio=filtro_anio,

        filtro_obligacion=filtro_obligacion,

        per_page=per_page
    )


# ============================================================
# NUEVO REPORTE
# ============================================================

@reportes_bp.route(
    '/reporte/nuevo/<int:obligacion_id>',
    methods=['GET', 'POST']
)
@login_required
def nuevo_reporte(obligacion_id):
    """
    Crea un nuevo reporte mensual para una obligación.

    Validaciones:
    - El usuario debe ser propietario del contrato.
    - Las fechas deben pertenecer al mes seleccionado.
    - La fecha inicial no puede ser posterior a la final.
    - Los meses deben ser consecutivos.
    - No se permiten reportes duplicados.
    """

    # --------------------------------------------------------
    # Obtener obligación
    # --------------------------------------------------------

    obligacion = (
        Obligacion.query
        .get_or_404(
            obligacion_id
        )
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    if (
        not contrato
        or contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # Verificar contrato finalizado
    # --------------------------------------------------------

    if contrato.etapa == 'Reporte Cerrado':

        flash(
            'Este contrato está finalizado '
            '(Reporte Cerrado). No se pueden crear '
            'nuevos reportes.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.reportes'
            )
        )

    # --------------------------------------------------------
    # Meses disponibles
    # --------------------------------------------------------

    meses = generar_meses_contrato(
        contrato.fecha_inicio,
        contrato.fecha_fin
    )

    # --------------------------------------------------------
    # Datos guardados temporalmente en sesión
    # --------------------------------------------------------

    form_data = {
        'mes': session.pop(
            'nuevo_rep_mes',
            ''
        ),

        'anio': session.pop(
            'nuevo_rep_anio',
            ''
        ),

        'fecha_inicio_reporte': session.pop(
            'nuevo_rep_fecha_inicio',
            ''
        ),

        'fecha_fin_reporte': session.pop(
            'nuevo_rep_fecha_fin',
            ''
        )
    }

    # ========================================================
    # POST
    # ========================================================

    if request.method == 'POST':

        try:

            mes = int(
                request.form[
                    'mes'
                ]
            )

            anio = int(
                request.form[
                    'anio'
                ]
            )

            fecha_inicio_rep = datetime.strptime(
                request.form[
                    'fecha_inicio_reporte'
                ],
                '%Y-%m-%d'
            ).date()

            fecha_fin_rep = datetime.strptime(
                request.form[
                    'fecha_fin_reporte'
                ],
                '%Y-%m-%d'
            ).date()

        except (
            KeyError,
            ValueError,
            TypeError
        ):

            flash(
                'Los datos del reporte no son válidos.',
                'danger'
            )

            return redirect(
                url_for(
                    'reportes.nuevo_reporte',
                    obligacion_id=obligacion_id
                )
            )

        # ----------------------------------------------------
        # Validar mes
        # ----------------------------------------------------

        if mes < 1 or mes > 12:

            flash(
                'El mes seleccionado no es válido.',
                'danger'
            )

            return redirect(
                url_for(
                    'reportes.nuevo_reporte',
                    obligacion_id=obligacion_id
                )
            )

        nombre_mes = obtener_nombre_mes(
            mes
        )

        # ----------------------------------------------------
        # Límites del mes
        # ----------------------------------------------------

        _, last_day = calendar.monthrange(
            anio,
            mes
        )

        primer_dia_mes = date(
            anio,
            mes,
            1
        )

        ultimo_dia_mes = date(
            anio,
            mes,
            last_day
        )

        # ----------------------------------------------------
        # Fecha inicial dentro del mes
        # ----------------------------------------------------

        if (
            fecha_inicio_rep < primer_dia_mes
            or fecha_inicio_rep > ultimo_dia_mes
        ):

            session[
                'nuevo_rep_mes'
            ] = str(mes)

            session[
                'nuevo_rep_anio'
            ] = str(anio)

            session[
                'nuevo_rep_fecha_inicio'
            ] = request.form[
                'fecha_inicio_reporte'
            ]

            session[
                'nuevo_rep_fecha_fin'
            ] = request.form[
                'fecha_fin_reporte'
            ]

            flash(
                f'La fecha de inicio debe estar dentro de '
                f'{nombre_mes} {anio} '
                f'('
                f'{primer_dia_mes.strftime("%d/%m/%Y")}'
                f' – '
                f'{ultimo_dia_mes.strftime("%d/%m/%Y")}'
                f').',
                'danger'
            )

            return redirect(
                url_for(
                    'reportes.nuevo_reporte',
                    obligacion_id=obligacion_id
                )
            )

        # ----------------------------------------------------
        # Fecha final dentro del mes
        # ----------------------------------------------------

        if (
            fecha_fin_rep < primer_dia_mes
            or fecha_fin_rep > ultimo_dia_mes
        ):

            session[
                'nuevo_rep_mes'
            ] = str(mes)

            session[
                'nuevo_rep_anio'
            ] = str(anio)

            session[
                'nuevo_rep_fecha_inicio'
            ] = request.form[
                'fecha_inicio_reporte'
            ]

            session[
                'nuevo_rep_fecha_fin'
            ] = request.form[
                'fecha_fin_reporte'
            ]

            flash(
                f'La fecha de fin debe estar dentro de '
                f'{nombre_mes} {anio} '
                f'('
                f'{primer_dia_mes.strftime("%d/%m/%Y")}'
                f' – '
                f'{ultimo_dia_mes.strftime("%d/%m/%Y")}'
                f').',
                'danger'
            )

            return redirect(
                url_for(
                    'reportes.nuevo_reporte',
                    obligacion_id=obligacion_id
                )
            )

        # ----------------------------------------------------
        # Fecha inicial <= fecha final
        # ----------------------------------------------------

        if fecha_inicio_rep > fecha_fin_rep:

            session[
                'nuevo_rep_mes'
            ] = str(mes)

            session[
                'nuevo_rep_anio'
            ] = str(anio)

            session[
                'nuevo_rep_fecha_inicio'
            ] = request.form[
                'fecha_inicio_reporte'
            ]

            session[
                'nuevo_rep_fecha_fin'
            ] = request.form[
                'fecha_fin_reporte'
            ]

            flash(
                'La fecha de inicio no puede ser posterior '
                'a la fecha de fin.',
                'danger'
            )

            return redirect(
                url_for(
                    'reportes.nuevo_reporte',
                    obligacion_id=obligacion_id
                )
            )

        # ====================================================
        # VALIDAR CONSECUTIVIDAD
        # ====================================================

        if mes == 1:

            mes_ant = 12
            anio_ant = anio - 1

        else:

            mes_ant = mes - 1
            anio_ant = anio

        fecha_mes_ant = date(
            anio_ant,
            mes_ant,
            1
        )

        fecha_inicio_contrato_mes = date(
            contrato.fecha_inicio.year,
            contrato.fecha_inicio.month,
            1
        )

        if fecha_mes_ant >= fecha_inicio_contrato_mes:

            reporte_anterior = (
                ReporteMensual.query
                .filter_by(
                    mes=mes_ant,
                    anio=anio_ant,
                    obligacion_id=obligacion_id
                )
                .first()
            )

            if not reporte_anterior:

                nombres_meses = [
                    '',
                    'Enero',
                    'Febrero',
                    'Marzo',
                    'Abril',
                    'Mayo',
                    'Junio',
                    'Julio',
                    'Agosto',
                    'Septiembre',
                    'Octubre',
                    'Noviembre',
                    'Diciembre'
                ]

                session[
                    'nuevo_rep_mes'
                ] = str(mes)

                session[
                    'nuevo_rep_anio'
                ] = str(anio)

                session[
                    'nuevo_rep_fecha_inicio'
                ] = request.form[
                    'fecha_inicio_reporte'
                ]

                session[
                    'nuevo_rep_fecha_fin'
                ] = request.form[
                    'fecha_fin_reporte'
                ]

                flash(
                    f'No puede saltar meses. Cree primero '
                    f'el reporte de '
                    f'{nombres_meses[mes_ant]} '
                    f'{anio_ant} antes de '
                    f'{nombre_mes} {anio}.',
                    'danger'
                )

                return redirect(
                    url_for(
                        'reportes.nuevo_reporte',
                        obligacion_id=obligacion_id
                    )
                )

        # ====================================================
        # VALIDAR DUPLICADO
        # ====================================================

        existente = (
            ReporteMensual.query
            .filter_by(
                mes=mes,
                anio=anio,
                obligacion_id=obligacion_id
            )
            .first()
        )

        if existente:

            session[
                'nuevo_rep_mes'
            ] = str(mes)

            session[
                'nuevo_rep_anio'
            ] = str(anio)

            session[
                'nuevo_rep_fecha_inicio'
            ] = request.form[
                'fecha_inicio_reporte'
            ]

            session[
                'nuevo_rep_fecha_fin'
            ] = request.form[
                'fecha_fin_reporte'
            ]

            flash(
                f'Ya existe un reporte para '
                f'{existente.nombre_mes} {anio}.',
                'warning'
            )

            return redirect(
                url_for(
                    'reportes.nuevo_reporte',
                    obligacion_id=obligacion_id
                )
            )

        # ====================================================
        # CREAR REPORTE
        # ====================================================

        reporte = ReporteMensual(
            mes=mes,
            anio=anio,
            fecha_inicio_reporte=fecha_inicio_rep,
            fecha_fin_reporte=fecha_fin_rep,
            obligacion_id=obligacion_id
        )

        db.session.add(
            reporte
        )

        db.session.commit()

        flash(
            f'Reporte de {reporte.nombre_mes} '
            f'{anio} creado.',
            'success'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=reporte.id
            )
        )

    # ========================================================
    # GET
    # ========================================================

    return render_template(
        'nuevo_reporte.html',

        obligacion=obligacion,

        meses=meses,

        contrato=contrato,

        form_data=form_data
    )


# ============================================================
# VER REPORTE
# ============================================================

@reportes_bp.route(
    '/reporte/<int:id>'
)
@login_required
def ver_reporte(id):
    """
    Visualiza un reporte mensual y sus evidencias.
    """

    reporte = (
        ReporteMensual.query
        .get_or_404(id)
    )

    obligacion = (
        reporte.obligacion
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    if (
        not contrato
        or contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # Evidencias
    # --------------------------------------------------------

    evidencias = (
        Evidencia.query
        .filter_by(
            reporte_id=id
        )
        .order_by(
            Evidencia.numero_actividad
        )
        .all()
    )

    # --------------------------------------------------------
    # API Key configurada
    #
    # IMPORTANTE:
    # _obtener_api_key() solamente recupera la clave.
    # NO realiza una consulta a Gemini.
    # --------------------------------------------------------

    api_key_configurada = bool(
        _obtener_api_key()
    )

    # --------------------------------------------------------
    # Datos temporales del formulario
    # --------------------------------------------------------

    form_data = {
        'anuncio_usuario': session.pop(
            'evidencia_anuncio',
            ''
        ),

        'fecha_actividad': session.pop(
            'evidencia_fecha',
            ''
        )
    }

    return render_template(
        'ver_reporte.html',

        reporte=reporte,

        obligacion=obligacion,

        contrato=contrato,

        evidencias=evidencias,

        api_key_configurada=api_key_configurada,

        form_data=form_data
    )

# ============================================================
# ANALISIS IA EN BACKGROUND
# ============================================================

def _analizar_ia_background(app, evidencia_id, imagen_path, api_key, obligacion_desc, anuncio):
    """
    Analiza la imagen con Gemini en un hilo separado
    y actualiza la evidencia cuando termina.
    """
    import time
    time.sleep(1)  # Esperar a que el commit principal termine

    with app.app_context():
        from models import db, Evidencia
        from vision_analyzer import analizar_imagen

        # SQLAlchemy requiere sesión limpia por thread
        db.session.remove()

        try:
            descripcion = analizar_imagen(
                imagen_path,
                api_key=api_key,
                contexto_obligacion=obligacion_desc,
                anuncio_usuario=anuncio
            )

            if descripcion:
                evidencia = Evidencia.query.get(evidencia_id)
                if evidencia:
                    evidencia.descripcion_visual_ia = descripcion
                    evidencia.descripcion_actividad = descripcion
                    db.session.commit()
                    print(f'[IA Background] OK evidencia {evidencia_id}')
                else:
                    print(f'[IA Background] Evidencia {evidencia_id} no encontrada')
            else:
                print(f'[IA Background] Evidencia {evidencia_id}: Gemini no retornó descripción')

        except Exception as e:
            print(f'[IA Background] ERROR evidencia {evidencia_id}: {e}')
            db.session.rollback()
        finally:
            db.session.remove()

# ============================================================
# SUBIR EVIDENCIA
# ============================================================

@reportes_bp.route(
    '/reporte/<int:id>/evidencia',
    methods=['POST']
)
@login_required
def subir_evidencia(id):
    """
    Registra una evidencia de actividad.

    El Blueprint se encarga de:
    - autenticación;
    - autorización;
    - validaciones HTTP;
    - validación del período;
    - análisis mediante IA.

    EvidenciaService se encarga de:
    - obtener número de actividad;
    - guardar imagen;
    - generar descripción;
    - crear la evidencia.
    """

    # --------------------------------------------------------
    # OBTENER REPORTE
    # --------------------------------------------------------

    reporte = (
        ReporteMensual.query
        .get_or_404(id)
    )

    obligacion = (
        reporte.obligacion
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # SEGURIDAD
    # --------------------------------------------------------

    if (
        not contrato
        or contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # REPORTE CERRADO
    # --------------------------------------------------------

    if reporte.cerrado:

        flash(
            'Este reporte está cerrado. '
            'No se pueden agregar más evidencias.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # CONTRATO FINALIZADO
    # --------------------------------------------------------

    if contrato.etapa == 'Reporte Cerrado':

        flash(
            'Este contrato está finalizado '
            '(Reporte Cerrado). No se pueden agregar '
            'más evidencias.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # API KEY
    # --------------------------------------------------------

    api_key = _obtener_api_key()

    # --------------------------------------------------------
    # GUARDAR DATOS TEMPORALMENTE
    # --------------------------------------------------------

    session[
        'evidencia_anuncio'
    ] = request.form.get(
        'anuncio_usuario',
        ''
    )

    session[
        'evidencia_fecha'
    ] = request.form.get(
        'fecha_actividad',
        ''
    )

    # --------------------------------------------------------
    # VERIFICAR ARCHIVO
    # --------------------------------------------------------

    if 'imagen' not in request.files:

        flash(
            'No se seleccionó ningún archivo.',
            'danger'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    file = request.files[
        'imagen'
    ]

    anuncio_usuario = request.form.get(
        'anuncio_usuario',
        ''
    ).strip()

    # --------------------------------------------------------
    # VALIDAR ANUNCIO
    # --------------------------------------------------------

    if not anuncio_usuario:

        flash(
            'Debe escribir un anuncio/contexto.',
            'danger'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # VALIDAR NOMBRE
    # --------------------------------------------------------

    if file.filename == '':

        flash(
            'No se seleccionó ningún archivo.',
            'danger'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # VALIDAR EXTENSIÓN
    # --------------------------------------------------------

    if not allowed_file(
        file.filename
    ):

        flash(
            'Formato no permitido.',
            'danger'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # FECHA DE ACTIVIDAD
    # --------------------------------------------------------

    fecha_actividad_str = request.form.get(
        'fecha_actividad',
        ''
    ).strip()

    try:

        if fecha_actividad_str:

            fecha_actividad = datetime.strptime(
                fecha_actividad_str,
                '%Y-%m-%d'
            ).date()

        else:

            fecha_actividad = date.today()

    except ValueError:

        flash(
            'La fecha de actividad no es válida.',
            'danger'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # VALIDAR FECHA CONTRA PERÍODO DEL REPORTE
    # --------------------------------------------------------

    if (
        fecha_actividad
        < reporte.fecha_inicio_reporte
        or
        fecha_actividad
        > reporte.fecha_fin_reporte
    ):

        flash(
            f'La fecha de la actividad '
            f'({fecha_actividad.strftime("%d/%m/%Y")}) '
            f'debe estar dentro del periodo del reporte: '
            f'{reporte.fecha_inicio_reporte.strftime("%d/%m/%Y")}'
            f' a '
            f'{reporte.fecha_fin_reporte.strftime("%d/%m/%Y")}.',
            'danger'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # ========================================================
    # CREAR EVIDENCIA MEDIANTE SERVICE
    # ========================================================

    try:

        # ----------------------------------------------------
        # CREAR EVIDENCIA INMEDIATAMENTE (sin esperar IA)
        # ----------------------------------------------------

        evidencia_service = EvidenciaService()

        evidencia = (
            evidencia_service.crear_evidencia(
                reporte=reporte,
                imagen=file,
                anuncio=anuncio_usuario,
                fecha=fecha_actividad,
                descripcion=None  # IA se procesa en background
            )
        )

        # ----------------------------------------------------
        # GUARDAR
        # ----------------------------------------------------

        db.session.commit()

        # ----------------------------------------------------
        # ANÁLISIS IA EN BACKGROUND (no bloquea al usuario)
        # ----------------------------------------------------

        if api_key and evidencia.imagen_path:

            app = current_app._get_current_object()

            thread = threading.Thread(
                target=_analizar_ia_background,
                args=(
                    app,
                    evidencia.id,
                    evidencia.imagen_path,
                    api_key,
                    obligacion.descripcion,
                    anuncio_usuario
                ),
                daemon=True
            )

            thread.start()

            flash(
                'Evidencia guardada. El análisis con IA '
                'se está procesando en segundo plano.',
                'info'
            )

        else:

            flash(
                f'Actividad '
                f'{evidencia.numero_actividad} '
                f'registrada.',
                'success'
            )

        # ----------------------------------------------------
        # CREAR EVIDENCIA
        # ----------------------------------------------------

                # ----------------------------------------------------
        # LIMPIAR SESIÓN
        # ----------------------------------------------------

        session.pop(
            'evidencia_anuncio',
            None
        )

        session.pop(
            'evidencia_fecha',
            None
        )
    except RequestEntityTooLarge:

        db.session.rollback()

        flash(
            'El archivo es demasiado grande. '
            'Máximo 16MB.',
            'danger'
        )

    except Exception as e:

        db.session.rollback()

        flash(
            f'Error: {str(e)}',
            'danger'
        )

    return redirect(
        url_for(
            'reportes.ver_reporte',
            id=id
        )
    )


# ============================================================
# ELIMINAR EVIDENCIA
# ============================================================

@reportes_bp.route(
    '/reporte/<int:id>/evidencia/<int:evidencia_id>/eliminar',
    methods=['POST']
)
@login_required
def eliminar_evidencia(
    id,
    evidencia_id
):
    """
    Elimina una evidencia y renumera las actividades
    restantes.
    """

    evidencia = (
        Evidencia.query
        .get_or_404(
            evidencia_id
        )
    )

    reporte = (
        ReporteMensual.query
        .get_or_404(id)
    )

    obligacion = (
        reporte.obligacion
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    if (
        not contrato
        or contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # REPORTE CERRADO
    # --------------------------------------------------------

    if reporte.cerrado:

        flash(
            'Este reporte está cerrado. '
            'No se pueden eliminar evidencias.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )
    # --------------------------------------------------------
    # Guardar número antes de eliminar
    # --------------------------------------------------------

    numero_eliminado = (
        evidencia.numero_actividad
    )

    # --------------------------------------------------------
    # Eliminar archivo físico
    # --------------------------------------------------------

    try:

        if (
            evidencia.imagen_path
            and
            os.path.exists(
                evidencia.imagen_path
            )
        ):

            os.remove(
                evidencia.imagen_path
            )

    except Exception:

        pass

    # --------------------------------------------------------
    # Eliminar registro
    # --------------------------------------------------------

    db.session.delete(
        evidencia
    )

    db.session.commit()

    # --------------------------------------------------------
    # Renumerar evidencias restantes
    # --------------------------------------------------------

    evidencias_restantes = (
        Evidencia.query
        .filter_by(
            reporte_id=id
        )
        .order_by(
            Evidencia.numero_actividad
        )
        .all()
    )

    for idx, ev in enumerate(
        evidencias_restantes,
        start=1
    ):

        ev.numero_actividad = idx

    db.session.commit()

    flash(
        f'Evidencia de Actividad '
        f'{numero_eliminado} eliminada.',
        'info'
    )

    return redirect(
        url_for(
            'reportes.ver_reporte',
            id=id
        )
    )


# ============================================================
# EDITAR EVIDENCIA
# ============================================================

@reportes_bp.route(
    '/reporte/<int:id>/evidencia/<int:evidencia_id>/editar',
    methods=['POST']
)
@login_required
def editar_evidencia(
    id,
    evidencia_id
):
    """
    Actualiza la descripción y fecha de una evidencia.
    """

    evidencia = (
        Evidencia.query
        .get_or_404(
            evidencia_id
        )
    )

    reporte = (
        ReporteMensual.query
        .get_or_404(id)
    )

    obligacion = (
        reporte.obligacion
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    if (
        not contrato
        or contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # REPORTE CERRADO
    # --------------------------------------------------------

    if reporte.cerrado:

        flash(
            'Este reporte está cerrado. '
            'No se pueden editar evidencias.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # Descripción
    # --------------------------------------------------------

    nueva_descripcion = request.form.get(
        'descripcion_actividad',
        ''
    ).strip()

    if nueva_descripcion:

        evidencia.descripcion_actividad = (
            nueva_descripcion
        )

    # --------------------------------------------------------
    # Fecha
    # --------------------------------------------------------

    fecha_actividad_str = request.form.get(
        'fecha_actividad',
        ''
    ).strip()

    if fecha_actividad_str:

        try:

            fecha_actividad = datetime.strptime(
                fecha_actividad_str,
                '%Y-%m-%d'
            ).date()

            # -----------------------------------------------
            # Validar que la fecha continúe dentro del
            # período del reporte.
            # -----------------------------------------------

            if (
                fecha_actividad
                < reporte.fecha_inicio_reporte
                or
                fecha_actividad
                > reporte.fecha_fin_reporte
            ):

                flash(
                    'La fecha de la actividad debe estar '
                    'dentro del período del reporte.',
                    'danger'
                )

                return redirect(
                    url_for(
                        'reportes.ver_reporte',
                        id=id
                    )
                )

            evidencia.fecha_actividad = (
                fecha_actividad
            )

        except ValueError:

            flash(
                'La fecha de actividad no es válida.',
                'danger'
            )

            return redirect(
                url_for(
                    'reportes.ver_reporte',
                    id=id
                )
            )

    # --------------------------------------------------------
    # Guardar
    # --------------------------------------------------------

    db.session.commit()

    flash(
        f'Actividad '
        f'{evidencia.numero_actividad} '
        f'actualizada.',
        'success'
    )

    return redirect(
        url_for(
            'reportes.ver_reporte',
            id=id
        )
    )


# ============================================================
# GENERAR PDF INDIVIDUAL
# ============================================================

@reportes_bp.route(
    '/reporte/<int:id>/pdf'
)
@login_required
def generar_pdf(id):
    """
    Genera y descarga el PDF de un reporte mensual.
    """

    reporte = (
        ReporteMensual.query
        .get_or_404(id)
    )

    obligacion = (
        reporte.obligacion
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    if (
        not contrato
        or contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # Evidencias
    # --------------------------------------------------------

    evidencias = (
        Evidencia.query
        .filter_by(
            reporte_id=id
        )
        .order_by(
            Evidencia.numero_actividad
        )
        .all()
    )

    # --------------------------------------------------------
    # No generar PDF vacío
    # --------------------------------------------------------

    if not evidencias:

        flash(
            'No se puede generar el PDF porque este '
            'reporte no tiene evidencias registradas. '
            'Agregue al menos una evidencia antes '
            'de descargar.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # Nombre y ruta
    # --------------------------------------------------------

    pdf_filename = (
        f'Reporte_Obligacion_'
        f'{obligacion.numero}_'
        f'{reporte.nombre_mes}_'
        f'{reporte.anio}.pdf'
    )

    pdf_path = os.path.join(
        current_app.config[
            'PDF_FOLDER'
        ],
        pdf_filename
    )

    # ========================================================
    # GENERAR
    # ========================================================

    try:

        generator = PDFGenerator(
            pdf_path
        )

        generator.generar_reporte(
            reporte,
            obligacion,
            evidencias,
            contrato
        )
        return send_from_directory(
            current_app.config[
                'PDF_FOLDER'
            ],
            pdf_filename,
            as_attachment=True
        )

    except Exception as e:

        flash(
            f'Error al generar PDF: {str(e)}',
            'danger'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )


# ============================================================
# ELIMINAR REPORTE
# ============================================================

@reportes_bp.route(
    '/reporte/<int:id>/eliminar',
    methods=['POST']
)
@login_required
def eliminar_reporte(id):
    """
    Elimina un reporte y sus archivos de evidencias.
    """

    reporte = (
        ReporteMensual.query
        .get_or_404(id)
    )

    obligacion = (
        reporte.obligacion
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    if (
        not contrato
        or contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # Eliminar archivos físicos
    # --------------------------------------------------------

    for ev in reporte.evidencias:

        try:

            if (
                ev.imagen_path
                and
                os.path.exists(
                    ev.imagen_path
                )
            ):

                os.remove(
                    ev.imagen_path
                )

        except Exception:

            pass

    # --------------------------------------------------------
    # Eliminar reporte
    # --------------------------------------------------------

    db.session.delete(
        reporte
    )

    db.session.commit()

    flash(
        'Reporte eliminado.',
        'info'
    )

    return redirect(
        url_for(
            'reportes.reportes'
        )
    )


# ============================================================
# SERVIR ARCHIVOS SUBIDOS
# ============================================================

@reportes_bp.route(
    '/uploads/<path:filename>'
)
@login_required
def uploaded_file(filename):
    """
    Sirve imágenes de evidencias.

    El sistema almacena actualmente rutas físicas completas
    en Evidencia.imagen_path. La plantilla envía únicamente
    el nombre del archivo.

    Esta función normaliza separadores para garantizar
    compatibilidad entre Windows y otros sistemas.
    """

    upload_folder = current_app.config.get(
        'UPLOAD_FOLDER'
    )

    if not upload_folder:
        return (
            'Carpeta de evidencias no configurada.',
            500
        )

    # --------------------------------------------------------
    # Normalizar separadores
    # --------------------------------------------------------

    filename = str(
        filename or ''
    ).replace(
        '\\',
        '/'
    )

    # --------------------------------------------------------
    # Evitar que llegue una ruta física completa
    # --------------------------------------------------------

    filename = os.path.basename(
        filename
    )

    if not filename:
        return (
            'Archivo no especificado.',
            404
        )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    filename = secure_filename(
        filename
    )

    if not filename:
        return (
            'Nombre de archivo no válido.',
            400
        )

    # --------------------------------------------------------
    # Ruta física
    # --------------------------------------------------------

    ruta_archivo = os.path.join(
        upload_folder,
        filename
    )

    # --------------------------------------------------------
    # Verificar existencia
    # --------------------------------------------------------

    if not os.path.isfile(
        ruta_archivo
    ):

        current_app.logger.warning(
            'Imagen de evidencia no encontrada: %s',
            ruta_archivo
        )

        return (
            'Imagen de evidencia no encontrada.',
            404
        )

    # --------------------------------------------------------
    # Servir imagen
    # --------------------------------------------------------

    return send_from_directory(
        upload_folder,
        filename
    )

# ============================================================
# DESCARGA MASIVA DE PDF POR MES
# ============================================================

@reportes_bp.route(
    '/reportes/descargar-mes',
    methods=['POST']
)
@login_required
def descargar_masivo_mes():
    """
    Genera un ZIP con todos los PDFs de las obligaciones
    correspondientes a un mes específico.

    Requisitos:
    - contrato activo;
    - todas las obligaciones deben tener reporte;
    - todos los reportes deben tener evidencias.
    """

    # --------------------------------------------------------
    # Detectar AJAX
    # --------------------------------------------------------

    es_ajax = (
        request.headers.get(
            'X-Requested-With'
        )
        == 'XMLHttpRequest'
    )

    def responder_error(
        mensaje,
        codigo=400
    ):

        if es_ajax:

            return jsonify(
                {
                    'error': mensaje
                }
            ), codigo

        flash(
            mensaje,
            'danger'
        )

        return redirect(
            url_for(
                'reportes.reportes'
            )
        )

    # --------------------------------------------------------
    # Obtener mes y año
    # --------------------------------------------------------

    try:

        mes = int(
            request.form.get(
                'mes',
                0
            )
        )

        anio = int(
            request.form.get(
                'anio',
                0
            )
        )

    except (
        TypeError,
        ValueError
    ):

        return responder_error(
            'Seleccione mes y año.'
        )

    if not mes or not anio:

        return responder_error(
            'Seleccione mes y año.'
        )

    if mes < 1 or mes > 12:

        return responder_error(
            'El mes seleccionado no es válido.'
        )

    # --------------------------------------------------------
    # Contrato activo
    # --------------------------------------------------------

    contrato = (
        Contrato.query
        .filter_by(
            activo=True,
            user_id=current_user.id
        )
        .first()
    )

    if not contrato:

        return responder_error(
            'No hay contrato activo.'
        )

    nombre_mes = obtener_nombre_mes(
        mes
    )

    # --------------------------------------------------------
    # Obligaciones
    # --------------------------------------------------------

    obligaciones = (
        Obligacion.query
        .filter_by(
            contrato_id=contrato.id
        )
        .order_by(
            Obligacion.numero
        )
        .all()
    )

    if not obligaciones:

        return responder_error(
            'No hay obligaciones registradas.'
        )

    # ========================================================
    # VALIDAR REPORTES Y EVIDENCIAS
    # ========================================================

    obligaciones_faltantes = []

    obligaciones_sin_evidencia = []

    reportes_validos = []

    for obligacion in obligaciones:

        reporte = (
            ReporteMensual.query
            .filter_by(
                mes=mes,
                anio=anio,
                obligacion_id=obligacion.id
            )
            .first()
        )

        # ----------------------------------------------------
        # Falta reporte
        # ----------------------------------------------------

        if not reporte:

            obligaciones_faltantes.append(
                f'No. {obligacion.numero}'
            )

        # ----------------------------------------------------
        # Reporte sin evidencia
        # ----------------------------------------------------

        elif not reporte.evidencias:

            obligaciones_sin_evidencia.append(
                f'No. {obligacion.numero}'
            )

        else:

            reportes_validos.append(
                reporte
            )

    # ========================================================
    # VALIDACIÓN FALLIDA
    # ========================================================

    if (
        obligaciones_faltantes
        or obligaciones_sin_evidencia
    ):

        mensajes = []

        if obligaciones_faltantes:

            mensajes.append(
                f'<strong>'
                f'Sin reporte para '
                f'{nombre_mes} {anio}:'
                f'</strong> '
                f'{", ".join(obligaciones_faltantes)}'
            )

        if obligaciones_sin_evidencia:

            mensajes.append(
                f'<strong>'
                f'Reporte sin evidencias para '
                f'{nombre_mes} {anio}:'
                f'</strong> '
                f'{", ".join(obligaciones_sin_evidencia)}'
            )

        error_html = (
            '<strong>'
            'No se puede descargar el ZIP.'
            '</strong><br>'
            'Todas las obligaciones deben tener '
            'un reporte con al menos una evidencia '
            'para el mes seleccionado.'
            '<br><br>'
            +
            '<br>'.join(
                mensajes
            )
        )

        return responder_error(
            error_html
        )

    # --------------------------------------------------------
    # Sin reportes válidos
    # --------------------------------------------------------

    if not reportes_validos:

        return responder_error(
            f'No hay reportes con evidencias '
            f'para {nombre_mes} {anio}.'
        )

    # ========================================================
    # GENERAR ZIP EN MEMORIA
    # ========================================================

    memory_zip = io.BytesIO()

    with zipfile.ZipFile(
        memory_zip,
        'w',
        zipfile.ZIP_DEFLATED
    ) as zf:

        for reporte in reportes_validos:

            obligacion = (
                reporte.obligacion
            )

            evidencias = (
                Evidencia.query
                .filter_by(
                    reporte_id=reporte.id
                )
                .order_by(
                    Evidencia.numero_actividad
                )
                .all()
            )

            pdf_filename = (
                f'Reporte_Obligacion_'
                f'{obligacion.numero}_'
                f'{reporte.nombre_mes}_'
                f'{reporte.anio}.pdf'
            )

            pdf_path = os.path.join(
                current_app.config[
                    'PDF_FOLDER'
                ],
                pdf_filename
            )

            try:

                generator = PDFGenerator(
                    pdf_path
                )

                generator.generar_reporte(
                    reporte,
                    obligacion,
                    evidencias,
                    contrato
                )

                zf.write(
                    pdf_path,
                    arcname=pdf_filename
                )

            except Exception as e:

                print(
                    'Error generando PDF para '
                    f'obligación {obligacion.numero}: '
                    f'{e}'
                )

                continue

    # --------------------------------------------------------
    # Preparar respuesta
    # --------------------------------------------------------

    memory_zip.seek(0)

    zip_filename = (
        f'Reportes_'
        f'{contrato.contratista or "Contrato"}_'
        f'{mes}_'
        f'{anio}.zip'
    )

    return send_file(
        memory_zip,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_filename
    )


# ============================================================
# EXCEL CONSOLIDADO
# ============================================================

@reportes_bp.route(
    '/reporte/consolidado/excel'
)
@login_required
def generar_excel_consolidado():
    """
    Genera un Excel consolidado de todas las obligaciones
    para un mes y año específicos.

    El resumen ejecutivo de las actividades se genera
    utilizando Gemini cuando existe una API Key configurada.
    """

    # --------------------------------------------------------
    # Importar openpyxl aquí
    #
    # Esto evita cargar la librería cuando no se necesita.
    # --------------------------------------------------------

    from openpyxl import (
        Workbook
    )

    from openpyxl.styles import (
        Font,
        Alignment,
        Border,
        Side,
        PatternFill
    )

    # --------------------------------------------------------
    # Detectar AJAX
    # --------------------------------------------------------

    es_ajax = (
        request.headers.get(
            'X-Requested-With'
        )
        == 'XMLHttpRequest'
    )

    def responder_error(
        mensaje,
        codigo=400
    ):

        if es_ajax:

            return jsonify(
                {
                    'error': mensaje
                }
            ), codigo

        flash(
            mensaje,
            'danger'
        )

        return redirect(
            url_for(
                'inicio.index'
            )
        )

    def responder_warning(
        mensaje
    ):

        if es_ajax:

            return jsonify(
                {
                    'error': mensaje
                }
            ), 400

        flash(
            mensaje,
            'warning'
        )

        return redirect(
            url_for(
                'inicio.index'
            )
        )

    try:

        # ====================================================
        # CONTRATO ACTIVO
        # ====================================================

        contrato = (
            Contrato.query
            .filter_by(
                activo=True,
                user_id=current_user.id
            )
            .first()
        )

        if not contrato:

            return responder_error(
                'No hay contrato activo configurado.'
            )

        # ====================================================
        # MES Y AÑO
        # ====================================================

        mes = request.args.get(
            'mes',
            type=int
        )

        anio = request.args.get(
            'anio',
            type=int
        )

        if not mes or not anio:

            return responder_warning(
                'Debe seleccionar mes y año '
                'para generar el consolidado.'
            )

        if mes < 1 or mes > 12:

            return responder_warning(
                'El mes seleccionado no es válido.'
            )

        nombre_mes = obtener_nombre_mes(
            mes
        )

        # ====================================================
        # OBLIGACIONES
        # ====================================================

        obligaciones = (
            Obligacion.query
            .filter_by(
                contrato_id=contrato.id
            )
            .order_by(
                Obligacion.numero
            )
            .all()
        )

        if not obligaciones:

            return responder_error(
                'No hay obligaciones registradas.'
            )

        # ====================================================
        # VALIDAR REPORTES Y EVIDENCIAS
        # ====================================================

        obligaciones_faltantes = []

        obligaciones_sin_evidencia = []

        reportes_validos = {}

        for obligacion in obligaciones:

            reporte = (
                ReporteMensual.query
                .filter_by(
                    mes=mes,
                    anio=anio,
                    obligacion_id=obligacion.id
                )
                .first()
            )

            if not reporte:

                obligaciones_faltantes.append(
                    f'No. {obligacion.numero}'
                )

            else:

                if not reporte.evidencias:

                    obligaciones_sin_evidencia.append(
                        f'No. {obligacion.numero}'
                    )

                else:

                    reportes_validos[
                        obligacion.id
                    ] = reporte

        # ====================================================
        # VALIDACIÓN
        # ====================================================

        if (
            obligaciones_faltantes
            or obligaciones_sin_evidencia
        ):

            mensajes = []

            if obligaciones_faltantes:

                mensajes.append(
                    f'<strong>'
                    f'Sin reporte para '
                    f'{nombre_mes} {anio}:'
                    f'</strong> '
                    f'{", ".join(obligaciones_faltantes)}'
                )

            if obligaciones_sin_evidencia:

                mensajes.append(
                    f'<strong>'
                    f'Reporte sin evidencias para '
                    f'{nombre_mes} {anio}:'
                    f'</strong> '
                    f'{", ".join(obligaciones_sin_evidencia)}'
                )

            error_html = (
                '<strong>'
                'No se puede generar el consolidado.'
                '</strong><br>'
                'Todas las obligaciones deben tener '
                'un reporte con al menos una evidencia '
                'para el mes seleccionado.'
                '<br><br>'
                +
                '<br>'.join(
                    mensajes
                )
            )

            return responder_error(
                error_html
            )

        # ====================================================
        # OBTENER API KEY
        # ====================================================

        api_key = _obtener_api_key()

        # ====================================================
        # CREAR LIBRO
        # ====================================================

        wb = Workbook()

        ws = wb.active

        ws.title = (
            f'Consolidado_'
            f'{nombre_mes}_'
            f'{anio}'
        )

        # ====================================================
        # ESTILOS
        # ====================================================

        titulo_fill = PatternFill(
            start_color='D9EAF7',
            end_color='D9EAF7',
            fill_type='solid'
        )

        encabezado_fill = PatternFill(
            start_color='2C3E50',
            end_color='2C3E50',
            fill_type='solid'
        )

        encabezado_font = Font(
            bold=True,
            color='FFFFFF'
        )

        titulo_font = Font(
            bold=True,
            size=14,
            color='2C3E50'
        )

        subtitulo_font = Font(
            italic=True,
            size=10,
            color='666666'
        )

        cell_align = Alignment(
            vertical='top',
            wrap_text=True
        )

        center_align = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

        thin_side = Side(
            style='thin',
            color='CCCCCC'
        )

        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        # ====================================================
        # INFORMACIÓN DEL DOCUMENTO
        # ====================================================

        ws['A1'] = (
            'CONSOLIDADO DE OBLIGACIONES CONTRACTUALES'
        )

        ws['A2'] = (
            f'Contrato: '
            f'{contrato.numero_contrato or ""}'
        )

        ws['A3'] = (
            f'Contratista: '
            f'{contrato.contratista or ""}'
        )

        ws['A4'] = (
            f'Periodo: '
            f'{nombre_mes} {anio}'
        )

        # ====================================================
        # ENCABEZADOS
        # ====================================================

        ws.append(
            [
                'Número',
                'Obligación',
                'Resumen Ejecutivo'
            ]
        )

        # ====================================================
        # FORMATO ENCABEZADOS
        # ====================================================

        for cell in ws[5]:

            cell.font = encabezado_font

            cell.fill = encabezado_fill

            cell.alignment = center_align

            cell.border = thin_border

        # ====================================================
        # PROCESAR OBLIGACIONES
        # ====================================================

        for obligacion in obligaciones:

            reporte = reportes_validos[
                obligacion.id
            ]

            evidencias = (
                Evidencia.query
                .filter_by(
                    reporte_id=reporte.id
                )
                .order_by(
                    Evidencia.numero_actividad
                )
                .all()
            )

            actividades_textos = [
                ev.descripcion_actividad
                for ev in evidencias
                if ev.descripcion_actividad
            ]

            # ------------------------------------------------
            # Resumen ejecutivo
            # ------------------------------------------------

            if actividades_textos:

                if api_key:

                    try:

                        texto_ejecutivo = (
                            consolidar_textos_ejecutivo(
                                actividades_textos,
                                api_key,
                                obligacion=obligacion.descripcion,
                                periodo=f'{nombre_mes} {anio}'
                            )
                        )

                    except Exception as e:

                        print(
                            'Error consolidando texto '
                            f'de obligación '
                            f'{obligacion.numero}: '
                            f'{e}'
                        )

                        texto_ejecutivo = (
                            'No fue posible generar '
                            'el resumen ejecutivo '
                            'con IA.'
                        )

                else:

                    texto_ejecutivo = (
                        '\n'.join(
                            actividades_textos
                        )
                    )

            else:

                texto_ejecutivo = (
                    'Sin actividades reportadas.'
                )

            # ------------------------------------------------
            # Agregar fila
            # ------------------------------------------------

            ws.append(
                [
                    obligacion.numero,
                    obligacion.descripcion,
                    texto_ejecutivo
                ]
            )

            row_idx = ws.max_row

            for col_idx in range(
                1,
                4
            ):

                cell = ws.cell(
                    row=row_idx,
                    column=col_idx
                )

                cell.alignment = cell_align

                cell.border = thin_border

        # ====================================================
        # ANCHOS DE COLUMNA
        # ====================================================

        ws.column_dimensions[
            'A'
        ].width = 18

        ws.column_dimensions[
            'B'
        ].width = 50

        ws.column_dimensions[
            'C'
        ].width = 90

        # ====================================================
        # ALTURA DE FILAS
        # ====================================================

        for row in ws.iter_rows(
            min_row=6,
            max_row=ws.max_row
        ):

            ws.row_dimensions[
                row[0].row
            ].height = 80

        # ====================================================
        # ESTILO TÍTULO
        # ====================================================

        for cell in ws[1]:

            cell.font = titulo_font

            cell.fill = titulo_fill

        for cell in ws[2]:

            cell.font = subtitulo_font

        for cell in ws[3]:

            cell.font = subtitulo_font

        for cell in ws[4]:

            cell.font = subtitulo_font

        # ====================================================
        # CONGELAR ENCABEZADOS
        # ====================================================

        ws.freeze_panes = 'A6'

        # ====================================================
        # GUARDAR EN MEMORIA
        # ====================================================

        output = io.BytesIO()

        wb.save(
            output
        )

        output.seek(0)

        # ====================================================
        # NOMBRE DEL ARCHIVO
        # ====================================================

        filename = (
            f'Consolidado_'
            f'{contrato.contratista or "Contrato"}_'
            f'{nombre_mes}_'
            f'{anio}.xlsx'
        )

        # ====================================================
        # DESCARGA
        # ====================================================

        return send_file(
            output,

            mimetype=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),

            as_attachment=True,

            download_name=filename
        )

    except Exception as e:

        import traceback

        traceback.print_exc()

        error_msg = (
            'Error inesperado al generar '
            f'el consolidado: {str(e)}'
        )

        if es_ajax:

            return jsonify(
                {
                    'error': error_msg
                }
            ), 500

        flash(
            error_msg,
            'danger'
        )

        return redirect(
            url_for(
                'inicio.index'
            )
        )
        # ============================================================
# CERRAR REPORTE
# ============================================================

@reportes_bp.route(
    '/reporte/<int:id>/cerrar',
    methods=['POST']
)
@login_required
def cerrar_reporte(id):
    """
    Cierra un reporte mensual.

    Condiciones:
    1. El reporte debe tener al menos una evidencia.
    2. Todas las obligaciones del contrato deben tener
       al menos un reporte con evidencia en ese mes.
    3. Solo se puede cerrar el mes anterior al actual
       (o meses anteriores).
    """

    reporte = (
        ReporteMensual.query
        .get_or_404(id)
    )

    obligacion = (
        reporte.obligacion
    )

    contrato = (
        Contrato.query
        .get(
            obligacion.contrato_id
        )
    )

    # --------------------------------------------------------
    # Seguridad
    # --------------------------------------------------------

    if (
        not contrato
        or contrato.user_id != current_user.id
    ):

        flash(
            'No tiene permiso.',
            'danger'
        )

        return redirect(
            url_for(
                'inicio.inicio'
            )
        )

    # --------------------------------------------------------
    # Ya está cerrado
    # --------------------------------------------------------

    if reporte.cerrado:

        flash(
            'Este reporte ya está cerrado.',
            'info'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # Solo cerrar meses anteriores al actual
    # --------------------------------------------------------

    hoy = date.today()

    if (
        reporte.anio > hoy.year
        or (
            reporte.anio == hoy.year
            and reporte.mes >= hoy.month
        )
    ):

        flash(
            'Solo se pueden cerrar reportes de meses '
            'anteriores al actual.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # El reporte debe tener al menos una evidencia
    # --------------------------------------------------------

    if not reporte.evidencias:

        flash(
            'No se puede cerrar un reporte sin evidencias.',
            'warning'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # Todas las obligaciones deben tener reporte con evidencia
    # en ese mes
    # --------------------------------------------------------

    obligaciones = (
        Obligacion.query
        .filter_by(
            contrato_id=contrato.id
        )
        .all()
    )

    obligaciones_faltantes = []

    for obl in obligaciones:

        rep = (
            ReporteMensual.query
            .filter_by(
                mes=reporte.mes,
                anio=reporte.anio,
                obligacion_id=obl.id
            )
            .first()
        )

        if not rep:

            obligaciones_faltantes.append(
                f'Obligación No. {obl.numero} '
                f'(sin reporte)'
            )

        elif not rep.evidencias:

            obligaciones_faltantes.append(
                f'Obligación No. {obl.numero} '
                f'(sin evidencias)'
            )

    if obligaciones_faltantes:

        flash(
            'No se puede cerrar el mes porque '
            'faltan evidencias en: '
            + ', '.join(obligaciones_faltantes),
            'danger'
        )

        return redirect(
            url_for(
                'reportes.ver_reporte',
                id=id
            )
        )

    # --------------------------------------------------------
    # Cerrar reporte
    # --------------------------------------------------------

    reporte.cerrado = True

    db.session.commit()

    flash(
        f'Reporte de {reporte.nombre_mes} {reporte.anio} '
        'cerrado exitosamente. '
        'Ahora es de solo lectura.',
        'success'
    )

    return redirect(
        url_for(
            'reportes.ver_reporte',
            id=id
        )
    )

# ============================================================
# CERRAR MES REPORTADO (TODAS LAS OBLIGACIONES)
# ============================================================

@reportes_bp.route(
    '/reportes/cerrar-mes',
    methods=['POST']
)
@login_required
def cerrar_mes_reportado():
    """
    Cierra todos los reportes de un mes específico.

    Validaciones:
    1. Mes y año son obligatorios.
    2. Solo se pueden cerrar meses anteriores al actual.
    3. Todas las obligaciones del contrato deben tener un reporte
       en ese mes con al menos una evidencia.
    """

    # --------------------------------------------------------
    # Mes y año obligatorios
    # --------------------------------------------------------

    try:
        mes = int(request.form.get('mes', 0))
        anio = int(request.form.get('anio', 0))
    except (TypeError, ValueError):
        flash('Debe seleccionar mes y año válidos.', 'danger')
        return redirect(url_for('reportes.reportes'))

    if not mes or not anio:
        flash('Mes y año son obligatorios.', 'danger')
        return redirect(url_for('reportes.reportes'))

    if mes < 1 or mes > 12:
        flash('El mes seleccionado no es válido.', 'danger')
        return redirect(url_for('reportes.reportes'))

    nombre_mes = obtener_nombre_mes(mes)

    # --------------------------------------------------------
    # Solo cerrar meses anteriores al actual
    # --------------------------------------------------------

    hoy = date.today()

    if anio > hoy.year or (anio == hoy.year and mes >= hoy.month):
        flash(
            f'Solo se pueden cerrar meses anteriores al actual '
            f'({obtener_nombre_mes(hoy.month)} {hoy.year}).',
            'warning'
        )
        return redirect(url_for('reportes.reportes'))

    # --------------------------------------------------------
    # Contrato activo
    # --------------------------------------------------------

    contrato = (
        Contrato.query
        .filter_by(activo=True, user_id=current_user.id)
        .first()
    )

    if not contrato:
        flash('No hay contrato activo.', 'warning')
        return redirect(url_for('reportes.reportes'))

    # --------------------------------------------------------
    # Obtener obligaciones
    # --------------------------------------------------------

    obligaciones = (
        Obligacion.query
        .filter_by(contrato_id=contrato.id)
        .all()
    )

    if not obligaciones:
        flash('No hay obligaciones registradas.', 'warning')
        return redirect(url_for('reportes.reportes'))

    # --------------------------------------------------------
    # Validar que cada obligación tenga reporte + evidencias
    # --------------------------------------------------------

    obligaciones_faltantes = []

    for obl in obligaciones:
        reporte = (
            ReporteMensual.query
            .filter_by(
                mes=mes,
                anio=anio,
                obligacion_id=obl.id
            )
            .first()
        )

        if not reporte:
            obligaciones_faltantes.append(
                f'Obligación No. {obl.numero} (sin reporte)'
            )
        elif not reporte.evidencias:
            obligaciones_faltantes.append(
                f'Obligación No. {obl.numero} (sin actividades)'
            )

    if obligaciones_faltantes:
        flash(
            'No se puede cerrar el mes porque faltan reportes o '
            'actividades en: ' + '; '.join(obligaciones_faltantes),
            'danger'
        )
        return redirect(url_for('reportes.reportes'))

    # --------------------------------------------------------
    # Cerrar todos los reportes del mes
    # --------------------------------------------------------

    reportes_mes = (
        ReporteMensual.query
        .join(Obligacion)
        .filter(
            Obligacion.contrato_id == contrato.id,
            ReporteMensual.mes == mes,
            ReporteMensual.anio == anio
        )
        .all()
    )

    for rep in reportes_mes:
        rep.cerrado = True

    db.session.commit()

    flash(
        f'Mes {nombre_mes} {anio} cerrado exitosamente. '
        f'Todos los reportes ({len(reportes_mes)}) son ahora de solo lectura.',
        'success'
    )

    return redirect(url_for('reportes.reportes'))
 
# ============================================================
# SELECCIONAR OBLIGACIÓN PARA NUEVO REPORTE
# ============================================================

@reportes_bp.route(
    '/reporte/nuevo'
)
@login_required
def nuevo_reporte_selector():
    """
    Muestra las obligaciones del contrato activo para que
    el usuario seleccione cuál desea reportar.
    """

    contrato = (
        Contrato.query
        .filter_by(
            activo=True,
            user_id=current_user.id
        )
        .first()
    )

    if not contrato:

        flash(
            'No hay contrato activo.',
            'warning'
        )

        return redirect(
            url_for(
                'contratos.contratos'
            )
        )

    obligaciones = (
        Obligacion.query
        .filter_by(
            contrato_id=contrato.id
        )
        .order_by(
            Obligacion.numero
        )
        .all()
    )

    if not obligaciones:

        flash(
            'No hay obligaciones registradas. '
            'Cree obligaciones primero.',
            'warning'
        )

        return redirect(
            url_for(
                'contratos.contratos'
            )
        )

    return render_template(
        'seleccionar_obligacion.html',
        obligaciones=obligaciones,
        contrato=contrato
    ) 
    
# ============================================================
# SUBIR EVIDENCIA VIA AJAX
# ============================================================

@reportes_bp.route(
    '/reporte/<int:id>/evidencia/ajax',
    methods=['POST']
)
@login_required
def subir_evidencia_ajax(id):
    """
    Subida de evidencia via AJAX.
    El analisis IA se ejecuta en background para no bloquear al usuario.
    """
    reporte = ReporteMensual.query.get_or_404(id)
    obligacion = reporte.obligacion
    contrato = Contrato.query.get(obligacion.contrato_id)

    if not contrato or contrato.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'No tiene permiso.'}), 403

    if reporte.cerrado:
        return jsonify({'success': False, 'error': 'Reporte cerrado.'}), 400

    if contrato.etapa == 'Reporte Cerrado':
        return jsonify({'success': False, 'error': 'Contrato finalizado.'}), 400

    api_key = _obtener_api_key()

    if 'imagen' not in request.files:
        return jsonify({'success': False, 'error': 'No se selecciono archivo.'}), 400

    file = request.files['imagen']
    anuncio_usuario = request.form.get('anuncio_usuario', '').strip()

    if not anuncio_usuario:
        return jsonify({'success': False, 'error': 'El anuncio/contexto es obligatorio.'}), 400

    if file.filename == '':
        return jsonify({'success': False, 'error': 'Archivo vacio.'}), 400

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Formato no permitido.'}), 400

    fecha_actividad_str = request.form.get('fecha_actividad', '').strip()
    try:
        if fecha_actividad_str:
            fecha_actividad = datetime.strptime(fecha_actividad_str, '%Y-%m-%d').date()
        else:
            fecha_actividad = date.today()
    except ValueError:
        return jsonify({'success': False, 'error': 'Fecha invalida.'}), 400

    if fecha_actividad < reporte.fecha_inicio_reporte or fecha_actividad > reporte.fecha_fin_reporte:
        return jsonify({
            'success': False,
            'error': f'La fecha debe estar entre {reporte.fecha_inicio_reporte.strftime("%d/%m/%Y")} y {reporte.fecha_fin_reporte.strftime("%d/%m/%Y")}.'
        }), 400

    try:
        evidencia_service = EvidenciaService()

        # ============================================================
        # FALLBACK PROFESIONAL: genera parrafo fluido con templates
        # mientras la IA se procesa en background
        # ============================================================
        descripcion_fallback = evidencia_service._generar_descripcion_actividad(
            reporte=reporte,
            anuncio=anuncio_usuario
        )

        evidencia = evidencia_service.crear_evidencia(
            reporte=reporte,
            imagen=file,
            anuncio=anuncio_usuario,
            fecha=fecha_actividad,
            descripcion=descripcion_fallback
        )

        db.session.commit()

        # Lanzar analisis IA en background (mismo patron que subir_evidencia POST)
        if api_key and evidencia.imagen_path:
            app = current_app._get_current_object()

            thread = threading.Thread(
                target=_analizar_ia_background,
                args=(
                    app,
                    evidencia.id,
                    evidencia.imagen_path,
                    api_key,
                    obligacion.descripcion,
                    anuncio_usuario
                ),
                daemon=True
            )
            thread.start()

        return jsonify({
            'success': True,
            'evidencia_id': evidencia.id,
            'numero_actividad': evidencia.numero_actividad,
            'descripcion_ia': bool(api_key and evidencia.imagen_path),
            'mensaje': 'Evidencia registrada correctamente.'
        })

    except RequestEntityTooLarge:
        db.session.rollback()
        return jsonify({'success': False, 'error': 'Archivo demasiado grande (max 16MB).'}), 413
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
