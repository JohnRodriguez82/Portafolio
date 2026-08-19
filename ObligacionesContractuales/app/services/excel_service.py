"""
Servicio para generación y lectura de archivos Excel.

Responsabilidades:

- Leer archivos Excel de carga masiva.
- Validar encabezados.
- Normalizar encabezados.
- Normalizar obligaciones.
- Normalizar fechas.
- Normalizar textos.
- Generar plantilla Excel para carga masiva.
- Generar plantilla Excel para reportes.

Este módulo NO depende de Flask.

IMPORTANTE
----------
El formato oficial de carga masiva es:

    Obligacion No.
    Descripcion Obligacion
    Anuncio / Contexto
    Fecha de la actividad
    Nombre Imagen

El lector acepta además encabezados antiguos/alternativos para
mantener compatibilidad con archivos Excel que ya existan.
"""

import io

from datetime import date, datetime

from openpyxl import (
    Workbook,
    load_workbook
)

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)


# ============================================================
# CONSTANTES
# ============================================================

EXCEL_MIMETYPE = (
    'application/vnd.openxmlformats-officedocument.'
    'spreadsheetml.sheet'
)


# ============================================================
# FORMATO OFICIAL DE CARGA MASIVA
# ============================================================

ENCABEZADOS_CARGA = [
    'Obligacion No.',
    'Descripcion Obligacion',
    'Anuncio / Contexto',
    'Fecha de la actividad',
    'Nombre Imagen'
]


# ============================================================
# ALIAS DE ENCABEZADOS
# ============================================================
#
# Las claves son nombres internos.
#
# Los valores son todas las variantes que el lector acepta.
#
# Después de _normalizar_encabezado(), por ejemplo:
#
# "Descripción Obligación"
# "DESCRIPCION OBLIGACION"
# "descripcion_obligacion"
#
# terminan siendo comparables.
# ============================================================

ALIASES_COLUMNAS = {

    'obligacion': [
        'Obligacion No.',
        'Obligacion No',
        'Obligacion',
        'Numero Obligacion',
        'Número Obligación',
        'No Obligacion',
        'No. Obligacion',
        'No de Obligacion',
        'No. de Obligacion'
    ],

    'descripcion': [
        'Descripcion Obligacion',
        'Descripción Obligación',
        'Descripcion',
        'Descripción'
    ],

    'anuncio': [
        'Anuncio / Contexto',
        'Anuncio/Contexto',
        'Anuncio',
        'Contexto'
    ],

    'fecha': [
        'Fecha de la actividad',
        'Fecha Actividad',
        'Fecha de Actividad',
        'Fecha'
    ],

    'nombre_imagen': [
        'Nombre Imagen',
        'Nombre de Imagen',
        'Imagen',
        'Nombre Archivo',
        'Nombre del Archivo',
        'Archivo Imagen'
    ]
}


# ============================================================
# CLAVES INTERNAS NORMALIZADAS
# ============================================================

CLAVES_COLUMNAS = {
    clave: [
        ExcelServiceAlias
        for ExcelServiceAlias in variantes
    ]
    for clave, variantes in ALIASES_COLUMNAS.items()
}


class ExcelService:
    """
    Servicio para trabajar con archivos Excel.

    El servicio no depende de Flask y puede utilizarse desde:

    - Blueprints.
    - Servicios.
    - Jobs en segundo plano.
    - Pruebas unitarias.
    """

    # ========================================================
    # LEER EXCEL
    # ========================================================

    def leer_excel(
        self,
        archivo_excel
    ):
        """
        Lee un Excel de carga masiva.

        Parámetros
        ----------
        archivo_excel:
            Puede ser:

            - Flask FileStorage.
            - BytesIO.
            - bytes.
            - Ruta de archivo.

        Retorna
        -------
        list[dict]

        Ejemplo:

        [
            {
                'obligacion': 1,
                'descripcion': 'Descripción...',
                'anuncio': 'Actividad realizada...',
                'fecha': date(2026, 8, 15),
                'nombre_imagen': 'evidencia1.jpg',
                '_fila_excel': 2
            }
        ]

        Notas
        -----
        - Las filas completamente vacías se ignoran.
        - La fecha queda normalizada como datetime.date.
        - El número de obligación queda normalizado.
        - Se conserva la fila original de Excel mediante
          '_fila_excel'.
        """

        if archivo_excel is None:

            raise ValueError(
                'No se recibió el archivo Excel.'
            )

        workbook = self._abrir_workbook(
            archivo_excel
        )

        try:

            worksheet = workbook.active

            filas = list(
                worksheet.iter_rows(
                    values_only=True
                )
            )

            if not filas:

                return []

            encabezados = [
                self._normalizar_encabezado(
                    valor
                )
                for valor in filas[0]
            ]

            mapa_columnas = (
                self._construir_mapa_columnas(
                    encabezados
                )
            )

            self._validar_encabezados(
                encabezados
            )

            resultado = []

            for numero_fila, fila in enumerate(
                filas[1:],
                start=2
            ):

                registro = (
                    self._convertir_fila(
                        fila=fila,
                        numero_fila=numero_fila,
                        encabezados=encabezados,
                        mapa_columnas=mapa_columnas
                    )
                )

                if registro is not None:

                    resultado.append(
                        registro
                    )

            return resultado

        finally:

            workbook.close()

    # ========================================================
    # ABRIR WORKBOOK
    # ========================================================

    @staticmethod
    def _abrir_workbook(
        archivo_excel
    ):
        """
        Abre un archivo Excel recibido como:

        - FileStorage.
        - BytesIO.
        - bytes.
        - ruta.

        data_only=True permite obtener el valor calculado de
        celdas de Excel cuando corresponda.
        """

        # ----------------------------------------------------
        # Flask FileStorage
        # ----------------------------------------------------

        if hasattr(
            archivo_excel,
            'stream'
        ):

            archivo_excel.stream.seek(0)

            return load_workbook(
                filename=archivo_excel.stream,
                data_only=True
            )

        # ----------------------------------------------------
        # Objetos con seek(), por ejemplo BytesIO
        # ----------------------------------------------------

        if hasattr(
            archivo_excel,
            'seek'
        ):

            archivo_excel.seek(0)

            return load_workbook(
                filename=archivo_excel,
                data_only=True
            )

        # ----------------------------------------------------
        # bytes o ruta
        # ----------------------------------------------------

        return load_workbook(
            filename=archivo_excel,
            data_only=True
        )

    # ========================================================
    # CONSTRUIR MAPA DE COLUMNAS
    # ========================================================

    @classmethod
    def _construir_mapa_columnas(
        cls,
        encabezados
    ):
        """
        Construye un mapa:

            {
                'obligacion': indice,
                'descripcion': indice,
                'anuncio': indice,
                'fecha': indice,
                'nombre_imagen': indice
            }

        Los índices corresponden a las posiciones de las
        columnas dentro de la hoja Excel.

        Ejemplo:

            A -> 0
            B -> 1
            C -> 2
            D -> 3
            E -> 4
        """

        mapa = {}

        # ----------------------------------------------------
        # Convertir aliases a valores normalizados
        # ----------------------------------------------------

        aliases_normalizados = {}

        for clave, variantes in ALIASES_COLUMNAS.items():

            aliases_normalizados[clave] = {
                cls._normalizar_encabezado(
                    variante
                )
                for variante in variantes
            }

        # ----------------------------------------------------
        # Buscar cada encabezado
        # ----------------------------------------------------

        for indice, encabezado in enumerate(
            encabezados
        ):

            for clave, aliases in (
                aliases_normalizados.items()
            ):

                if encabezado in aliases:

                    # Si existe duplicado, conservar la primera
                    # aparición para evitar comportamiento ambiguo.

                    if clave not in mapa:

                        mapa[clave] = indice

                    break

        return mapa

    # ========================================================
    # VALIDAR ENCABEZADOS
    # ========================================================

    @classmethod
    def _validar_encabezados(
        cls,
        encabezados
    ):
        """
        Valida que el Excel tenga las cinco columnas requeridas.

        La validación NO exige que estén en A/B/C/D/E exactamente.

        Esto permite que un usuario mueva columnas dentro del Excel
        sin romper la lectura.

        Sí exige que existan todos los campos necesarios.
        """

        mapa = cls._construir_mapa_columnas(
            encabezados
        )

        requeridos = [
            'obligacion',
            'descripcion',
            'anuncio',
            'fecha',
            'nombre_imagen'
        ]

        faltantes = [
            clave
            for clave in requeridos
            if clave not in mapa
        ]

        if faltantes:

            nombres_faltantes = []

            nombres_legibles = {
                'obligacion': 'Obligacion No.',
                'descripcion': 'Descripcion Obligacion',
                'anuncio': 'Anuncio / Contexto',
                'fecha': 'Fecha de la actividad',
                'nombre_imagen': 'Nombre Imagen'
            }

            for faltante in faltantes:

                nombres_faltantes.append(
                    nombres_legibles.get(
                        faltante,
                        faltante
                    )
                )

            raise ValueError(
                'El archivo Excel no contiene '
                'los encabezados requeridos. '
                f'Faltan: {", ".join(nombres_faltantes)}'
            )

    # ========================================================
    # CONVERTIR FILA
    # ========================================================

    @classmethod
    def _convertir_fila(
        cls,
        fila,
        numero_fila,
        encabezados,
        mapa_columnas=None
    ):
        """
        Convierte una fila del Excel en un diccionario.

        El método utiliza el mapa de columnas y NO depende de
        que la obligación esté necesariamente en A, la descripción
        en B, etc.

        Esto hace el lector mucho más resistente a cambios del
        usuario en la hoja.
        """

        if mapa_columnas is None:

            mapa_columnas = (
                cls._construir_mapa_columnas(
                    encabezados
                )
            )

        # ----------------------------------------------------
        # Obtener valores por nombre lógico
        # ----------------------------------------------------

        obligacion = cls._valor_columna(
            fila,
            mapa_columnas,
            'obligacion'
        )

        descripcion = cls._valor_columna(
            fila,
            mapa_columnas,
            'descripcion'
        )

        anuncio = cls._valor_columna(
            fila,
            mapa_columnas,
            'anuncio'
        )

        fecha = cls._valor_columna(
            fila,
            mapa_columnas,
            'fecha'
        )

        nombre_imagen = cls._valor_columna(
            fila,
            mapa_columnas,
            'nombre_imagen'
        )

        # ----------------------------------------------------
        # Ignorar filas completamente vacías
        # ----------------------------------------------------

        valores = [
            obligacion,
            descripcion,
            anuncio,
            fecha,
            nombre_imagen
        ]

        if all(
            cls._esta_vacio(valor)
            for valor in valores
        ):

            return None

        # ----------------------------------------------------
        # Normalización
        # ----------------------------------------------------

        return {
            'obligacion': (
                cls._normalizar_obligacion(
                    obligacion
                )
            ),

            'descripcion': (
                cls._texto(
                    descripcion
                )
            ),

            'anuncio': (
                cls._texto(
                    anuncio
                )
            ),

            'fecha': (
                cls._normalizar_fecha(
                    fecha
                )
            ),

            'nombre_imagen': (
                cls._texto(
                    nombre_imagen
                )
            ),

            '_fila_excel': numero_fila
        }

    # ========================================================
    # OBTENER VALOR DE COLUMNA
    # ========================================================

    @staticmethod
    def _valor_columna(
        fila,
        mapa_columnas,
        clave
    ):
        """
        Obtiene un valor de una fila mediante la clave lógica.

        Si la columna no existe, retorna None.

        Se mantiene como método separado para que el flujo de
        lectura sea claro y fácilmente comprobable mediante tests.
        """

        indice = mapa_columnas.get(
            clave
        )

        if indice is None:

            return None

        if indice >= len(fila):

            return None

        return fila[indice]

    # ========================================================
    # NORMALIZAR OBLIGACIÓN
    # ========================================================

    @staticmethod
    def _normalizar_obligacion(
        valor
    ):
        """
        Convierte el número de obligación a un formato manejable.

        Acepta:

            1
            1.0
            '1'
            '1.0'
            '1,0'
            ' 1 '

        Retorna:

            1

        Si el valor no representa un número entero, se conserva
        como texto para que el servicio superior pueda decidir
        cómo manejarlo.
        """

        if valor is None:

            return None

        # ----------------------------------------------------
        # int
        # ----------------------------------------------------

        if isinstance(
            valor,
            int
        ):

            return valor

        # ----------------------------------------------------
        # float
        # ----------------------------------------------------

        if isinstance(
            valor,
            float
        ):

            if valor.is_integer():

                return int(
                    valor
                )

            return valor

        # ----------------------------------------------------
        # Texto
        # ----------------------------------------------------

        texto = str(
            valor
        ).strip()

        if not texto:

            return None

        # ----------------------------------------------------
        # Intentar convertir números
        # ----------------------------------------------------

        texto_numerico = (
            texto
            .replace(
                ',',
                '.'
            )
        )

        try:

            numero = float(
                texto_numerico
            )

            if numero.is_integer():

                return int(
                    numero
                )

        except (
            ValueError,
            TypeError
        ):

            pass

        # ----------------------------------------------------
        # Si no es numérico, conservar texto
        # ----------------------------------------------------

        return texto

    # ========================================================
    # NORMALIZAR FECHA
    # ========================================================

    @staticmethod
    def _normalizar_fecha(
        valor
    ):
        """
        Convierte fechas de Excel a datetime.date.

        Acepta:

        - datetime
        - date
        - YYYY-MM-DD
        - DD/MM/YYYY
        - DD-MM-YYYY
        - YYYY/MM/DD

        IMPORTANTE:
        Excel normalmente entrega las celdas con formato fecha
        como datetime/date. Por eso se comprueba primero el tipo
        Python antes de intentar interpretar texto.
        """

        if valor is None:

            return None

        # ----------------------------------------------------
        # datetime
        # ----------------------------------------------------

        if isinstance(
            valor,
            datetime
        ):

            return valor.date()

        # ----------------------------------------------------
        # date
        # ----------------------------------------------------

        if isinstance(
            valor,
            date
        ):

            return valor

        # ----------------------------------------------------
        # Texto
        # ----------------------------------------------------

        texto = str(
            valor
        ).strip()

        if not texto:

            return None

        formatos = [
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%d-%m-%Y',
            '%Y/%m/%d'
        ]

        for formato in formatos:

            try:

                return datetime.strptime(
                    texto,
                    formato
                ).date()

            except ValueError:

                continue

        raise ValueError(
            f'Fecha no válida: {valor}'
        )

    # ========================================================
    # NORMALIZAR TEXTO
    # ========================================================

    @staticmethod
    def _texto(
        valor
    ):
        """
        Convierte un valor a texto limpio.

        None -> ''

        Los demás valores se convierten a str y se eliminan
        espacios al principio y al final.
        """

        if valor is None:

            return ''

        return str(
            valor
        ).strip()

    # ========================================================
    # VALIDAR VACÍO
    # ========================================================

    @staticmethod
    def _esta_vacio(
        valor
    ):
        """
        Determina si un valor está vacío.

        Considera vacíos:

            None
            ''
            '   '
        """

        if valor is None:

            return True

        if isinstance(
            valor,
            str
        ):

            return not valor.strip()

        return False

    # ========================================================
    # NORMALIZAR ENCABEZADO
    # ========================================================

    @staticmethod
    def _normalizar_encabezado(
        valor
    ):
        """
        Normaliza nombres de columnas.

        Permite tolerar:

        - Mayúsculas/minúsculas.
        - Tildes.
        - Espacios múltiples.
        - Guiones.
        - Guiones bajos.

        Ejemplos:

            'Descripción Obligación'
                ->
            'descripcion obligacion'

            'DESCRIPCION_OBLIGACION'
                ->
            'descripcion obligacion'

            'Anuncio / Contexto'
                ->
            'anuncio / contexto'
        """

        if valor is None:

            return ''

        texto = str(
            valor
        ).strip().lower()

        reemplazos = {
            'á': 'a',
            'é': 'e',
            'í': 'i',
            'ó': 'o',
            'ú': 'u',
            'ü': 'u'
        }

        for origen, destino in (
            reemplazos.items()
        ):

            texto = texto.replace(
                origen,
                destino
            )

        texto = (
            texto
            .replace(
                '_',
                ' '
            )
            .replace(
                '-',
                ' '
            )
        )

        return ' '.join(
            texto.split()
        )

    # ========================================================
    # GENERAR PLANTILLA MASIVA
    # ========================================================

    @staticmethod
    def generar_plantilla_masiva(
        obligaciones,
        mes,
        anio
    ):
        """
        Genera la plantilla oficial para carga masiva mensual.

        Columnas:

            A - Obligacion No.
            B - Descripcion Obligacion
            C - Anuncio / Contexto
            D - Fecha de la actividad
            E - Nombre Imagen

        Las columnas A y B son informativas y se generan a partir
        de las obligaciones recibidas.

        C, D y E quedan disponibles para diligenciar.
        """

        wb = Workbook()

        ws = wb.active

        ws.title = (
            f'Carga_{int(mes):02d}_{anio}'
        )

        # ----------------------------------------------------
        # Estilos
        # ----------------------------------------------------

        encabezado_fill = PatternFill(
            'solid',
            fgColor='0D6EFD'
        )

        encabezado_font = Font(
            bold=True,
            color='FFFFFF'
        )

        encabezado_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

        borde = Border(
            bottom=Side(
                style='thin',
                color='CCCCCC'
            )
        )

        # ----------------------------------------------------
        # Encabezados
        # ----------------------------------------------------

        for columna, encabezado in enumerate(
            ENCABEZADOS_CARGA,
            start=1
        ):

            celda = ws.cell(
                row=1,
                column=columna,
                value=encabezado
            )

            celda.fill = (
                encabezado_fill
            )

            celda.font = (
                encabezado_font
            )

            celda.alignment = (
                encabezado_alignment
            )

            celda.border = (
                borde
            )

        # ----------------------------------------------------
        # Obligaciones
        # ----------------------------------------------------

        fila = 2

        for obligacion in (
            obligaciones or []
        ):

            numero = getattr(
                obligacion,
                'numero',
                ''
            )

            descripcion = getattr(
                obligacion,
                'descripcion',
                ''
            )

            # ------------------------------------------------
            # Columna A
            # ------------------------------------------------

            ws.cell(
                row=fila,
                column=1,
                value=numero
            )

            # ------------------------------------------------
            # Columna B
            # ------------------------------------------------

            ws.cell(
                row=fila,
                column=2,
                value=descripcion
            )

            # ------------------------------------------------
            # Columnas C/D/E
            #
            # Se dejan vacías deliberadamente.
            # ------------------------------------------------

            fila += 1

        # ----------------------------------------------------
        # Anchos
        # ----------------------------------------------------

        ws.column_dimensions[
            'A'
        ].width = 18

        ws.column_dimensions[
            'B'
        ].width = 70

        ws.column_dimensions[
            'C'
        ].width = 55

        ws.column_dimensions[
            'D'
        ].width = 22

        ws.column_dimensions[
            'E'
        ].width = 35

        # ----------------------------------------------------
        # Filtro
        # ----------------------------------------------------

        ws.auto_filter.ref = (
            f'A1:E{max(fila - 1, 1)}'
        )

        # ----------------------------------------------------
        # Congelar encabezados
        # ----------------------------------------------------

        ws.freeze_panes = 'A2'

        # ----------------------------------------------------
        # Altura encabezado
        # ----------------------------------------------------

        ws.row_dimensions[
            1
        ].height = 32

        # ====================================================
        # HOJA DE INSTRUCCIONES
        # ====================================================

        ws_instr = wb.create_sheet(
            'Instrucciones'
        )

        instrucciones = [

            [
                'INSTRUCCIONES DE CARGA MASIVA POR MES'
            ],

            [''],

            [
                'FORMATO OFICIAL DEL ARCHIVO'
            ],

            [
                'A: Obligacion No.'
            ],

            [
                'B: Descripcion Obligacion'
            ],

            [
                'C: Anuncio / Contexto'
            ],

            [
                'D: Fecha de la actividad'
            ],

            [
                'E: Nombre Imagen'
            ],

            [''],

            [
                '1. NO modifique los encabezados de la fila 1.'
            ],

            [
                '2. Las columnas A y B son generadas por el sistema '
                'y no deben modificarse.'
            ],

            [
                '3. En la columna C escriba el anuncio o contexto '
                'de la actividad.'
            ],

            [
                '4. En la columna D indique la fecha de la '
                'actividad.'
            ],

            [
                '5. Se aceptan fechas YYYY-MM-DD, DD/MM/YYYY '
                'y DD-MM-YYYY.'
            ],

            [
                '6. En la columna E escriba el nombre EXACTO '
                'del archivo de imagen, incluyendo extension.'
            ],

            [
                '7. Ejemplo de imagen: evidencia1.jpg'
            ],

            [
                '8. Puede insertar varias filas para una misma '
                'obligacion cuando tenga multiples evidencias.'
            ],

            [
                '9. Puede eliminar filas de obligaciones que '
                'no tengan actividades durante el periodo.'
            ],

            [
                '10. Las imagenes deben cargarse junto con el '
                'archivo Excel en el formulario web.'
            ],

            [''],

            [
                'REGLAS IMPORTANTES'
            ],

            [
                '- La fecha debe pertenecer al mes y año '
                'seleccionados.'
            ],

            [
                '- El nombre de imagen en el Excel debe '
                'coincidir con el archivo cargado.'
            ],

            [
                '- Si no desea adjuntar imagen, deje la columna E '
                'vacia.'
            ],

            [
                '- Una fila sin imagen puede crear una actividad '
                'sin evidencia visual.'
            ],

            [
                '- Puede existir mas de una actividad para '
                'una misma obligacion.'
            ],

            [
                '- El sistema conserva la numeracion de actividades '
                'existentes.'
            ],

            [
                '- Si Gemini esta configurado, las imagenes '
                'pueden ser analizadas automaticamente.'
            ]
        ]

        for fila_instruccion in instrucciones:

            ws_instr.append(
                fila_instruccion
            )

        ws_instr.column_dimensions[
            'A'
        ].width = 110

        # ----------------------------------------------------
        # Estilo de títulos de instrucciones
        # ----------------------------------------------------

        ws_instr['A1'].font = Font(
            bold=True,
            size=14,
            color='FFFFFF'
        )

        ws_instr['A1'].fill = PatternFill(
            'solid',
            fgColor='0D6EFD'
        )

        ws_instr['A3'].font = Font(
            bold=True,
            size=12
        )

        ws_instr['A21'].font = Font(
            bold=True,
            size=12
        )

        # ----------------------------------------------------
        # Wrap
        # ----------------------------------------------------

        for fila_instr in ws_instr.iter_rows():

            for celda in fila_instr:

                celda.alignment = Alignment(
                    vertical='top',
                    wrap_text=True
                )

        # ----------------------------------------------------
        # Archivo en memoria
        # ----------------------------------------------------

        output = io.BytesIO()

        wb.save(
            output
        )

        output.seek(0)

        return output

    # ========================================================
    # GENERAR PLANTILLA REPORTE
    # ========================================================

    @staticmethod
    def generar_plantilla_reporte():
        """
        Genera una plantilla Excel básica para registrar
        actividades de un reporte específico.

        Esta plantilla corresponde al flujo individual,
        no al flujo de carga masiva mensual.
        """

        wb = Workbook()

        ws = wb.active

        ws.title = 'Carga Masiva'

        encabezados = [
            'Anuncio',
            'Fecha',
            'Nombre Imagen'
        ]

        # ----------------------------------------------------
        # Encabezados
        # ----------------------------------------------------

        for columna, encabezado in enumerate(
            encabezados,
            start=1
        ):

            celda = ws.cell(
                row=1,
                column=columna,
                value=encabezado
            )

            celda.font = Font(
                bold=True
            )

            celda.fill = PatternFill(
                'solid',
                fgColor='0D6EFD'
            )

            celda.font = Font(
                bold=True,
                color='FFFFFF'
            )

            celda.alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

        # ----------------------------------------------------
        # Anchos
        # ----------------------------------------------------

        ws.column_dimensions[
            'A'
        ].width = 60

        ws.column_dimensions[
            'B'
        ].width = 20

        ws.column_dimensions[
            'C'
        ].width = 35

        # ----------------------------------------------------
        # Congelar encabezado
        # ----------------------------------------------------

        ws.freeze_panes = 'A2'

        # ----------------------------------------------------
        # Archivo en memoria
        # ----------------------------------------------------

        output = io.BytesIO()

        wb.save(
            output
        )

        output.seek(0)

        return output


# ============================================================
# FUNCIONES DE COMPATIBILIDAD
# ============================================================

def leer_excel(
    archivo_excel
):
    """
    Función de compatibilidad.

    Permite utilizar:

        leer_excel(archivo)

    además de:

        ExcelService().leer_excel(archivo)
    """

    return ExcelService().leer_excel(
        archivo_excel
    )


def generar_plantilla_masiva(
    obligaciones,
    mes,
    anio
):
    """
    Función de compatibilidad para generación de plantilla masiva.
    """

    return ExcelService.generar_plantilla_masiva(
        obligaciones,
        mes,
        anio
    )


def generar_plantilla_reporte():
    """
    Función de compatibilidad para generación de plantilla
    de reporte individual.
    """

    return ExcelService.generar_plantilla_reporte()
