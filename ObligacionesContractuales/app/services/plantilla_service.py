"""
Servicio para generación de plantillas Excel.

Responsabilidades:
- Construir plantilla de carga masiva mensual.
- Agregar obligaciones del contrato.
- Definir encabezados compatibles con ExcelService.
- Aplicar formato al archivo Excel.
- Agregar instrucciones de uso.
- Prellenar la fecha con el día 15 del mes seleccionado.
- Retornar el workbook listo para descarga.

IMPORTANTE:
Este módulo NO procesa la carga.
Únicamente genera archivos Excel compatibles con:
    app.services.excel_service.ExcelService
    app.blueprints.cargas
"""

from io import BytesIO
from datetime import date
import calendar

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter


# ============================================================
# CONSTANTES
# ============================================================

ENCABEZADOS_CARGA_MASIVA = [
    'Obligacion No.',
    'Descripcion Obligacion',
    'Anuncio / Contexto',
    'Fecha de la actividad',
    'Nombre Imagen',
]

# Alias de compatibilidad.
#
# Algunos módulos/procesos antiguos pueden importar ENCABEZADOS.
# Se conserva para evitar errores de importación.
ENCABEZADOS = ENCABEZADOS_CARGA_MASIVA


# ============================================================
# ESTILOS
# ============================================================

COLOR_ENCABEZADO = '1F4E78'
COLOR_ENCABEZADO_SECUNDARIO = '0D6EFD'
COLOR_BORDE = 'B7C9D6'
COLOR_FONDO_INSTRUCCIONES = 'EAF2F8'
COLOR_ADVERTENCIA = 'FFF2CC'

FUENTE_ENCABEZADO = Font(
    bold=True,
    color='FFFFFF',
    size=11,
)

FUENTE_TITULO = Font(
    bold=True,
    color='FFFFFF',
    size=14,
)

FUENTE_SUBTITULO = Font(
    bold=True,
    color='1F1F1F',
    size=11,
)

FUENTE_NORMAL = Font(
    color='1F1F1F',
    size=10,
)

FUENTE_ADVERTENCIA = Font(
    bold=True,
    color='7F6000',
    size=10,
)

RELLENO_ENCABEZADO = PatternFill(
    fill_type='solid',
    fgColor=COLOR_ENCABEZADO,
)

RELLENO_TITULO = PatternFill(
    fill_type='solid',
    fgColor=COLOR_ENCABEZADO,
)

RELLENO_INSTRUCCIONES = PatternFill(
    fill_type='solid',
    fgColor=COLOR_FONDO_INSTRUCCIONES,
)

RELLENO_ADVERTENCIA = PatternFill(
    fill_type='solid',
    fgColor=COLOR_ADVERTENCIA,
)

BORDE_CELDA = Border(
    left=Side(
        style='thin',
        color=COLOR_BORDE,
    ),
    right=Side(
        style='thin',
        color=COLOR_BORDE,
    ),
    top=Side(
        style='thin',
        color=COLOR_BORDE,
    ),
    bottom=Side(
        style='thin',
        color=COLOR_BORDE,
    ),
)

BORDE_ENCABEZADO = Border(
    left=Side(
        style='thin',
        color='FFFFFF',
    ),
    right=Side(
        style='thin',
        color='FFFFFF',
    ),
    top=Side(
        style='thin',
        color='FFFFFF',
    ),
    bottom=Side(
        style='thin',
        color='FFFFFF',
    ),
)


# ============================================================
# CREAR PLANTILLA
# ============================================================

def crear_plantilla(
    obligaciones,
    mes,
    anio,
):
    """
    Genera un workbook Excel para carga masiva mensual.

    Args:
        obligaciones:
            Lista de objetos Obligacion.

        mes:
            Mes numérico, 1 a 12.

        anio:
            Año del reporte.

    Returns:
        BytesIO:
            Archivo Excel en memoria.

    Raises:
        ValueError:
            Si mes o año no son válidos.
    """

    mes = _validar_mes(mes)
    anio = _validar_anio(anio)

    obligaciones = obligaciones or []

    wb = Workbook()

    ws = wb.active

    ws.title = (
        f'Carga_{mes:02d}_{anio}'
    )

    # --------------------------------------------------------
    # CONFIGURACIÓN DE HOJA
    # --------------------------------------------------------

    ws.sheet_view.showGridLines = False

    ws.freeze_panes = 'A2'

    # --------------------------------------------------------
    # ENCABEZADOS
    # --------------------------------------------------------

    for columna, encabezado in enumerate(
        ENCABEZADOS_CARGA_MASIVA,
        start=1,
    ):
        celda = ws.cell(
            row=1,
            column=columna,
            value=encabezado,
        )

        _estilizar_encabezado(
            celda
        )

    ws.row_dimensions[1].height = 32

    # --------------------------------------------------------
    # OBLIGACIONES
    # --------------------------------------------------------

    fila = 2

    for obligacion in obligaciones:

        numero = getattr(
            obligacion,
            'numero',
            '',
        )

        descripcion = getattr(
            obligacion,
            'descripcion',
            '',
        )

        # A - Número de obligación
        celda_numero = ws.cell(
            row=fila,
            column=1,
            value=numero,
        )

        # B - Descripción
        celda_descripcion = ws.cell(
            row=fila,
            column=2,
            value=descripcion,
        )

        # C - Anuncio / Contexto
        ws.cell(
            row=fila,
            column=3,
            value='',
        )

        # ----------------------------------------------------
        # D - Fecha de actividad
        #
        # Se coloca una fecha REAL de Excel, no texto.
        # Esto permite que openpyxl la recupere como date.
        # ----------------------------------------------------

        fecha_predeterminada = date(
            anio,
            mes,
            15,
        )

        celda_fecha = ws.cell(
            row=fila,
            column=4,
            value=fecha_predeterminada,
        )

        celda_fecha.number_format = 'yyyy-mm-dd'

        # E - Nombre de imagen
        ws.cell(
            row=fila,
            column=5,
            value='',
        )

        # ----------------------------------------------------
        # FORMATO DE FILA
        # ----------------------------------------------------

        for columna in range(
            1,
            6,
        ):
            celda = ws.cell(
                row=fila,
                column=columna,
            )

            celda.border = BORDE_CELDA

            celda.alignment = Alignment(
                vertical='top',
                wrap_text=True,
            )

            if columna == 1:
                celda_numero.alignment = Alignment(
                    horizontal='center',
                    vertical='top',
                )

            if columna == 2:
                celda_descripcion.alignment = Alignment(
                    vertical='top',
                    wrap_text=True,
                )

            if columna == 3:
                celda.alignment = Alignment(
                    vertical='top',
                    wrap_text=True,
                )

            if columna == 4:
                celda_fecha.alignment = Alignment(
                    horizontal='center',
                    vertical='top',
                )

            if columna == 5:
                celda.alignment = Alignment(
                    vertical='top',
                    wrap_text=True,
                )

        fila += 1

    # --------------------------------------------------------
    # FORMATO DE COLUMNAS
    # --------------------------------------------------------

    _ajustar_columnas(
        ws
    )

    # --------------------------------------------------------
    # FILTRO
    # --------------------------------------------------------

    if ws.max_row >= 1:
        ws.auto_filter.ref = (
            f'A1:E{max(ws.max_row, 2)}'
        )

    # --------------------------------------------------------
    # HOJA DE INSTRUCCIONES
    # --------------------------------------------------------

    _crear_hoja_instrucciones(
        wb,
        mes,
        anio,
    )

    # --------------------------------------------------------
    # GUARDAR EN MEMORIA
    # --------------------------------------------------------

    output = BytesIO()

    wb.save(
        output
    )

    output.seek(0)

    return output


# ============================================================
# HOJA DE INSTRUCCIONES
# ============================================================

def _crear_hoja_instrucciones(
    wb,
    mes,
    anio,
):
    """
    Crea la hoja de instrucciones para el usuario.
    """

    ws = wb.create_sheet(
        'Instrucciones'
    )

    ws.sheet_view.showGridLines = False

    # --------------------------------------------------------
    # TÍTULO
    # --------------------------------------------------------

    ws.merge_cells(
        'A1:E1'
    )

    titulo = ws['A1']

    titulo.value = (
        'INSTRUCCIONES DE CARGA MASIVA POR MES'
    )

    titulo.font = FUENTE_TITULO

    titulo.fill = RELLENO_TITULO

    titulo.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )

    titulo.border = BORDE_ENCABEZADO

    ws.row_dimensions[1].height = 30

    # --------------------------------------------------------
    # INFORMACIÓN DEL PERIODO
    # --------------------------------------------------------

    ws.merge_cells(
        'A3:E3'
    )

    periodo = ws['A3']

    periodo.value = (
        f'Periodo seleccionado: '
        f'{mes:02d}/{anio}'
    )

    periodo.font = Font(
        bold=True,
        color='1F1F1F',
        size=11,
    )

    periodo.fill = RELLENO_INSTRUCCIONES

    periodo.alignment = Alignment(
        horizontal='left',
        vertical='center',
    )

    # --------------------------------------------------------
    # INSTRUCCIONES
    # --------------------------------------------------------

    instrucciones = [
        (
            '1.',
            'NO modifique los encabezados '
            'de la fila 1.'
        ),
        (
            '2.',
            'NO modifique las columnas A y B. '
            'Estas contienen el número y la '
            'descripción de la obligación.'
        ),
        (
            '3.',
            'En la columna C escriba el anuncio '
            'o contexto de la actividad.'
        ),
        (
            '4.',
            'En la columna D indique la fecha '
            'real de la actividad.'
        ),
        (
            '5.',
            'La fecha debe pertenecer al mes '
            'y año seleccionados.'
        ),
        (
            '6.',
            'La columna E debe contener el nombre '
            'EXACTO del archivo de imagen, '
            'incluyendo su extensión.'
        ),
        (
            '7.',
            'Ejemplo de nombre de imagen: '
            'evidencia1.jpg'
        ),
        (
            '8.',
            'Puede insertar varias filas para '
            'una misma obligación si existen '
            'varias actividades o evidencias.'
        ),
        (
            '9.',
            'Puede eliminar las obligaciones '
            'que no tengan actividades durante '
            'el periodo.'
        ),
        (
            '10.',
            'Las imágenes deben cargarse junto '
            'con este archivo Excel en el campo '
            'de archivos múltiples del formulario.'
        ),
    ]

    fila = 5

    for numero, texto in instrucciones:

        ws.cell(
            row=fila,
            column=1,
            value=numero,
        )

        ws.cell(
            row=fila,
            column=2,
            value=texto,
        )

        ws.merge_cells(
            start_row=fila,
            start_column=2,
            end_row=fila,
            end_column=5,
        )

        numero_celda = ws.cell(
            row=fila,
            column=1,
        )

        texto_celda = ws.cell(
            row=fila,
            column=2,
        )

        numero_celda.font = Font(
            bold=True,
            color=COLOR_ENCABEZADO,
        )

        texto_celda.font = FUENTE_NORMAL

        numero_celda.alignment = Alignment(
            horizontal='center',
            vertical='top',
        )

        texto_celda.alignment = Alignment(
            vertical='top',
            wrap_text=True,
        )

        numero_celda.border = BORDE_CELDA

        texto_celda.border = BORDE_CELDA

        ws.row_dimensions[fila].height = 30

        fila += 1

    # --------------------------------------------------------
    # REGLAS IMPORTANTES
    # --------------------------------------------------------

    fila += 1

    ws.merge_cells(
        start_row=fila,
        start_column=1,
        end_row=fila,
        end_column=5,
    )

    titulo_reglas = ws.cell(
        row=fila,
        column=1,
        value='REGLAS IMPORTANTES',
    )

    titulo_reglas.font = FUENTE_SUBTITULO

    titulo_reglas.fill = RELLENO_ADVERTENCIA

    titulo_reglas.alignment = Alignment(
        vertical='center',
        wrap_text=True,
    )

    titulo_reglas.border = BORDE_CELDA

    fila += 1

    reglas = [
        (
            'La fecha debe pertenecer al mes '
            'y año seleccionados.'
        ),
        (
            'El nombre de la imagen en el Excel '
            'debe coincidir exactamente con el '
            'archivo que se adjunta.'
        ),
        (
            'Si no se adjunta imagen, la columna E '
            'puede dejarse vacía. La actividad se '
            'creará sin evidencia visual.'
        ),
        (
            'Si se utiliza Gemini, las imágenes '
            'pueden ser analizadas automáticamente '
            'por el sistema.'
        ),
        (
            'El sistema genera o reutiliza el '
            'ReporteMensual correspondiente a '
            'cada obligación.'
        ),
    ]

    for regla in reglas:

        ws.merge_cells(
            start_row=fila,
            start_column=1,
            end_row=fila,
            end_column=5,
        )

        celda = ws.cell(
            row=fila,
            column=1,
            value=f'• {regla}',
        )

        celda.font = FUENTE_ADVERTENCIA

        celda.fill = RELLENO_ADVERTENCIA

        celda.alignment = Alignment(
            vertical='top',
            wrap_text=True,
        )

        celda.border = BORDE_CELDA

        ws.row_dimensions[fila].height = 32

        fila += 1

    # --------------------------------------------------------
    # EJEMPLO
    # --------------------------------------------------------

    fila += 1

    ws.merge_cells(
        start_row=fila,
        start_column=1,
        end_row=fila,
        end_column=5,
    )

    ejemplo_titulo = ws.cell(
        row=fila,
        column=1,
        value='EJEMPLO DE REGISTRO',
    )

    ejemplo_titulo.font = FUENTE_SUBTITULO

    ejemplo_titulo.fill = RELLENO_INSTRUCCIONES

    ejemplo_titulo.border = BORDE_CELDA

    fila += 1

    ejemplos = [
        (
            'Obligacion No.',
            'Descripcion Obligacion',
            'Anuncio / Contexto',
            'Fecha de la actividad',
            'Nombre Imagen',
        ),
        (
            '1',
            'Ejemplo de obligación',
            'Actividad realizada durante el periodo',
            f'{anio}-{mes:02d}-15',
            'evidencia1.jpg',
        ),
    ]

    for fila_ejemplo in ejemplos:

        for columna, valor in enumerate(
            fila_ejemplo,
            start=1,
        ):
            celda = ws.cell(
                row=fila,
                column=columna,
                value=valor,
            )

            celda.border = BORDE_CELDA

            celda.alignment = Alignment(
                vertical='top',
                wrap_text=True,
            )

            if fila == (
                fila
                # Esta condición se reemplaza abajo
            ):
                pass

        fila += 1

    # --------------------------------------------------------
    # CORREGIR ESTILO DE ENCABEZADO DEL EJEMPLO
    # --------------------------------------------------------

    fila_encabezado_ejemplo = fila - 2

    for columna in range(
        1,
        6,
    ):
        celda = ws.cell(
            row=fila_encabezado_ejemplo,
            column=columna,
        )

        celda.font = Font(
            bold=True,
            color='FFFFFF',
            size=9,
        )

        celda.fill = RELLENO_ENCABEZADO

        celda.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
        )

    # --------------------------------------------------------
    # ANCHOS
    # --------------------------------------------------------

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 28

    ws.freeze_panes = 'A5'


# ============================================================
# ESTILO DE ENCABEZADO
# ============================================================

def _estilizar_encabezado(
    cell,
):
    """
    Aplica formato al encabezado de una columna.
    """

    cell.font = FUENTE_ENCABEZADO

    cell.fill = RELLENO_ENCABEZADO

    cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )

    cell.border = BORDE_ENCABEZADO


# ============================================================
# AJUSTAR COLUMNAS
# ============================================================

def _ajustar_columnas(
    ws,
):
    """
    Define anchos adecuados para las cinco columnas.
    """

    anchos = {
        'A': 18,
        'B': 70,
        'C': 55,
        'D': 22,
        'E': 35,
    }

    for columna, ancho in anchos.items():

        ws.column_dimensions[
            columna
        ].width = ancho


# ============================================================
# VALIDAR MES
# ============================================================

def _validar_mes(
    mes,
):
    """
    Valida y convierte el mes a entero.
    """

    try:
        mes = int(
            mes
        )
    except (
        TypeError,
        ValueError,
    ) as exc:

        raise ValueError(
            f'Mes inválido: {mes}'
        ) from exc

    if mes < 1 or mes > 12:

        raise ValueError(
            f'El mes debe estar entre 1 y 12. '
            f'Recibido: {mes}'
        )

    return mes


# ============================================================
# VALIDAR AÑO
# ============================================================

def _validar_anio(
    anio,
):
    """
    Valida y convierte el año a entero.
    """

    try:
        anio = int(
            anio
        )
    except (
        TypeError,
        ValueError,
    ) as exc:

        raise ValueError(
            f'Año inválido: {anio}'
        ) from exc

    if anio < 1900 or anio > 9999:

        raise ValueError(
            f'El año está fuera de rango: {anio}'
        )

    return anio


# ============================================================
# NOMBRE DE ARCHIVO
# ============================================================

def nombre_archivo_plantilla(
    mes,
    anio,
):
    """
    Genera el nombre estándar de descarga.
    """

    mes = _validar_mes(
        mes
    )

    anio = _validar_anio(
        anio
    )

    return (
        f'Plantilla_CargaMasiva_'
        f'{mes:02d}_{anio}.xlsx'
    )


# ============================================================
# ALIAS DE COMPATIBILIDAD
# ============================================================

def generar_plantilla_masiva(
    obligaciones,
    mes,
    anio,
):
    """
    Alias compatible para código que utilice
    generar_plantilla_masiva().
    """

    return crear_plantilla(
        obligaciones=obligaciones,
        mes=mes,
        anio=anio,
    )


def generar_plantilla(
    obligaciones,
    mes,
    anio,
):
    """
    Alias genérico para generación de plantilla.
    """

    return crear_plantilla(
        obligaciones=obligaciones,
        mes=mes,
        anio=anio,
    )


# ============================================================
# INFORMACIÓN DE LA PLANTILLA
# ============================================================

def obtener_encabezados():
    """
    Retorna una copia de los encabezados oficiales.

    Se retorna una copia para evitar que otro módulo
    modifique accidentalmente la constante global.
    """

    return list(
        ENCABEZADOS_CARGA_MASIVA
    )


# ============================================================
# EXPORTS
# ============================================================

__all__ = [
    'ENCABEZADOS',
    'ENCABEZADOS_CARGA_MASIVA',
    'crear_plantilla',
    'generar_plantilla',
    'generar_plantilla_masiva',
    'nombre_archivo_plantilla',
    'obtener_encabezados',
]
