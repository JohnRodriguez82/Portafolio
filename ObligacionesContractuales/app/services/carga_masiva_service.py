"""
Servicio para procesamiento de cargas masivas mensuales.

Responsabilidades:
- Procesar una carga masiva mensual.
- Validar las filas del Excel.
- Crear o reutilizar reportes mensuales.
- Crear evidencias.
- Procesar imágenes y Gemini.
- Actualizar el progreso.
- Limpiar archivos temporales.

Este módulo no define rutas Flask.
"""

import os
import calendar
import time
import threading

from datetime import datetime, date

from flask import current_app

from werkzeug.utils import secure_filename

from openpyxl import load_workbook

from models import (
    db,
    Obligacion,
    ReporteMensual,
    Evidencia
)

from vision_analyzer import analizar_imagen


# ============================================================
# CONFIGURACIÓN GEMINI
# ============================================================

GEMINI_MIN_INTERVAL = 4.1

_gemini_last_call = 0.0
_gemini_lock = threading.Lock()


# ============================================================
# RATE LIMITER GEMINI
# ============================================================

def esperar_rate_limit_gemini():
    """
    Respeta el intervalo mínimo entre llamadas a Gemini.
    """

    global _gemini_last_call

    with _gemini_lock:

        ahora = time.time()

        transcurrido = (
            ahora - _gemini_last_call
        )

        if transcurrido < GEMINI_MIN_INTERVAL:

            esperar = (
                GEMINI_MIN_INTERVAL
                - transcurrido
            )

            time.sleep(
                esperar
            )

        _gemini_last_call = time.time()


# ============================================================
# PROCESAMIENTO PRINCIPAL
# ============================================================

def procesar_carga_masiva_job(
    app,
    job_id,
    contrato,
    obligaciones,
    mes,
    anio,
    excel_path,
    imagenes_subidas,
    api_key,
    actualizar_progreso
):
    """
    Procesa una carga masiva mensual en segundo plano.

    El blueprint se encarga de:
    - recibir archivos,
    - crear el job,
    - iniciar el hilo,
    - enviar progreso mediante SSE.

    Este servicio se encarga exclusivamente de la lógica
    de procesamiento.
    """

    try:

        with app.app_context():

            resultado = _procesar_excel(
                job_id=job_id,
                contrato=contrato,
                obligaciones=obligaciones,
                mes=mes,
                anio=anio,
                excel_path=excel_path,
                imagenes_subidas=imagenes_subidas,
                api_key=api_key,
                actualizar_progreso=actualizar_progreso
            )

            actualizar_progreso(
                job_id,
                'completado',
                100,
                (
                    'Proceso finalizado. '
                    f"{resultado['exitosos']} "
                    'evidencias cargadas.'
                ),
                resultado={
                    'exitosos': resultado['exitosos'],
                    'mes': mes,
                    'anio': anio
                },
                errores=resultado['errores']
            )

    except Exception as exc:

        import traceback

        traceback.print_exc()

        actualizar_progreso(
            job_id,
            'error',
            0,
            (
                'Error inesperado: '
                f'{str(exc)}'
            )
        )

    finally:

        _limpiar_archivo(
            excel_path
        )


# ============================================================
# PROCESAR EXCEL
# ============================================================

def _procesar_excel(
    job_id,
    contrato,
    obligaciones,
    mes,
    anio,
    excel_path,
    imagenes_subidas,
    api_key,
    actualizar_progreso
):
    """
    Lee el Excel y procesa todas sus filas.
    """

    wb = load_workbook(
        excel_path
    )

    ws = wb.active

    _validar_encabezados(
        ws
    )

    filas_validas = _obtener_filas_validas(
        ws
    )

    if not filas_validas:

        raise ValueError(
            'No se encontraron filas válidas en el Excel.'
        )

    total_filas = len(
        filas_validas
    )

    errores = []

    exitosos = 0

    reportes_cache = {}

    evidencias_por_reporte = {}

    fechas_mes = _obtener_fechas_mes(
        mes,
        anio
    )

    imagenes_disponibles = dict(
        imagenes_subidas
    )

    for posicion, (
        fila_excel,
        row
    ) in enumerate(
        filas_validas,
        start=1
    ):

        porcentaje = int(
            (
                (posicion - 1)
                / total_filas
            )
            * 100
        )

        actualizar_progreso(
            job_id,
            'procesando',
            porcentaje,
            (
                f'Procesando fila '
                f'{fila_excel} '
                f'({posicion}/{total_filas})...'
            )
        )

        resultado_fila = _procesar_fila(
            job_id=job_id,
            fila_excel=fila_excel,
            row=row,
            contrato=contrato,
            mes=mes,
            anio=anio,
            fechas_mes=fechas_mes,
            api_key=api_key,
            imagenes_disponibles=imagenes_disponibles,
            reportes_cache=reportes_cache,
            evidencias_por_reporte=evidencias_por_reporte,
            actualizar_progreso=actualizar_progreso
        )

        if resultado_fila['exitoso']:

            exitosos += 1

        errores.extend(
            resultado_fila['errores']
        )

    db.session.commit()

    _limpiar_archivos_temporales(
        imagenes_disponibles.values()
    )

    return {
        'exitosos': exitosos,
        'errores': errores
    }


# ============================================================
# VALIDAR ENCABEZADOS
# ============================================================

def _validar_encabezados(ws):
    """
    Valida los encabezados obligatorios del Excel.
    """

    headers = [
        cell.value
        for cell in ws[1]
    ]

    expected = [
        'Obligacion No.',
        'Descripcion Obligacion',
        'Anuncio / Contexto',
        'Fecha de la actividad',
        'Nombre Imagen'
    ]

    if headers[:5] != expected:

        raise ValueError(
            'Encabezados incorrectos. '
            f'Esperado: {expected}'
        )


# ============================================================
# OBTENER FILAS VÁLIDAS
# ============================================================

def _obtener_filas_validas(ws):
    """
    Obtiene las filas que contienen información.
    """

    filas = []

    for idx, row in enumerate(
        ws.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        if (
            not row[0]
            and
            not row[2]
        ):
            continue

        filas.append(
            (
                idx,
                row
            )
        )

    return filas


# ============================================================
# PROCESAR FILA
# ============================================================

def _procesar_fila(
    job_id,
    fila_excel,
    row,
    contrato,
    mes,
    anio,
    fechas_mes,
    api_key,
    imagenes_disponibles,
    reportes_cache,
    evidencias_por_reporte,
    actualizar_progreso
):
    """
    Procesa una única fila del Excel.
    """

    errores = []

    obl_num = row[0]

    anuncio = str(
        row[2] or ''
    ).strip()

    fecha_str = str(
        row[3] or ''
    ).strip()

    nombre_imagen = str(
        row[4] or ''
    ).strip()

    # --------------------------------------------------------
    # OBLIGACIÓN
    # --------------------------------------------------------

    try:

        obl_num_int = int(
            obl_num
        )

    except (
        ValueError,
        TypeError
    ):

        errores.append(
            f'Fila {fila_excel}: '
            f'Número de obligación inválido '
            f'({obl_num}).'
        )

        return {
            'exitoso': False,
            'errores': errores
        }

    obligacion = (
        Obligacion.query
        .filter_by(
            numero=obl_num_int,
            contrato_id=contrato.id
        )
        .first()
    )

    if not obligacion:

        errores.append(
            f'Fila {fila_excel}: '
            f'Obligación No. '
            f'{obl_num_int} '
            'no encontrada.'
        )

        return {
            'exitoso': False,
            'errores': errores
        }

    # --------------------------------------------------------
    # ANUNCIO
    # --------------------------------------------------------

    if not anuncio:

        errores.append(
            f'Fila {fila_excel}: '
            'Anuncio vacío.'
        )

        return {
            'exitoso': False,
            'errores': errores
        }

    # --------------------------------------------------------
    # FECHA
    # --------------------------------------------------------

    fecha_actividad = _parsear_fecha(
        fecha_str
    )

    if fecha_str and not fecha_actividad:

        errores.append(
            f'Fila {fila_excel}: '
            f'Fecha inválida ({fecha_str}).'
        )

        return {
            'exitoso': False,
            'errores': errores
        }

    if fecha_actividad:

        if (
            fecha_actividad
            < fechas_mes['inicio']
            or
            fecha_actividad
            > fechas_mes['fin']
        ):

            errores.append(
                f'Fila {fila_excel}: '
                f'Fecha {fecha_str} '
                f'fuera del mes '
                f'{mes}/{anio}.'
            )

            return {
                'exitoso': False,
                'errores': errores
            }

    else:

        fecha_actividad = date(
            anio,
            mes,
            15
        )

    # --------------------------------------------------------
    # REPORTE
    # --------------------------------------------------------

    reporte = _obtener_reporte(
        obligacion=obligacion,
        mes=mes,
        anio=anio,
        fechas_mes=fechas_mes,
        reportes_cache=reportes_cache,
        evidencias_por_reporte=evidencias_por_reporte
    )

    # --------------------------------------------------------
    # IMAGEN
    # --------------------------------------------------------

    imagen_path = ''

    if nombre_imagen:

        actualizar_progreso(
            job_id,
            'procesando',
            0,
            (
                f'Fila {fila_excel}: '
                f'Buscando imagen '
                f'"{nombre_imagen}"...'
            )
        )

        imagen_path = _procesar_imagen(
            nombre_imagen=nombre_imagen,
            reporte=reporte,
            imagenes_disponibles=imagenes_disponibles
        )

        if not imagen_path:

            errores.append(
                f'Fila {fila_excel}: '
                f'Imagen "{nombre_imagen}" '
                'no encontrada.'
            )

    # --------------------------------------------------------
    # NÚMERO DE ACTIVIDAD
    # --------------------------------------------------------

    evidencias_por_reporte[
        reporte.id
    ] += 1

    numero_actividad = (
        evidencias_por_reporte[
            reporte.id
        ]
    )

    # --------------------------------------------------------
    # CREAR EVIDENCIA
    # --------------------------------------------------------

    evidencia = Evidencia(
        numero_actividad=numero_actividad,
        imagen_path=imagen_path,
        anuncio_usuario=anuncio,
        descripcion_visual_ia=None,
        descripcion_actividad='',
        fecha_actividad=fecha_actividad,
        reporte_id=reporte.id
    )

    # --------------------------------------------------------
    # GEMINI
    # --------------------------------------------------------

    if (
        imagen_path
        and
        api_key
    ):

        actualizar_progreso(
            job_id,
            'procesando',
            0,
            (
                f'Fila {fila_excel}: '
                'Analizando con Gemini...'
            )
        )

        try:

            esperar_rate_limit_gemini()

            descripcion_visual = (
                analizar_imagen(
                    imagen_path,
                    api_key
                )
            )

            if descripcion_visual:

                evidencia.descripcion_visual_ia = (
                    descripcion_visual
                )

        except Exception as exc:

            print(
                '[CargaMasiva] '
                f'Error IA fila {fila_excel}: '
                f'{exc}'
            )

            errores.append(
                f'Fila {fila_excel}: '
                'Error al analizar imagen '
                f'con IA ({str(exc)[:60]}).'
            )

    # --------------------------------------------------------
    # DESCRIPCIÓN AUTOMÁTICA
    # --------------------------------------------------------

    evidencia.descripcion_actividad = (
        evidencia.generar_descripcion_automatica(
            obligacion
        )
    )

    db.session.add(
        evidencia
    )

    return {
        'exitoso': True,
        'errores': errores
    }


# ============================================================
# OBTENER REPORTE
# ============================================================

def _obtener_reporte(
    obligacion,
    mes,
    anio,
    fechas_mes,
    reportes_cache,
    evidencias_por_reporte
):
    """
    Obtiene un reporte existente o crea uno nuevo.
    """

    cache_key = (
        obligacion.id,
        mes,
        anio
    )

    if cache_key in reportes_cache:

        return reportes_cache[
            cache_key
        ]

    reporte = (
        ReporteMensual.query
        .filter_by(
            mes=mes,
            anio=anio,
            obligacion_id=obligacion.id
        )
        .first()
    )

    if not reporte:

        reporte = ReporteMensual(
            mes=mes,
            anio=anio,
            fecha_inicio_reporte=(
                fechas_mes['inicio']
            ),
            fecha_fin_reporte=(
                fechas_mes['fin']
            ),
            obligacion_id=obligacion.id
        )

        db.session.add(
            reporte
        )

        db.session.flush()

    reportes_cache[
        cache_key
    ] = reporte

    ultima = (
        Evidencia.query
        .filter_by(
            reporte_id=reporte.id
        )
        .order_by(
            Evidencia
            .numero_actividad
            .desc()
        )
        .first()
    )

    evidencias_por_reporte[
        reporte.id
    ] = (
        ultima.numero_actividad
        if ultima
        else 0
    )

    return reporte


# ============================================================
# PROCESAR IMAGEN
# ============================================================

def _procesar_imagen(
    nombre_imagen,
    reporte,
    imagenes_disponibles
):
    """
    Busca una imagen subida y la mueve al directorio
    definitivo de evidencias.
    """

    tmp_src = None

    # --------------------------------------------------------
    # NOMBRE EXACTO
    # --------------------------------------------------------

    if nombre_imagen in imagenes_disponibles:

        tmp_src = (
            imagenes_disponibles.pop(
                nombre_imagen
            )
        )

    # --------------------------------------------------------
    # NOMBRE SEGURO
    # --------------------------------------------------------

    else:

        safe_name = secure_filename(
            nombre_imagen
        )

        if safe_name in imagenes_disponibles:

            tmp_src = (
                imagenes_disponibles.pop(
                    safe_name
                )
            )

    if not tmp_src:

        return ''

    # --------------------------------------------------------
    # NOMBRE FINAL
    # --------------------------------------------------------

    final_name = secure_filename(
        (
            f'evidencia_'
            f'{reporte.id}_'
            f'{datetime.now().strftime("%Y%m%d_%H%M%S")}_'
            f'{nombre_imagen}'
        )
    )

    final_path = os.path.join(
        current_app.config[
            'UPLOAD_FOLDER'
        ],
        final_name
    )

    # --------------------------------------------------------
    # MOVER ARCHIVO
    # --------------------------------------------------------

    os.rename(
        tmp_src,
        final_path
    )

    return final_path


# ============================================================
# FECHAS DEL MES
# ============================================================

def _obtener_fechas_mes(
    mes,
    anio
):
    """
    Retorna las fechas inicial y final del mes.
    """

    _, last_day = calendar.monthrange(
        anio,
        mes
    )

    return {
        'inicio': date(
            anio,
            mes,
            1
        ),
        'fin': date(
            anio,
            mes,
            last_day
        )
    }


# ============================================================
# PARSEAR FECHA
# ============================================================

def _parsear_fecha(
    fecha
):
    """
    Convierte una fecha a datetime.date.

    Formatos soportados:
    - YYYY-MM-DD
    - DD/MM/YYYY
    - DD-MM-YYYY
    """

    if not fecha:

        return None

    # --------------------------------------------------------
    # SI OPENPYXL YA ENTREGA DATE
    # --------------------------------------------------------

    if isinstance(
        fecha,
        datetime
    ):

        return fecha.date()

    if isinstance(
        fecha,
        date
    ):

        return fecha

    fecha = str(
        fecha
    ).strip()

    formatos = (
        '%Y-%m-%d',
        '%d/%m/%Y',
        '%d-%m-%Y'
    )

    for formato in formatos:

        try:

            return datetime.strptime(
                fecha,
                formato
            ).date()

        except ValueError:

            continue

    return None


# ============================================================
# LIMPIEZA DE ARCHIVO
# ============================================================

def _limpiar_archivo(
    archivo_path
):
    """
    Elimina un archivo temporal.
    """

    try:

        if (
            archivo_path
            and
            os.path.exists(
                archivo_path
            )
        ):

            os.remove(
                archivo_path
            )

    except Exception:

        pass


# ============================================================
# LIMPIEZA DE IMÁGENES
# ============================================================

def _limpiar_archivos_temporales(
    archivos
):
    """
    Elimina las imágenes temporales
    que no fueron utilizadas.
    """

    for archivo_path in archivos:

        _limpiar_archivo(
            archivo_path
        )
