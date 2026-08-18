"""
Reporte de Cumplimiento de Obligaciones Contractuales
Aplicacion web en Flask con autenticacion, multi-usuario, descarga masiva y carga masiva.
"""
import os
import io
import zipfile
import calendar
from datetime import datetime, date
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, session, send_file, Response, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
from authlib.integrations.flask_client import OAuth

from config import Config, BASE_DIR
from models import db, Usuario, Contrato, Obligacion, ReporteMensual, Evidencia
from pdf_generator import PDFGenerator
from cryptography.fernet import Fernet

from vision_analyzer import analizar_imagen, verificar_api_key, _limpiar_key, consolidar_textos_ejecutivo

app = Flask(__name__)
app.config.from_object(Config)
app.config['UPLOAD_FOLDER'] = Config.UPLOAD_FOLDER
app.config['PDF_FOLDER'] = Config.PDF_FOLDER
app.config['MAX_CONTENT_LENGTH'] = Config.MAX_CONTENT_LENGTH

# Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicie sesion para acceder.'
login_manager.login_message_category = 'warning'

# OAuth Google
oauth = OAuth(app)

_GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '').strip()
_GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '').strip()

# Log de depuracion (solo si faltan credenciales)
if not _GOOGLE_CLIENT_ID or not _GOOGLE_CLIENT_SECRET:
    print("[ADVERTENCIA] Credenciales de Google OAuth no configuradas.")
    print("[ADVERTENCIA] Credenciales de Google OAuth no configuradas.")
    print("="*60)
    print("El inicio de sesion con Google NO funcionara.")
    print("Solucion: configure el archivo .env con GOOGLE_CLIENT_ID y GOOGLE_CLIENT_SECRET")
    print("O ejecute: python diagnostico_google.py")
    print("="*60 + "\n")

google = oauth.register(
    name='google',
    client_id=_GOOGLE_CLIENT_ID,
    client_secret=_GOOGLE_CLIENT_SECRET,
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={'scope': 'openid email profile'},
)

db.init_app(app)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def generar_meses_contrato(fecha_inicio, fecha_fin):
    meses = []
    current = date(fecha_inicio.year, fecha_inicio.month, 1)
    end = date(fecha_fin.year, fecha_fin.month, 1)
    nombres_meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    while current <= end:
        meses.append((current.month, current.year, nombres_meses[current.month]))
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)
    return meses


def migrar_db():
    with app.app_context():
        inspector = db.inspect(db.engine)
        tables = inspector.get_table_names()

        # Crear todas las tablas si no existen
        if 'usuario' not in tables:
            db.create_all()
            return

        # Migrar contrato: agregar columnas si no existen
        if 'contrato' in tables:
            cols = [c['name'] for c in inspector.get_columns('contrato')]
            with db.engine.connect() as conn:
                if 'user_id' not in cols:
                    conn.execute(db.text("ALTER TABLE contrato ADD COLUMN user_id INTEGER"))
                    conn.commit()
                if 'contratista' not in cols:
                    conn.execute(db.text("ALTER TABLE contrato ADD COLUMN contratista VARCHAR(200)"))
                    conn.commit()
                if 'numero_contrato' not in cols:
                    conn.execute(db.text("ALTER TABLE contrato ADD COLUMN numero_contrato VARCHAR(100)"))
                    conn.commit()
                if 'etapa' not in cols:
                    conn.execute(db.text("ALTER TABLE contrato ADD COLUMN etapa VARCHAR(50) DEFAULT 'Reporte en Proceso'"))
                    conn.execute(db.text("UPDATE contrato SET etapa = 'Reporte en Proceso' WHERE etapa IS NULL"))
                    conn.commit()

        # Corregir: eliminar columna etapa erronea de usuario (si existe)
        if 'usuario' in tables:
            cols_usuario = [c['name'] for c in inspector.get_columns('usuario')]
            with db.engine.connect() as conn:
                if 'etapa' in cols_usuario:
                    conn.execute(db.text("ALTER TABLE usuario DROP COLUMN etapa"))
                    conn.commit()


def _obtener_api_key():
    """Obtiene la API key de Gemini del usuario actual (DB desencriptada o session para compatibilidad)."""
    if current_user.is_authenticated:
        if current_user.gemini_api_key:
            return _decrypt_api_key(current_user.gemini_api_key)
    return session.get('gemini_api_key', '')


def _guardar_api_key(api_key):
    """Guarda la API key de Gemini encriptada en el usuario actual y en session."""
    encrypted = _encrypt_api_key(api_key)
    if current_user.is_authenticated:
        current_user.gemini_api_key = encrypted
        db.session.commit()
    if api_key:
        session['gemini_api_key'] = api_key
        session.permanent = True
    else:
        session.pop('gemini_api_key', None)


def _limpiar_key_externa(api_key):
    """Limpia espacios y saltos de linea de la API key."""
    if not api_key:
        return None
    return api_key.strip().replace("\n", "").replace("\r", "").replace(" ", "")



@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))


# ============================================================
# AUTENTICACION
# ============================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        remember = bool(request.form.get('remember'))

        if not email or not password:
            flash('Complete todos los campos.', 'danger')
            return redirect(url_for('login'))

        usuario = Usuario.query.filter_by(email=email).first()
        if usuario and usuario.check_password(password):
            login_user(usuario, remember=remember)
            next_page = request.args.get('next')
            flash(f'Bienvenido, {usuario.nombre or usuario.email}!', 'success')
            return redirect(next_page or url_for('inicio'))
        else:
            flash('Correo o contrasena incorrectos.', 'danger')
            return redirect(url_for('login'))

    google_configurado = bool(_GOOGLE_CLIENT_ID and _GOOGLE_CLIENT_SECRET)
    return render_template('login.html', google_configurado=google_configurado)


@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if current_user.is_authenticated:
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirmar = request.form.get('confirmar_password', '')

        if not nombre or not email or not password:
            flash('Complete todos los campos obligatorios.', 'danger')
            return redirect(url_for('registro'))

        if password != confirmar:
            flash('Las contrasenas no coinciden.', 'danger')
            return redirect(url_for('registro'))

        if len(password) < 6:
            flash('La contrasena debe tener al menos 6 caracteres.', 'danger')
            return redirect(url_for('registro'))

        existente = Usuario.query.filter_by(email=email).first()
        if existente:
            flash('Ya existe una cuenta con este correo.', 'danger')
            return redirect(url_for('registro'))

        nuevo = Usuario(email=email, nombre=nombre)
        nuevo.set_password(password)
        db.session.add(nuevo)
        db.session.commit()
        flash('Cuenta creada exitosamente. Inicie sesion.', 'success')
        return redirect(url_for('login'))

    google_configurado = bool(_GOOGLE_CLIENT_ID and _GOOGLE_CLIENT_SECRET)
    return render_template('registro.html', google_configurado=google_configurado)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    flash('Sesion cerrada.', 'info')
    return redirect(url_for('login'))


@app.route('/auth/google')
def auth_google():
    redirect_uri = url_for('auth_google_callback', _external=True)
    return google.authorize_redirect(redirect_uri)


@app.route('/auth/google/callback')
def auth_google_callback():
    try:
        token = google.authorize_access_token()
        resp = google.get('https://www.googleapis.com/oauth2/v3/userinfo')
        user_info = resp.json()

        email = user_info.get('email', '').lower()
        google_id = user_info.get('sub')
        nombre = user_info.get('name', email)
        avatar = user_info.get('picture', '')

        if not email:
            flash('No se pudo obtener el correo de Google.', 'danger')
            return redirect(url_for('login'))

        usuario = Usuario.query.filter_by(email=email).first()
        if not usuario:
            usuario = Usuario(
                email=email,
                nombre=nombre,
                auth_google=True,
                google_id=google_id,
                avatar_url=avatar
            )
            db.session.add(usuario)
            db.session.commit()
        else:
            if not usuario.auth_google:
                usuario.auth_google = True
                usuario.google_id = google_id
                usuario.avatar_url = avatar
                db.session.commit()

        login_user(usuario)
        flash(f'Bienvenido, {usuario.nombre or usuario.email}!', 'success')
        return redirect(url_for('inicio'))

    except Exception as e:
        error_msg = str(e)
        if 'invalid_client' in error_msg.lower():
            flash(
                'Error: El Client Secret de Google es invalido. '
                'Verifique que haya copiado el valor correcto de "Secreto de cliente" (no el ID de cliente). '
                'Ejecute: python diagnostico_google.py',
                'danger'
            )
        elif 'redirect_uri' in error_msg.lower():
            flash(
                'Error: La URI de redireccionamiento no coincide. '
                'En Google Cloud Console, agregue exactamente: http://127.0.0.1:5000/auth/google/callback',
                'danger'
            )
        elif 'unauthorized_client' in error_msg.lower():
            flash(
                'Error: El Client ID no es valido para aplicaciones web. '
                'Cree un ID de cliente OAuth 2.0 de tipo "Aplicacion web" en Google Cloud Console.',
                'danger'
            )
        else:
            flash(f'Error al autenticar con Google: {error_msg}', 'danger')
        return redirect(url_for('login'))


# ============================================================
# RUTAS PRINCIPALES (protegidas)
# ============================================================

@app.route('/inicio')
@login_required
def inicio():
    """Vista de inicio / landing page del sistema."""
    return render_template('inicio.html')

@app.route('/')
@login_required
def index():
    contrato = Contrato.query.filter_by(activo=True, user_id=current_user.id).first()
    obligaciones = []
    meses = []
    reportes_count = 0
    api_key_configurada = bool(_obtener_api_key())
    reportes_por_obligacion_mes = {}
    meses_con_reporte = set()

    # Paginación y filtro de obligaciones
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = request.args.get('search', '', type=str).strip()

    if contrato:
        query = Obligacion.query.filter_by(contrato_id=contrato.id)
        if search:
            query = query.filter(
                db.or_(
                    Obligacion.numero.cast(db.String).ilike(f'%{search}%'),
                    Obligacion.descripcion.ilike(f'%{search}%')
                )
            )
        obligaciones_pag = query.order_by(Obligacion.numero).paginate(
            page=page, per_page=per_page, error_out=False
        )
        obligaciones = obligaciones_pag.items
        meses = generar_meses_contrato(contrato.fecha_inicio, contrato.fecha_fin)
        reportes_count = ReporteMensual.query.join(Obligacion).filter(
            Obligacion.contrato_id == contrato.id
        ).count()

        for obl in obligaciones:
            reportes_obl = ReporteMensual.query.filter_by(obligacion_id=obl.id).all()
            reportes_por_obligacion_mes[obl.id] = {}
            for rep in reportes_obl:
                reportes_por_obligacion_mes[obl.id][(rep.mes, rep.anio)] = rep.id
                meses_con_reporte.add((rep.mes, rep.anio))

    return render_template('index.html',
                         contrato=contrato,
                         obligaciones=obligaciones,
                         obligaciones_pag=obligaciones_pag if contrato else None,
                         meses=meses,
                         reportes_count=reportes_count,
                         api_key_configurada=api_key_configurada,
                         reportes_por_obligacion_mes=reportes_por_obligacion_mes,
                         meses_con_reporte=meses_con_reporte,
                         search=search,
                         per_page=per_page)



# ============================================================
# GESTION DE CONTRATOS (multi-contrato)
# ============================================================

@app.route('/contratos')
@login_required
def contratos():
    """Vista de gestion de contratos con grilla, seleccion, acciones y obligaciones."""
    contratos_list = Contrato.query.filter_by(user_id=current_user.id).order_by(Contrato.fecha_creacion.desc()).all()

    # Preparar datos para JSON en el frontend (edicion)
    contratos_datos = []
    for c in contratos_list:
        contratos_datos.append({
            'id': c.id,
            'contratista': c.contratista or '',
            'numero_contrato': c.numero_contrato or '',
            'fecha_inicio': c.fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_fin': c.fecha_fin.strftime('%Y-%m-%d'),
            'activo': c.activo,
            'etapa': c.etapa
        })

    # ── Obligaciones del contrato activo ──
    contrato = Contrato.query.filter_by(activo=True, user_id=current_user.id).first()

    # Paginación y filtro de obligaciones
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = request.args.get('search', '', type=str).strip()

    obligaciones = []
    obligaciones_pag = None
    if contrato:
        query = Obligacion.query.filter_by(contrato_id=contrato.id)
        if search:
            query = query.filter(
                db.or_(
                    Obligacion.numero.cast(db.String).ilike(f'%{search}%'),
                    Obligacion.descripcion.ilike(f'%{search}%')
                )
            )
        obligaciones_pag = query.order_by(Obligacion.numero).paginate(
            page=page, per_page=per_page, error_out=False
        )
        obligaciones = obligaciones_pag.items

    return render_template('contratos.html',
                         contratos=contratos_list,
                         contratos_datos=contratos_datos,
                         generar_meses=generar_meses_contrato,
                         contrato=contrato,
                         obligaciones=obligaciones,
                         obligaciones_pag=obligaciones_pag,
                         search=search,
                         per_page=per_page,
                         obl_numero_error=session.pop('obl_numero_error', ''),
                         obl_descripcion_error=session.pop('obl_descripcion_error', ''))



@app.route('/contrato/nuevo', methods=['POST'])
@login_required
def contrato_nuevo():
    """Crea un nuevo contrato. Por defecto queda Inactivo."""
    contratista = request.form.get('contratista', '').strip()
    numero_contrato = request.form.get('numero_contrato', '').strip()

    if not contratista or not numero_contrato:
        flash('Contratista y Numero de contrato son obligatorios.', 'danger')
        return redirect(url_for('contratos'))

    fecha_inicio = datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date()
    fecha_fin = datetime.strptime(request.form['fecha_fin'], '%Y-%m-%d').date()

    nuevo = Contrato(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        contratista=contratista,
        numero_contrato=numero_contrato,
        activo=False,  # Nuevo contrato inactivo por defecto
        etapa='Reporte en Proceso',
        user_id=current_user.id
    )
    db.session.add(nuevo)
    db.session.commit()
    flash(f'Contrato "{contratista}" creado exitosamente. Recuerde activarlo para usarlo.', 'success')
    return redirect(url_for('contratos'))


@app.route('/contrato/<int:id>/editar', methods=['POST'])
@login_required
def contrato_editar(id):
    """Edita un contrato existente."""
    contrato = Contrato.query.get_or_404(id)
    if contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('contratos'))

    contratista = request.form.get('contratista', '').strip()
    numero_contrato = request.form.get('numero_contrato', '').strip()

    if not contratista or not numero_contrato:
        flash('Contratista y Numero de contrato son obligatorios.', 'danger')
        return redirect(url_for('contratos'))

    contrato.contratista = contratista
    contrato.numero_contrato = numero_contrato
    contrato.fecha_inicio = datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date()
    contrato.fecha_fin = datetime.strptime(request.form['fecha_fin'], '%Y-%m-%d').date()
    db.session.commit()
    flash('Contrato actualizado correctamente.', 'success')
    return redirect(url_for('contratos'))


@app.route('/contrato/<int:id>/eliminar', methods=['POST'])
@login_required
def contrato_eliminar(id):
    """Elimina un contrato solo si sus obligaciones NO tienen evidencias."""
    contrato = Contrato.query.get_or_404(id)
    if contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('contratos'))

    # Validacion en cascada: verificar si alguna obligacion tiene evidencias
    tiene_evidencias = False
    obligaciones_con_evidencias = []

    for obl in contrato.obligaciones:
        for rep in obl.reportes:
            if rep.evidencias:
                tiene_evidencias = True
                obligaciones_con_evidencias.append(f'Obligacion No. {obl.numero}')
                break

    if tiene_evidencias:
        flash(
            f'No se puede eliminar el contrato porque tiene evidencias registradas. '
            f'Elimine primero las evidencias de: {", ".join(obligaciones_con_evidencias)}.',
            'danger'
        )
        return redirect(url_for('contratos'))

    # Si no tiene evidencias, eliminar en cascada (obligaciones y reportes se borran por cascade)
    db.session.delete(contrato)
    db.session.commit()
    flash('Contrato eliminado correctamente.', 'success')
    return redirect(url_for('contratos'))


@app.route('/contrato/<int:id>/cambiar-estado', methods=['POST'])
@login_required
def contrato_cambiar_estado(id):
    """Activa o desactiva un contrato. Solo puede haber UN contrato activo por usuario."""
    contrato = Contrato.query.get_or_404(id)
    if contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('contratos'))

    if contrato.activo:
        # Desactivar: validar que no sea el unico contrato del usuario
        total_contratos = Contrato.query.filter_by(user_id=current_user.id).count()
        if total_contratos <= 1:
            flash('No puede inactivar el unico contrato existente. Cree otro contrato primero.', 'danger')
            return redirect(url_for('contratos'))

        contrato.activo = False
        db.session.commit()
        flash(f'Contrato "{contrato.contratista}" desactivado.', 'info')
    else:
        # Activar: primero desactivar todos los demas del usuario
        Contrato.query.filter_by(user_id=current_user.id, activo=True).update({'activo': False})
        contrato.activo = True
        db.session.commit()
        flash(f'Contrato "{contrato.contratista}" activado. Ahora puede gestionar sus obligaciones y reportes.', 'success')

    return redirect(url_for('contratos'))


@app.route('/contrato/<int:id>/finalizar', methods=['POST'])
@login_required
def contrato_finalizar(id):
    """
    Finaliza un contrato (cambia etapa a 'Reporte Cerrado').
    Solo permite si:
    1. Existen obligaciones registradas.
    2. Cada obligacion tiene un reporte para CADA MES del contrato.
    3. Cada reporte tiene al menos 1 evidencia.
    """
    contrato = Contrato.query.get_or_404(id)
    if contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('contratos'))

    obligaciones = Obligacion.query.filter_by(contrato_id=contrato.id).all()
    if not obligaciones:
        flash('No se puede finalizar el contrato porque no tiene obligaciones registradas.', 'danger')
        return redirect(url_for('contratos'))

    meses_contrato = generar_meses_contrato(contrato.fecha_inicio, contrato.fecha_fin)
    # meses_contrato = [(mes_num, anio, nombre), ...]

    errores = []

    for obl in obligaciones:
        reportes = ReporteMensual.query.filter_by(obligacion_id=obl.id).all()
        # Mapa de reportes existentes: {(mes, anio): reporte}
        reportes_map = {(rep.mes, rep.anio): rep for rep in reportes}

        meses_faltantes = []
        meses_sin_evidencia = []

        for mes_num, anio, nombre in meses_contrato:
            if (mes_num, anio) not in reportes_map:
                meses_faltantes.append(nombre + ' ' + str(anio))
            else:
                rep = reportes_map[(mes_num, anio)]
                if not rep.evidencias:
                    meses_sin_evidencia.append(nombre + ' ' + str(anio))

        if meses_faltantes:
            errores.append(
                f'Obligacion No. {obl.numero}: falta(n) reporte(s) para {", ".join(meses_faltantes)}.'
            )
        if meses_sin_evidencia:
            errores.append(
                f'Obligacion No. {obl.numero}: reporte(s) sin evidencia en {", ".join(meses_sin_evidencia)}.'
            )

    if errores:
        lista_html = '<ul class="mb-0">' + ''.join([f'<li>{e}</li>' for e in errores]) + '</ul>'
        flash(
            '<strong>No se puede finalizar el contrato.</strong><br>'
            'Cada obligacion debe tener un reporte con al menos una evidencia para <strong>TODOS los meses</strong> del contrato.<br><br>'
            f'<strong>Detalles encontrados ({len(errores)}):</strong>{lista_html}',
            'danger'
        )
        return redirect(url_for('contratos'))

    contrato.etapa = 'Reporte Cerrado'
    db.session.commit()
    flash(
        f'Contrato "{contrato.contratista}" finalizado exitosamente. '
        f'Etapa: Reporte Cerrado. No se podran agregar mas evidencias.',
        'success'
    )
    return redirect(url_for('contratos'))
# ── Encriptación de API Key ──
_ENCRYPTION_KEY = None
_FERNET = None


def _get_encryption_key():
    """Obtiene o genera la clave maestra de encriptación."""
    global _ENCRYPTION_KEY, _FERNET
    if _FERNET is not None:
        return _FERNET

    # 1. Intentar desde variable de entorno
    key_env = os.environ.get('ENCRYPTION_KEY', '').strip()
    if key_env:
        try:
            _FERNET = Fernet(key_env.encode())
            return _FERNET
        except Exception:
            pass

    # 2. Intentar desde archivo .key
    key_file = os.path.join(BASE_DIR, '.encryption_key')
    if os.path.exists(key_file):
        with open(key_file, 'rb') as f:
            key_data = f.read().strip()
        try:
            _FERNET = Fernet(key_data)
            return _FERNET
        except Exception:
            pass

    # 3. Generar nueva clave y guardarla
    key = Fernet.generate_key()
    with open(key_file, 'wb') as f:
        f.write(key)
    print(f"[INFO] Clave de encriptación generada y guardada en: {key_file}")
    print(f"[INFO] Guarde esta clave en su backup. Si la pierde, las API keys encriptadas serán irrecuperables.")
    _FERNET = Fernet(key)
    return _FERNET


def _encrypt_api_key(plain_key):
    """Encripta una API key en texto plano."""
    if not plain_key:
        return ''
    fernet = _get_encryption_key()
    return fernet.encrypt(plain_key.encode()).decode()


def _decrypt_api_key(encrypted_key):
    """Desencripta una API key. Si falla, asume texto plano (backward compat)."""
    if not encrypted_key:
        return ''
    fernet = _get_encryption_key()
    try:
        return fernet.decrypt(encrypted_key.encode()).decode()
    except Exception:
        # Backward compatibility: key guardada en texto plano antes de la encriptación
        return encrypted_key



@app.route('/config', methods=['GET', 'POST'])
@login_required
def config():
    contrato = Contrato.query.filter_by(activo=True, user_id=current_user.id).first()
    api_key = _obtener_api_key()
    api_key_valida = False
    api_key_error = ''
    modelo_disponible = ''
    if api_key:
        api_key_valida, resultado = verificar_api_key(api_key)
        if api_key_valida:
            modelo_disponible = resultado
        else:
            api_key_error = resultado

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'agregar_obligacion':
            if not contrato:
                flash('Primero debe configurar el contrato.', 'danger')
                return redirect(url_for('contratos'))

            numero = int(request.form['numero'])
            descripcion = request.form['descripcion']

            # Validar que el numero de obligacion no se repita en este contrato
            existente = Obligacion.query.filter_by(numero=numero, contrato_id=contrato.id).first()
            if existente:
                session['obl_numero_error'] = numero
                session['obl_descripcion_error'] = descripcion
                flash(f'Ya existe una obligacion con el numero {numero} en este contrato. Use otro numero.', 'danger')
                return redirect(url_for('contratos'))

            obligacion = Obligacion(numero=numero, descripcion=descripcion, contrato_id=contrato.id)
            db.session.add(obligacion)
            db.session.commit()
            flash(f'Obligacion No. {numero} agregada.', 'success')
            return redirect(url_for('contratos'))

        elif action == 'guardar_api_key':
            api_key_raw = request.form.get('gemini_api_key', '')
            api_key = _limpiar_key_externa(api_key_raw)
            if api_key:
                valida, mensaje = verificar_api_key(api_key)
                if valida:
                    _guardar_api_key(api_key)
                    flash('API key de Gemini configurada y verificada correctamente.', 'success')
                else:
                    flash(f'Error: {mensaje}', 'danger')
            else:
                _guardar_api_key('')
                flash('API key eliminada.', 'info')
            return redirect(url_for('config'))

    return render_template('config.html',
                         contrato=contrato,
                         api_key=api_key,
                         api_key_valida=api_key_valida,
                         api_key_error=api_key_error,
                         modelo_disponible=modelo_disponible)


@app.route('/obligacion/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_obligacion(id):
    obligacion = Obligacion.query.get_or_404(id)
    contrato = Contrato.query.get(obligacion.contrato_id)
    if contrato and contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('contratos'))
    db.session.delete(obligacion)
    db.session.commit()
    flash('Obligacion eliminada.', 'info')
    return redirect(url_for('contratos'))


@app.route('/obligacion/<int:id>/editar', methods=['POST'])
@login_required
def editar_obligacion(id):
    obligacion = Obligacion.query.get_or_404(id)
    contrato = Contrato.query.get(obligacion.contrato_id)
    if contrato and contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('contratos'))
    nueva_descripcion = request.form.get('descripcion', '').strip()
    if nueva_descripcion:
        obligacion.descripcion = nueva_descripcion
        db.session.commit()
        flash('Descripcion de la obligacion actualizada.', 'success')
    else:
        flash('La descripcion no puede estar vacia.', 'danger')
    return redirect(url_for('config'))


@app.route('/reportes')
@login_required
def reportes():
    contrato = Contrato.query.filter_by(activo=True, user_id=current_user.id).first()

    # Filtros y paginación
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = request.args.get('search', '', type=str).strip()
    filtro_mes = request.args.get('filtro_mes', '', type=str)
    filtro_anio = request.args.get('filtro_anio', '', type=str)
    filtro_obligacion = request.args.get('filtro_obligacion', '', type=str)

    reportes_list = []
    reportes_pag = None
    obligaciones_list = []

    if contrato:
        obligaciones_list = Obligacion.query.filter_by(contrato_id=contrato.id).order_by(Obligacion.numero).all()
        query = ReporteMensual.query.join(Obligacion).filter(
            Obligacion.contrato_id == contrato.id
        )

        if search:
            query = query.filter(
                db.or_(
                    Obligacion.descripcion.ilike(f'%{search}%'),
                    Obligacion.numero.cast(db.String).ilike(f'%{search}%')
                )
            )

        if filtro_mes:
            query = query.filter(ReporteMensual.mes == int(filtro_mes))
        if filtro_anio:
            query = query.filter(ReporteMensual.anio == int(filtro_anio))
        if filtro_obligacion:
            query = query.filter(Obligacion.id == int(filtro_obligacion))

        reportes_pag = query.order_by(ReporteMensual.anio.desc(), ReporteMensual.mes.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        reportes_list = reportes_pag.items

    return render_template('reportes.html',
                         reportes=reportes_list,
                         reportes_pag=reportes_pag,
                         contrato=contrato,
                         obligaciones_list=obligaciones_list,
                         search=search,
                         filtro_mes=filtro_mes,
                         filtro_anio=filtro_anio,
                         filtro_obligacion=filtro_obligacion,
                         per_page=per_page)


@app.route('/reporte/nuevo/<int:obligacion_id>', methods=['GET', 'POST'])
@login_required
def nuevo_reporte(obligacion_id):
    obligacion = Obligacion.query.get_or_404(obligacion_id)
    contrato = Contrato.query.get(obligacion.contrato_id)
    if not contrato or contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('inicio'))

    meses = generar_meses_contrato(contrato.fecha_inicio, contrato.fecha_fin)

    form_data = {
        'mes': session.pop('nuevo_rep_mes', ''),
        'anio': session.pop('nuevo_rep_anio', ''),
        'fecha_inicio_reporte': session.pop('nuevo_rep_fecha_inicio', ''),
        'fecha_fin_reporte': session.pop('nuevo_rep_fecha_fin', ''),
    }

    if request.method == 'POST':
        mes = int(request.form['mes'])
        anio = int(request.form['anio'])
        fecha_inicio_rep = datetime.strptime(request.form['fecha_inicio_reporte'], '%Y-%m-%d').date()
        fecha_fin_rep = datetime.strptime(request.form['fecha_fin_reporte'], '%Y-%m-%d').date()

        nombres_meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        nombre_mes = nombres_meses[mes]

        # ── Validación: fechas dentro del mes a reportar ──
        _, last_day = calendar.monthrange(anio, mes)
        primer_dia_mes = date(anio, mes, 1)
        ultimo_dia_mes = date(anio, mes, last_day)

        if fecha_inicio_rep < primer_dia_mes or fecha_inicio_rep > ultimo_dia_mes:
            session['nuevo_rep_mes'] = str(mes)
            session['nuevo_rep_anio'] = str(anio)
            session['nuevo_rep_fecha_inicio'] = request.form['fecha_inicio_reporte']
            session['nuevo_rep_fecha_fin'] = request.form['fecha_fin_reporte']
            flash(f'La fecha de inicio debe estar dentro de {nombre_mes} {anio} '
                  f'({primer_dia_mes.strftime("%d/%m/%Y")} – {ultimo_dia_mes.strftime("%d/%m/%Y")}).', 'danger')
            return redirect(url_for('nuevo_reporte', obligacion_id=obligacion_id))

        if fecha_fin_rep < primer_dia_mes or fecha_fin_rep > ultimo_dia_mes:
            session['nuevo_rep_mes'] = str(mes)
            session['nuevo_rep_anio'] = str(anio)
            session['nuevo_rep_fecha_inicio'] = request.form['fecha_inicio_reporte']
            session['nuevo_rep_fecha_fin'] = request.form['fecha_fin_reporte']
            flash(f'La fecha de fin debe estar dentro de {nombre_mes} {anio} '
                  f'({primer_dia_mes.strftime("%d/%m/%Y")} – {ultimo_dia_mes.strftime("%d/%m/%Y")}).', 'danger')
            return redirect(url_for('nuevo_reporte', obligacion_id=obligacion_id))

        if fecha_inicio_rep > fecha_fin_rep:
            session['nuevo_rep_mes'] = str(mes)
            session['nuevo_rep_anio'] = str(anio)
            session['nuevo_rep_fecha_inicio'] = request.form['fecha_inicio_reporte']
            session['nuevo_rep_fecha_fin'] = request.form['fecha_fin_reporte']
            flash('La fecha de inicio no puede ser posterior a la fecha de fin.', 'danger')
            return redirect(url_for('nuevo_reporte', obligacion_id=obligacion_id))

        # ── Validación: consecutividad de meses ──
        if mes == 1:
            mes_ant, anio_ant = 12, anio - 1
        else:
            mes_ant, anio_ant = mes - 1, anio

        fecha_mes_ant = date(anio_ant, mes_ant, 1)
        fecha_inicio_contrato_mes = date(contrato.fecha_inicio.year, contrato.fecha_inicio.month, 1)

        if fecha_mes_ant >= fecha_inicio_contrato_mes:
            reporte_anterior = ReporteMensual.query.filter_by(
                mes=mes_ant, anio=anio_ant, obligacion_id=obligacion_id
            ).first()

            if not reporte_anterior:
                session['nuevo_rep_mes'] = str(mes)
                session['nuevo_rep_anio'] = str(anio)
                session['nuevo_rep_fecha_inicio'] = request.form['fecha_inicio_reporte']
                session['nuevo_rep_fecha_fin'] = request.form['fecha_fin_reporte']
                flash(f'No puede saltar meses. Cree primero el reporte de '
                      f'{nombres_meses[mes_ant]} {anio_ant} antes de {nombre_mes} {anio}.', 'danger')
                return redirect(url_for('nuevo_reporte', obligacion_id=obligacion_id))

        # ── Validación: reporte duplicado ──
        existente = ReporteMensual.query.filter_by(
            mes=mes, anio=anio, obligacion_id=obligacion_id
        ).first()

        if existente:
            session['nuevo_rep_mes'] = str(mes)
            session['nuevo_rep_anio'] = str(anio)
            session['nuevo_rep_fecha_inicio'] = request.form['fecha_inicio_reporte']
            session['nuevo_rep_fecha_fin'] = request.form['fecha_fin_reporte']
            flash(f'Ya existe un reporte para {existente.nombre_mes} {anio}.', 'warning')
            return redirect(url_for('nuevo_reporte', obligacion_id=obligacion_id))

        reporte = ReporteMensual(
            mes=mes, anio=anio,
            fecha_inicio_reporte=fecha_inicio_rep,
            fecha_fin_reporte=fecha_fin_rep,
            obligacion_id=obligacion_id
        )
        db.session.add(reporte)
        db.session.commit()
        flash(f'Reporte de {reporte.nombre_mes} {anio} creado.', 'success')
        return redirect(url_for('ver_reporte', id=reporte.id))

    return render_template('nuevo_reporte.html',
                         obligacion=obligacion,
                         meses=meses,
                         contrato=contrato,
                         form_data=form_data)


@app.route('/reporte/<int:id>')
@login_required
def ver_reporte(id):
    reporte = ReporteMensual.query.get_or_404(id)
    obligacion = reporte.obligacion
    contrato = Contrato.query.get(obligacion.contrato_id)
    if not contrato or contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('inicio'))

    evidencias = Evidencia.query.filter_by(reporte_id=id).order_by(Evidencia.numero_actividad).all()
    api_key_configurada = bool(_obtener_api_key())

    form_data = {
        'anuncio_usuario': session.pop('evidencia_anuncio', ''),
        'fecha_actividad': session.pop('evidencia_fecha', ''),
    }

    return render_template('ver_reporte.html',
                         reporte=reporte,
                         obligacion=obligacion,
                         contrato=contrato,
                         evidencias=evidencias,
                         api_key_configurada=api_key_configurada,
                         form_data=form_data)


@app.route('/reporte/<int:id>/evidencia', methods=['POST'])
@login_required
def subir_evidencia(id):
    reporte = ReporteMensual.query.get_or_404(id)
    obligacion = reporte.obligacion
    contrato = Contrato.query.get(obligacion.contrato_id)
    if not contrato or contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('inicio'))

    if contrato.etapa == 'Reporte Cerrado':
        flash('Este contrato esta finalizado (Reporte Cerrado). No se pueden agregar mas evidencias.', 'warning')
        return redirect(url_for('ver_reporte', id=id))

    api_key = _obtener_api_key()
    session['evidencia_anuncio'] = request.form.get('anuncio_usuario', '')
    session['evidencia_fecha'] = request.form.get('fecha_actividad', '')

    if 'imagen' not in request.files:
        flash('No se selecciono ningun archivo.', 'danger')
        return redirect(url_for('ver_reporte', id=id))

    file = request.files['imagen']
    anuncio_usuario = request.form.get('anuncio_usuario', '').strip()

    if not anuncio_usuario:
        flash('Debe escribir un anuncio/contexto.', 'danger')
        return redirect(url_for('ver_reporte', id=id))

    if file.filename == '':
        flash('No se selecciono ningun archivo.', 'danger')
        return redirect(url_for('ver_reporte', id=id))

    if file and allowed_file(file.filename):
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = secure_filename(f"evidencia_{reporte.id}_{timestamp}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            ultima_evidencia = Evidencia.query.filter_by(reporte_id=reporte.id).order_by(Evidencia.numero_actividad.desc()).first()
            numero_actividad = (ultima_evidencia.numero_actividad + 1) if ultima_evidencia else 1
            from datetime import date
            fecha_actividad_str = request.form.get('fecha_actividad', '').strip()
            if fecha_actividad_str:
                fecha_actividad = datetime.strptime(fecha_actividad_str, '%Y-%m-%d').date()
            else:
                fecha_actividad = date.today()

            if fecha_actividad < reporte.fecha_inicio_reporte or fecha_actividad > reporte.fecha_fin_reporte:
                flash(f'La fecha de la actividad ({fecha_actividad.strftime("%d/%m/%Y")}) debe estar dentro del periodo del reporte: {reporte.fecha_inicio_reporte.strftime("%d/%m/%Y")} a {reporte.fecha_fin_reporte.strftime("%d/%m/%Y")}.', 'danger')
                return redirect(url_for('ver_reporte', id=id))

            evidencia = Evidencia(
                numero_actividad=numero_actividad,
                imagen_path=filepath,
                anuncio_usuario=anuncio_usuario,
                descripcion_visual_ia=None,
                descripcion_actividad='',
                fecha_actividad=fecha_actividad,
                reporte_id=reporte.id
            )
            descripcion_visual = None
            if api_key:
                flash('Analizando imagen con IA...', 'info')
                descripcion_visual = analizar_imagen(filepath, api_key)
                if descripcion_visual:
                    evidencia.descripcion_visual_ia = descripcion_visual
                    flash(f'IA detecto: "{descripcion_visual[:80]}..."', 'success')
                else:
                    flash('No se pudo analizar la imagen con IA.', 'warning')
            evidencia.descripcion_actividad = evidencia.generar_descripcion_automatica(obligacion)
            db.session.add(evidencia)
            db.session.commit()

            session.pop('evidencia_anuncio', None)
            session.pop('evidencia_fecha', None)
            flash(f'Actividad {numero_actividad} registrada.', 'success')

        except RequestEntityTooLarge:
            flash('El archivo es demasiado grande. Maximo 16MB.', 'danger')
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
    else:
        flash('Formato no permitido.', 'danger')

    return redirect(url_for('ver_reporte', id=id))


@app.route('/reporte/<int:id>/evidencia/<int:evidencia_id>/eliminar', methods=['POST'])
@login_required
def eliminar_evidencia(id, evidencia_id):
    evidencia = Evidencia.query.get_or_404(evidencia_id)
    reporte = ReporteMensual.query.get_or_404(id)
    obligacion = reporte.obligacion
    contrato = Contrato.query.get(obligacion.contrato_id)
    if not contrato or contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('inicio'))

    numero_eliminado = evidencia.numero_actividad
    try:
        if os.path.exists(evidencia.imagen_path):
            os.remove(evidencia.imagen_path)
    except Exception:
        pass
    db.session.delete(evidencia)
    db.session.commit()
    evidencias_restantes = Evidencia.query.filter_by(reporte_id=id).order_by(Evidencia.numero_actividad).all()
    for idx, ev in enumerate(evidencias_restantes, start=1):
        ev.numero_actividad = idx
    db.session.commit()
    flash(f'Evidencia de Actividad {numero_eliminado} eliminada.', 'info')
    return redirect(url_for('ver_reporte', id=id))


@app.route('/reporte/<int:id>/evidencia/<int:evidencia_id>/editar', methods=['POST'])
@login_required
def editar_evidencia(id, evidencia_id):
    evidencia = Evidencia.query.get_or_404(evidencia_id)
    reporte = ReporteMensual.query.get_or_404(id)
    obligacion = reporte.obligacion
    contrato = Contrato.query.get(obligacion.contrato_id)
    if not contrato or contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('inicio'))

    nueva_descripcion = request.form.get('descripcion_actividad', '').strip()
    if nueva_descripcion:
        evidencia.descripcion_actividad = nueva_descripcion
    fecha_actividad_str = request.form.get('fecha_actividad', '').strip()
    if fecha_actividad_str:
        from datetime import datetime as dt
        evidencia.fecha_actividad = dt.strptime(fecha_actividad_str, '%Y-%m-%d').date()
    db.session.commit()
    flash(f'Actividad {evidencia.numero_actividad} actualizada.', 'success')
    return redirect(url_for('ver_reporte', id=id))


@app.route('/reporte/<int:id>/pdf')
@login_required
def generar_pdf(id):
    reporte = ReporteMensual.query.get_or_404(id)
    obligacion = reporte.obligacion
    contrato = Contrato.query.get(obligacion.contrato_id)
    if not contrato or contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('inicio'))

    evidencias = Evidencia.query.filter_by(reporte_id=id).order_by(Evidencia.numero_actividad).all()

    # ── Validación: no generar PDF vacío ──
    if not evidencias:
        flash('No se puede generar el PDF porque este reporte no tiene evidencias registradas. '
              'Agregue al menos una evidencia antes de descargar.', 'warning')
        return redirect(url_for('ver_reporte', id=id))

    pdf_filename = f"Reporte_Obligacion_{obligacion.numero}_{reporte.nombre_mes}_{reporte.anio}.pdf"
    pdf_path = os.path.join(app.config['PDF_FOLDER'], pdf_filename)
    try:
        generator = PDFGenerator(pdf_path)
        generator.generar_reporte(reporte, obligacion, evidencias, contrato)
        flash(f'PDF generado: {pdf_filename}', 'success')
        return send_from_directory(app.config['PDF_FOLDER'], pdf_filename, as_attachment=True)
    except Exception as e:
        flash(f'Error al generar PDF: {str(e)}', 'danger')
        return redirect(url_for('ver_reporte', id=id))



@app.route('/reporte/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_reporte(id):
    reporte = ReporteMensual.query.get_or_404(id)
    obligacion = reporte.obligacion
    contrato = Contrato.query.get(obligacion.contrato_id)
    if not contrato or contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('inicio'))

    for ev in reporte.evidencias:
        try:
            if os.path.exists(ev.imagen_path):
                os.remove(ev.imagen_path)
        except Exception:
            pass
    db.session.delete(reporte)
    db.session.commit()
    flash('Reporte eliminado.', 'info')
    return redirect(url_for('reportes'))


@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# ============================================================
# DESCARGA MASIVA DE PDFs POR MES
# ============================================================

@app.route('/reportes/descargar-mes', methods=['POST'])
@login_required
def descargar_masivo_mes():
    """Descarga todos los PDFs de un mes especifico en un ZIP. Valida que todas las obligaciones tengan evidencias."""

    # Detectar si es petición AJAX/fetch
    es_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    def responder_error(mensaje, codigo=400):
        if es_ajax:
            return jsonify({'error': mensaje}), codigo
        flash(mensaje, 'danger')
        return redirect(url_for('reportes'))

    mes = int(request.form.get('mes', 0))
    anio = int(request.form.get('anio', 0))

    if not mes or not anio:
        return responder_error('Seleccione mes y anio.')

    contrato = Contrato.query.filter_by(activo=True, user_id=current_user.id).first()
    if not contrato:
        return responder_error('No hay contrato activo.')

    nombres_meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    nombre_mes = nombres_meses[mes]

    # Obtener todas las obligaciones del contrato
    obligaciones = Obligacion.query.filter_by(contrato_id=contrato.id).order_by(Obligacion.numero).all()

    # ── Validación: todas las obligaciones deben tener reporte con evidencia ──
    obligaciones_faltantes = []
    obligaciones_sin_evidencia = []
    reportes_validos = []

    for obl in obligaciones:
        reporte = ReporteMensual.query.filter_by(
            mes=mes, anio=anio, obligacion_id=obl.id
        ).first()

        if not reporte:
            obligaciones_faltantes.append(f'No. {obl.numero}')
        elif not reporte.evidencias:
            obligaciones_sin_evidencia.append(f'No. {obl.numero}')
        else:
            reportes_validos.append(reporte)

    if obligaciones_faltantes or obligaciones_sin_evidencia:
        mensajes = []
        if obligaciones_faltantes:
            mensajes.append(f'<strong>Sin reporte para {nombre_mes} {anio}:</strong> {", ".join(obligaciones_faltantes)}')
        if obligaciones_sin_evidencia:
            mensajes.append(f'<strong>Reporte sin evidencias para {nombre_mes} {anio}:</strong> {", ".join(obligaciones_sin_evidencia)}')

        error_html = (
            '<strong>No se puede descargar el ZIP.</strong><br>'
            'Todas las obligaciones deben tener un reporte con al menos una evidencia para el mes seleccionado.<br><br>'
            + '<br>'.join(mensajes)
        )
        return responder_error(error_html)

    if not reportes_validos:
        return responder_error(f'No hay reportes con evidencias para {nombre_mes} {anio}.')

    # Generar PDFs y empaquetar en ZIP en memoria
    memory_zip = io.BytesIO()
    with zipfile.ZipFile(memory_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        for rep in reportes_validos:
            obligacion = rep.obligacion
            evidencias = Evidencia.query.filter_by(reporte_id=rep.id).order_by(Evidencia.numero_actividad).all()
            pdf_filename = f"Reporte_Obligacion_{obligacion.numero}_{rep.nombre_mes}_{rep.anio}.pdf"
            pdf_path = os.path.join(app.config['PDF_FOLDER'], pdf_filename)

            try:
                generator = PDFGenerator(pdf_path)
                generator.generar_reporte(rep, obligacion, evidencias, contrato)
                zf.write(pdf_path, arcname=pdf_filename)
            except Exception as e:
                print(f"Error generando PDF para obligacion {obligacion.numero}: {e}")
                continue

    memory_zip.seek(0)
    zip_filename = f"Reportes_{contrato.contratista or 'Contrato'}_{mes}_{anio}.zip"

    return send_file(
        memory_zip,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_filename
    )



# ============================================================
# CARGA MASIVA DE EVIDENCIAS DESDE EXCEL
# ============================================================

@app.route('/reporte/<int:id>/carga-masiva', methods=['GET', 'POST'])
@login_required
def carga_masiva_evidencias(id):
    """Carga masiva de evidencias desde Excel + imagenes."""
    reporte = ReporteMensual.query.get_or_404(id)
    obligacion = reporte.obligacion
    contrato = Contrato.query.get(obligacion.contrato_id)
    if not contrato or contrato.user_id != current_user.id:
        flash('No tiene permiso.', 'danger')
        return redirect(url_for('inicio'))

    if contrato.etapa == 'Reporte Cerrado':
        flash('Este contrato esta finalizado (Reporte Cerrado). No se pueden agregar mas evidencias.', 'warning')
        return redirect(url_for('ver_reporte', id=id))

    if request.method == 'POST':
        if 'archivo_excel' not in request.files:
            flash('No se selecciono el archivo Excel.', 'danger')
            return redirect(url_for('carga_masiva_evidencias', id=id))

        archivo_excel = request.files['archivo_excel']
        if archivo_excel.filename == '':
            flash('No se selecciono archivo Excel.', 'danger')
            return redirect(url_for('carga_masiva_evidencias', id=id))

        # Validar extension Excel
        if not archivo_excel.filename.endswith(('.xlsx', '.xls')):
            flash('El archivo debe ser Excel (.xlsx o .xls).', 'danger')
            return redirect(url_for('carga_masiva_evidencias', id=id))

        try:
            from openpyxl import load_workbook
            wb = load_workbook(archivo_excel)
            ws = wb.active

            # Validar encabezados
            headers = [cell.value for cell in ws[1]]
            expected = ['Anuncio / Contexto', 'Fecha de la actividad']
            if headers[:2] != expected:
                flash(f'Encabezados incorrectos. Se esperaba: {expected}. Encontrado: {headers[:2]}', 'danger')
                return redirect(url_for('carga_masiva_evidencias', id=id))

            api_key = _obtener_api_key()
            exitosos = 0
            errores = []

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                anuncio = str(row[0] or '').strip()
                fecha_str = str(row[1] or '').strip()

                if not anuncio:
                    errores.append(f'Fila {idx}: Anuncio vacio.')
                    continue
                if not fecha_str:
                    errores.append(f'Fila {idx}: Fecha vacia.')
                    continue

                try:
                    fecha_actividad = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                except ValueError:
                    try:
                        fecha_actividad = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                    except ValueError:
                        errores.append(f'Fila {idx}: Fecha invalida ({fecha_str}). Use YYYY-MM-DD o DD/MM/YYYY.')
                        continue

                if fecha_actividad < reporte.fecha_inicio_reporte or fecha_actividad > reporte.fecha_fin_reporte:
                    errores.append(f'Fila {idx}: Fecha {fecha_str} fuera del periodo del reporte.')
                    continue

                # Para carga masiva sin imagen individual, usamos un placeholder
                # El usuario debe cargar las imagenes despues o usar el metodo normal
                # Aqui generamos la actividad sin imagen (texto solo)
                ultima_evidencia = Evidencia.query.filter_by(reporte_id=reporte.id).order_by(Evidencia.numero_actividad.desc()).first()
                numero_actividad = (ultima_evidencia.numero_actividad + 1) if ultima_evidencia else 1

                evidencia = Evidencia(
                    numero_actividad=numero_actividad,
                    imagen_path='',  # Sin imagen en carga masiva Excel
                    anuncio_usuario=anuncio,
                    descripcion_visual_ia=None,
                    descripcion_actividad='',
                    fecha_actividad=fecha_actividad,
                    reporte_id=reporte.id
                )
                evidencia.descripcion_actividad = evidencia.generar_descripcion_automatica(obligacion)
                db.session.add(evidencia)
                exitosos += 1

            db.session.commit()

            if exitosos > 0:
                flash(f'{exitosos} actividades cargadas exitosamente desde Excel.', 'success')
            if errores:
                for err in errores[:5]:
                    flash(err, 'warning')
                if len(errores) > 5:
                    flash(f'... y {len(errores) - 5} errores mas.', 'warning')

        except Exception as e:
            flash(f'Error al procesar Excel: {str(e)}', 'danger')

        return redirect(url_for('ver_reporte', id=id))

    return render_template('carga_masiva.html', reporte=reporte, obligacion=obligacion, contrato=contrato)


# ============================================================
# EXCEL CONSOLIDADO
# ============================================================

@app.route('/reporte/consolidado/excel')
@login_required
def generar_excel_consolidado():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Detectar si es petición AJAX/fetch
    es_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    def responder_error(mensaje, codigo=400):
        if es_ajax:
            return jsonify({'error': mensaje}), codigo
        flash(mensaje, 'danger')
        return redirect(url_for('index'))

    def responder_warning(mensaje):
        if es_ajax:
            return jsonify({'error': mensaje}), 400
        flash(mensaje, 'warning')
        return redirect(url_for('index'))

    try:
        contrato = Contrato.query.filter_by(activo=True, user_id=current_user.id).first()
        if not contrato:
            return responder_error('No hay contrato activo configurado.')

        # Obtener mes y año solicitados
        mes = request.args.get('mes', type=int)
        anio = request.args.get('anio', type=int)

        if not mes or not anio:
            return responder_warning('Debe seleccionar mes y año para generar el consolidado.')

        nombres_meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        nombre_mes = nombres_meses[mes]

        obligaciones = Obligacion.query.filter_by(contrato_id=contrato.id).order_by(Obligacion.numero).all()
        if not obligaciones:
            return responder_error('No hay obligaciones registradas.')

        # ── Validación: todas las obligaciones deben tener reporte con evidencia para el mes/año ──
        obligaciones_faltantes = []
        obligaciones_sin_evidencia = []
        reportes_validos = {}

        for obl in obligaciones:
            reporte = ReporteMensual.query.filter_by(
                mes=mes, anio=anio, obligacion_id=obl.id
            ).first()

            if not reporte:
                obligaciones_faltantes.append(f'No. {obl.numero}')
            else:
                if not reporte.evidencias:
                    obligaciones_sin_evidencia.append(f'No. {obl.numero}')
                else:
                    reportes_validos[obl.id] = reporte

        if obligaciones_faltantes or obligaciones_sin_evidencia:
            mensajes = []
            if obligaciones_faltantes:
                mensajes.append(f'<strong>Sin reporte para {nombre_mes} {anio}:</strong> {", ".join(obligaciones_faltantes)}')
            if obligaciones_sin_evidencia:
                mensajes.append(f'<strong>Reporte sin evidencias para {nombre_mes} {anio}:</strong> {", ".join(obligaciones_sin_evidencia)}')

            error_html = (
                '<strong>No se puede generar el consolidado.</strong><br>'
                'Todas las obligaciones deben tener un reporte con al menos una evidencia para el mes seleccionado.<br><br>'
                + '<br>'.join(mensajes)
            )
            return responder_error(error_html)

        # ── Generar Excel consolidado para el mes/año específico ──
        wb = Workbook()
        ws = wb.active
        ws.title = f"Consolidado_{nombre_mes}_{anio}"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Encabezado del consolidado
        ws.append([f"REPORTE CONSOLIDADO - {contrato.contratista or 'Contrato'}"])
        ws.append([f"Mes: {nombre_mes} {anio}  |  N° Contrato: {contrato.numero_contrato or 'N/A'}"])
        ws.append([])

        headers = ["No. Obligacion", "Obligacion Contractual", "Texto Ejecutivo - Actividades Reportadas"]
        ws.append(headers)
        for cell in ws[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        api_key = _obtener_api_key()

        for obl in obligaciones:
            reporte = reportes_validos[obl.id]
            evidencias = Evidencia.query.filter_by(reporte_id=reporte.id).order_by(
                Evidencia.numero_actividad
            ).all()

            actividades_textos = [ev.descripcion_actividad for ev in evidencias]

            if actividades_textos:
                texto_ejecutivo = consolidar_textos_ejecutivo(actividades_textos, api_key)
            else:
                texto_ejecutivo = "Sin actividades reportadas."

            row = [obl.numero, obl.descripcion, texto_ejecutivo]
            ws.append(row)
            row_idx = ws.max_row
            for col_idx in range(1, 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = cell_align
                cell.border = thin_border

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 90

        # Ajustar altura de filas de datos
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 80

        # Estilo para el título
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14, color="2c3e50")
        for cell in ws[2]:
            cell.font = Font(italic=True, size=10, color="666666")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Consolidado_{contrato.contratista or 'Contrato'}_{nombre_mes}_{anio}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        error_msg = f'Error inesperado al generar el consolidado: {str(e)}'
        if es_ajax:
            return jsonify({'error': error_msg}), 500
        flash(error_msg, 'danger')
        return redirect(url_for('index'))



# ============================================================
# CARGA MASIVA POR MES CON PROGRESO SSE Y RATE LIMITER GEMINI
# ============================================================

import threading
import time
import uuid
import json

# --- Estado global para jobs de progreso (SSE) ---
jobs_lock = threading.Lock()
jobs_progreso = {}   # job_id -> dict con estado, porcentaje, mensaje, resultado

# --- Rate limiter para Gemini (tier gratuito: 15 req/min) ---
_gemini_last_call = 0.0
_gemini_lock = threading.Lock()
GEMINI_MIN_INTERVAL = 4.1  # segundos entre llamadas (15/min = 4s, damos margen)


def _esperar_rate_limit_gemini():
    """Espera el tiempo necesario para no exceder 15 peticiones/minuto a Gemini."""
    global _gemini_last_call
    with _gemini_lock:
        ahora = time.time()
        transcurrido = ahora - _gemini_last_call
        if transcurrido < GEMINI_MIN_INTERVAL:
            esperar = GEMINI_MIN_INTERVAL - transcurrido
            time.sleep(esperar)
        _gemini_last_call = time.time()


def _actualizar_job(job_id, estado, porcentaje, mensaje, resultado=None, errores=None):
    """Actualiza el estado de un job de forma thread-safe."""
    with jobs_lock:
        if job_id not in jobs_progreso:
            jobs_progreso[job_id] = {}
        jobs_progreso[job_id].update({
            'estado': estado,
            'porcentaje': porcentaje,
            'mensaje': mensaje,
            'timestamp': time.time()
        })
        if resultado is not None:
            jobs_progreso[job_id]['resultado'] = resultado
        if errores is not None:
            jobs_progreso[job_id]['errores'] = errores


def _procesar_carga_masiva_job(job_id, contrato, obligaciones, mes, anio,
                                excel_path, imagenes_subidas, api_key):
    """Procesa la carga masiva en background y actualiza progreso."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        ws = wb.active

        # Validar encabezados
        headers = [cell.value for cell in ws[1]]
        expected = ['Obligacion No.', 'Descripcion Obligacion', 'Anuncio / Contexto', 'Fecha de la actividad', 'Nombre Imagen']
        if headers[:5] != expected:
            _actualizar_job(job_id, 'error', 0, f'Encabezados incorrectos. Esperado: {expected}')
            return

        # Contar filas válidas primero
        filas_validas = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] and not row[2]:
                continue
            filas_validas.append((idx, row))

        total_filas = len(filas_validas)
        if total_filas == 0:
            _actualizar_job(job_id, 'error', 0, 'No se encontraron filas válidas en el Excel.')
            return

        exitosos = 0
        errores = []
        reportes_cache = {}
        evidencias_por_reporte = {}

        _, last_day = calendar.monthrange(anio, mes)
        fecha_inicio_mes = date(anio, mes, 1)
        fecha_fin_mes = date(anio, mes, last_day)

        # Copiar imagenes_subidas para no mutar el original
        imagenes_disponibles = dict(imagenes_subidas)

        for i, (idx, row) in enumerate(filas_validas):
            porcentaje = int((i / total_filas) * 100)
            _actualizar_job(job_id, 'procesando', porcentaje,
                           f'Procesando fila {idx} ({i+1}/{total_filas})...')

            obl_num = row[0]
            anuncio = str(row[2] or '').strip()
            fecha_str = str(row[3] or '').strip()
            nombre_imagen = str(row[4] or '').strip()

            try:
                obl_num_int = int(obl_num)
            except (ValueError, TypeError):
                errores.append(f'Fila {idx}: Numero de obligacion invalido ({obl_num}).')
                continue

            obligacion = Obligacion.query.filter_by(numero=obl_num_int, contrato_id=contrato.id).first()
            if not obligacion:
                errores.append(f'Fila {idx}: Obligacion No. {obl_num_int} no encontrada.')
                continue

            if not anuncio:
                errores.append(f'Fila {idx}: Anuncio vacio.')
                continue

            # Parse fecha
            fecha_actividad = None
            if fecha_str:
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        fecha_actividad = datetime.strptime(fecha_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if fecha_actividad is None:
                    errores.append(f'Fila {idx}: Fecha invalida ({fecha_str}).')
                    continue
                if fecha_actividad < fecha_inicio_mes or fecha_actividad > fecha_fin_mes:
                    errores.append(f'Fila {idx}: Fecha {fecha_str} fuera del mes {mes}/{anio}.')
                    continue
            else:
                fecha_actividad = date(anio, mes, 15)

            # Buscar o crear reporte
            cache_key = (obligacion.id, mes, anio)
            if cache_key not in reportes_cache:
                reporte = ReporteMensual.query.filter_by(
                    mes=mes, anio=anio, obligacion_id=obligacion.id
                ).first()
                if not reporte:
                    reporte = ReporteMensual(
                        mes=mes, anio=anio,
                        fecha_inicio_reporte=fecha_inicio_mes,
                        fecha_fin_reporte=fecha_fin_mes,
                        obligacion_id=obligacion.id
                    )
                    db.session.add(reporte)
                    db.session.commit()
                reportes_cache[cache_key] = reporte
                ultima = Evidencia.query.filter_by(reporte_id=reporte.id).order_by(Evidencia.numero_actividad.desc()).first()
                evidencias_por_reporte[reporte.id] = ultima.numero_actividad if ultima else 0
            else:
                reporte = reportes_cache[cache_key]

            # Buscar imagen
            imagen_path = ''
            if nombre_imagen:
                _actualizar_job(job_id, 'procesando', porcentaje,
                               f'Fila {idx}: Buscando imagen "{nombre_imagen}"...')
                if nombre_imagen in imagenes_disponibles:
                    tmp_src = imagenes_disponibles.pop(nombre_imagen)
                    final_name = secure_filename(f"evidencia_{reporte.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{nombre_imagen}")
                    final_path = os.path.join(app.config['UPLOAD_FOLDER'], final_name)
                    os.rename(tmp_src, final_path)
                    imagen_path = final_path
                else:
                    safe_name = secure_filename(nombre_imagen)
                    if safe_name in imagenes_disponibles:
                        tmp_src = imagenes_disponibles.pop(safe_name)
                        final_name = secure_filename(f"evidencia_{reporte.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{nombre_imagen}")
                        final_path = os.path.join(app.config['UPLOAD_FOLDER'], final_name)
                        os.rename(tmp_src, final_path)
                        imagen_path = final_path
                    else:
                        errores.append(f'Fila {idx}: Imagen "{nombre_imagen}" no encontrada.')

            # Crear evidencia
            evidencias_por_reporte[reporte.id] += 1
            numero_actividad = evidencias_por_reporte[reporte.id]

            evidencia = Evidencia(
                numero_actividad=numero_actividad,
                imagen_path=imagen_path,
                anuncio_usuario=anuncio,
                descripcion_visual_ia=None,
                descripcion_actividad='',
                fecha_actividad=fecha_actividad,
                reporte_id=reporte.id
            )

            # Analizar con IA si hay imagen y API key (con rate limiter)
            if imagen_path and api_key:
                _actualizar_job(job_id, 'procesando', porcentaje,
                               f'Fila {idx}: Analizando con Gemini (esperando rate limit)...')
                _esperar_rate_limit_gemini()
                _actualizar_job(job_id, 'procesando', porcentaje,
                               f'Fila {idx}: Analizando imagen con IA...')
                try:
                    descripcion_visual = analizar_imagen(imagen_path, api_key)
                    if descripcion_visual:
                        evidencia.descripcion_visual_ia = descripcion_visual
                except Exception as e:
                    print(f"[CargaMasiva] Error IA fila {idx}: {e}")
                    errores.append(f'Fila {idx}: Error al analizar imagen con IA ({str(e)[:60]}).')

            evidencia.descripcion_actividad = evidencia.generar_descripcion_automatica(obligacion)
            db.session.add(evidencia)
            exitosos += 1

        db.session.commit()

        # Limpiar temporales no usados
        for tmp_path in imagenes_disponibles.values():
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass

        # Limpiar Excel temporal
        try:
            if os.path.exists(excel_path):
                os.remove(excel_path)
        except Exception:
            pass

        _actualizar_job(job_id, 'completado', 100,
                       f'Proceso finalizado. {exitosos} evidencias cargadas.',
                       resultado={'exitosos': exitosos, 'mes': mes, 'anio': anio},
                       errores=errores)

    except Exception as e:
        import traceback
        traceback.print_exc()
        _actualizar_job(job_id, 'error', 0, f'Error inesperado: {str(e)}')


def generar_plantilla_masiva(contrato, obligaciones, mes, anio):
    """Genera plantilla Excel con obligaciones para carga masiva por mes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"Carga_{mes:02d}_{anio}"

    headers = ['Obligacion No.', 'Descripcion Obligacion', 'Anuncio / Contexto', 'Fecha de la actividad', 'Nombre Imagen']
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for obl in obligaciones:
        ws.append([
            obl.numero,
            obl.descripcion,
            '',
            f"{anio}-{mes:02d}-15",
            ''
        ])

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 28
    ws.freeze_panes = 'A2'

    ws_instr = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES DE CARGA MASIVA POR MES"],
        [""],
        ["1. NO modifique los encabezados de columna (fila 1)."],
        ["2. NO modifique las columnas A y B (Obligacion No. y Descripcion)."],
        ["3. En la columna C escriba el anuncio o contexto de la actividad (solo para el sistema)."],
        ["4. En la columna D indique la fecha en formato YYYY-MM-DD o DD/MM/YYYY."],
        ["5. En la columna E escriba el nombre EXACTO del archivo de imagen, incluyendo extension (ej: evidencia1.jpg)."],
        ["6. Puede INSERTAR mas filas para la misma obligacion si tiene multiples evidencias."],
        ["7. Puede ELIMINAR las filas de obligaciones que no tengan evidencias este mes."],
        ["8. Las imagenes deben cargarse JUNTO con el Excel en el formulario web (campo de archivos multiples)."],
        [""],
        ["REGLAS IMPORTANTES:"],
        ["- La fecha debe pertenecer al mes y año seleccionados."],
        ["- El nombre de imagen en el Excel debe coincidir EXACTAMENTE con el archivo subido."],
        ["- Si no adjunta imagen, deje la columna E vacia; se creara la actividad sin evidencia visual."],
        ["- El sistema creara automaticamente los reportes mensuales por obligacion si no existen."],
        ["- Si tiene API key de Gemini configurada, analizara automaticamente cada imagen."],
        ["- NOTA: El tier gratuito de Gemini permite 15 imagenes/minuto. Si sube mas, el sistema las procesara automaticamente con pausas."],
    ]
    for row in instrucciones:
        ws_instr.append(row)
    ws_instr.column_dimensions['A'].width = 100

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Plantilla_CargaMasiva_{mes:02d}_{anio}_{contrato.contratista or 'Contrato'}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/carga-masiva-mes', methods=['GET', 'POST'])
@login_required
def carga_masiva_mes():
    """Carga masiva de evidencias por mes para TODAS las obligaciones del contrato."""
    contrato = Contrato.query.filter_by(activo=True, user_id=current_user.id).first()
    if not contrato:
        flash('No hay contrato activo configurado.', 'danger')
        return redirect(url_for('inicio'))

    if contrato.etapa == 'Reporte Cerrado':
        flash('El contrato activo esta finalizado (Reporte Cerrado). No se pueden agregar mas evidencias.', 'warning')
        return redirect(url_for('reportes'))

    obligaciones = Obligacion.query.filter_by(contrato_id=contrato.id).order_by(Obligacion.numero).all()
    meses = generar_meses_contrato(contrato.fecha_inicio, contrato.fecha_fin)

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'descargar_plantilla':
            mes = int(request.form.get('mes', 0))
            anio = int(request.form.get('anio', 0))
            if not mes or not anio:
                flash('Seleccione mes y año.', 'danger')
                return redirect(url_for('carga_masiva_mes'))
            return generar_plantilla_masiva(contrato, obligaciones, mes, anio)

        elif action == 'cargar_masivo':
            mes = int(request.form.get('mes', 0))
            anio = int(request.form.get('anio', 0))
            if not mes or not anio:
                return {'error': 'Seleccione mes y año.'}, 400

            if 'archivo_excel' not in request.files:
                return {'error': 'No se seleccionó el archivo Excel.'}, 400

            archivo_excel = request.files['archivo_excel']
            if archivo_excel.filename == '':
                return {'error': 'No se seleccionó archivo Excel.'}, 400

            if not archivo_excel.filename.endswith(('.xlsx', '.xls')):
                return {'error': 'El archivo debe ser Excel (.xlsx o .xls).'}, 400

            # Guardar Excel temporalmente
            job_id = str(uuid.uuid4())
            tmp_ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_tmp_name = secure_filename(f"tmp_excel_{job_id}_{tmp_ts}.xlsx")
            excel_tmp_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_tmp_name)
            archivo_excel.save(excel_tmp_path)

            # Guardar imagenes temporalmente
            imagenes_subidas = {}
            imagenes_files = request.files.getlist('imagenes')
            for img_file in imagenes_files:
                if img_file and img_file.filename and allowed_file(img_file.filename):
                    tmp_name = secure_filename(f"tmp_{job_id}_{img_file.filename}")
                    tmp_path = os.path.join(app.config['UPLOAD_FOLDER'], tmp_name)
                    img_file.save(tmp_path)
                    imagenes_subidas[img_file.filename] = tmp_path
                    imagenes_subidas[secure_filename(img_file.filename)] = tmp_path

            api_key = _obtener_api_key()

            # Inicializar job
            _actualizar_job(job_id, 'iniciado', 0, 'Iniciando procesamiento...')

            # Iniciar procesamiento en background
            thread = threading.Thread(
                target=_procesar_carga_masiva_job,
                args=(job_id, contrato, obligaciones, mes, anio,
                      excel_tmp_path, imagenes_subidas, api_key)
            )
            thread.daemon = True
            thread.start()

            return {'job_id': job_id, 'status': 'started'}

    api_key_configurada = bool(_obtener_api_key())
    return render_template('carga_masiva_mes.html', contrato=contrato, obligaciones=obligaciones, meses=meses, api_key_configurada=api_key_configurada)


@app.route('/carga-masiva-mes/progreso/<job_id>')
@login_required
def carga_masiva_progreso(job_id):
    """Server-Sent Events: envia progreso del job en tiempo real."""
    def event_stream():
        ultimo_estado = None
        while True:
            with jobs_lock:
                job = jobs_progreso.get(job_id, {})
            estado = job.get('estado', 'desconocido')
            porcentaje = job.get('porcentaje', 0)
            mensaje = job.get('mensaje', 'Procesando...')
            errores = job.get('errores', [])
            resultado = job.get('resultado')

            # Solo enviar si cambió algo
            if estado != ultimo_estado or True:  # enviar siempre para keepalive
                data = {
                    'estado': estado,
                    'porcentaje': porcentaje,
                    'mensaje': mensaje,
                    'errores': errores[:5]  # enviar solo primeros 5
                }
                if resultado:
                    data['resultado'] = resultado
                yield f"data: {json.dumps(data)}\n\n"
                ultimo_estado = estado

            if estado in ('completado', 'error'):
                # Limpiar job despues de un tiempo
                time.sleep(2)
                with jobs_lock:
                    jobs_progreso.pop(job_id, None)
                break

            time.sleep(0.5)

    return Response(event_stream(), mimetype='text/event-stream')

# ============================================================
# PLANTILLA EXCEL PARA CARGA MASIVA
# ============================================================

@app.route('/reporte/<int:id>/plantilla-excel')
@login_required
def descargar_plantilla_excel(id):
    """Descarga plantilla Excel para carga masiva."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Carga Masiva"

    headers = ['Anuncio / Contexto', 'Fecha de la actividad']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    ws.append(['Presentacion del estado de avance de proyectos', '2026-07-15'])
    ws.append(['Revision de solicitudes de ajuste tecnicos', '2026-07-20'])
    ws.append(['Elaboracion del plan de trabajo mensual', '2026-07-25'])

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Plantilla_Carga_Masiva_{id}.xlsx"
    )


@app.errorhandler(413)
def too_large(e):
    flash('El archivo es demasiado grande.', 'danger')
    return redirect(request.referrer or url_for('index'))


with app.app_context():
    db.create_all()
    migrar_db()


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
